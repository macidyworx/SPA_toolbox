"""
xlsx_reader.py - Reader for .xlsx files using openpyxl.
"""

from openpyxl import load_workbook

from Finders.File_sorter.cell_utils import parse_cell_ref
from Finders.File_sorter.readers.base_reader import BaseReader


class XlsxReader(BaseReader):
    """Reads .xlsx files using openpyxl."""

    def read_cell(self, filepath, sheet, cell_ref):
        row, col = parse_cell_ref(cell_ref)
        wb = load_workbook(filepath, read_only=True, data_only=True)
        try:
            ws = self._get_sheet(wb, sheet)
            value = ws.cell(row=row, column=col).value
            return self._normalize(value)
        finally:
            wb.close()

    def scan_area(self, filepath, sheet, max_rows=20, max_cols=30):
        wb = load_workbook(filepath, read_only=True, data_only=True)
        try:
            ws = self._get_sheet(wb, sheet)
            results = []
            for row in ws.iter_rows(
                min_row=1, max_row=max_rows,
                min_col=1, max_col=max_cols,
            ):
                for cell in row:
                    normalized = self._normalize(cell.value)
                    if normalized:
                        results.append(normalized)
            return results
        finally:
            wb.close()

    @staticmethod
    def _get_sheet(wb, sheet):
        if isinstance(sheet, int):
            return wb.worksheets[sheet]
        return wb[sheet]
