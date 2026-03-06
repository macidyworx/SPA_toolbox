"""
test_patonline_finder.py - Test suite for PATonline-FINDER.py

Tests header detection, file categorization, and file movement logic.
"""

import pytest
import os
import sys
import tempfile
import shutil
from pathlib import Path

# Add project root to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from openpyxl import Workbook
from Finders.PATonline_FINDER import find_headers, PATonlineFinder


# === HELPER FUNCTIONS ===
def create_test_xlsx(headers_dict, filename):
    """
    Create a test Excel file with specified headers.

    Args:
        headers_dict: Dict mapping (col, row) -> header_text
        filename: Path to save the Excel file

    Returns:
        Path to created file
    """
    wb = Workbook()
    ws = wb.active

    # Write headers to specified locations
    for (col, row), header_text in headers_dict.items():
        ws[f'{col}{row}'] = header_text

    wb.save(filename)
    return filename


# === TEST SUITE ===
class TestHeaderDetection:
    """Tests for find_headers() function."""

    def test_headers_in_first_row(self):
        """Test detecting headers in row 1."""
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test.xlsx")
            create_test_xlsx({
                ('A', 1): 'Family name',
                ('B', 1): 'Given name',
                ('C', 1): 'Unique ID',
                ('D', 1): 'Username',
            }, filepath)

            wb_openpyxl = __import__('openpyxl').load_workbook(filepath)
            ws = wb_openpyxl.active
            result = find_headers(ws)
            wb_openpyxl.close()

            assert result['has_family_name'] is True
            assert result['has_given_name'] is True
            assert result['has_unique_id'] is True
            assert result['has_username'] is True

    def test_headers_in_different_rows(self):
        """Test detecting headers scattered across different rows."""
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test.xlsx")
            create_test_xlsx({
                ('A', 1): 'Family name',
                ('B', 5): 'Given name',
                ('C', 10): 'Unique ID',
                ('D', 20): 'Username',
            }, filepath)

            wb_openpyxl = __import__('openpyxl').load_workbook(filepath)
            ws = wb_openpyxl.active
            result = find_headers(ws)
            wb_openpyxl.close()

            assert result['has_family_name'] is True
            assert result['has_given_name'] is True
            assert result['has_unique_id'] is True
            assert result['has_username'] is True

    def test_headers_in_different_columns(self):
        """Test detecting headers in columns A, B, M (outer columns)."""
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test.xlsx")
            create_test_xlsx({
                ('A', 5): 'Family name',
                ('G', 10): 'Given name',
                ('M', 15): 'Unique ID',
                ('D', 8): 'Username',
            }, filepath)

            wb_openpyxl = __import__('openpyxl').load_workbook(filepath)
            ws = wb_openpyxl.active
            result = find_headers(ws)
            wb_openpyxl.close()

            assert result['has_family_name'] is True
            assert result['has_given_name'] is True
            assert result['has_unique_id'] is True
            assert result['has_username'] is True

    def test_case_insensitive_matching(self):
        """Test that header matching is case-insensitive."""
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test.xlsx")
            create_test_xlsx({
                ('A', 1): 'family name',  # lowercase
                ('B', 1): 'GIVEN NAME',  # uppercase
                ('C', 1): 'Unique ID',  # mixed case
                ('D', 1): 'USERNAME',  # uppercase
            }, filepath)

            wb_openpyxl = __import__('openpyxl').load_workbook(filepath)
            ws = wb_openpyxl.active
            result = find_headers(ws)
            wb_openpyxl.close()

            assert result['has_family_name'] is True
            assert result['has_given_name'] is True
            assert result['has_unique_id'] is True
            assert result['has_username'] is True

    def test_whitespace_handling(self):
        """Test that leading/trailing whitespace is handled."""
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test.xlsx")
            create_test_xlsx({
                ('A', 1): '  Family name  ',
                ('B', 1): '\tGiven name\t',
                ('C', 1): ' Unique ID ',
                ('D', 1): '  Username  ',
            }, filepath)

            wb_openpyxl = __import__('openpyxl').load_workbook(filepath)
            ws = wb_openpyxl.active
            result = find_headers(ws)
            wb_openpyxl.close()

            assert result['has_family_name'] is True
            assert result['has_given_name'] is True
            assert result['has_unique_id'] is True
            assert result['has_username'] is True

    def test_partial_headers(self):
        """Test file with only some headers present."""
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test.xlsx")
            create_test_xlsx({
                ('A', 1): 'Family name',
                ('B', 1): 'Given name',
                # Missing Unique ID and Username
            }, filepath)

            wb_openpyxl = __import__('openpyxl').load_workbook(filepath)
            ws = wb_openpyxl.active
            result = find_headers(ws)
            wb_openpyxl.close()

            assert result['has_family_name'] is True
            assert result['has_given_name'] is True
            assert result['has_unique_id'] is False
            assert result['has_username'] is False

    def test_empty_file(self):
        """Test file with no headers in A1-M20."""
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test.xlsx")
            create_test_xlsx({}, filepath)

            wb_openpyxl = __import__('openpyxl').load_workbook(filepath)
            ws = wb_openpyxl.active
            result = find_headers(ws)
            wb_openpyxl.close()

            assert result['has_family_name'] is False
            assert result['has_given_name'] is False
            assert result['has_unique_id'] is False
            assert result['has_username'] is False


class TestCategorization:
    """Tests for PATonlineFinder.categorize_file() method."""

    def test_categorize_all_fields(self):
        """Test categorization of file with all four fields."""
        with tempfile.TemporaryDirectory() as tmpdir:
            finder = PATonlineFinder()
            filepath = os.path.join(tmpdir, "test.xlsx")
            create_test_xlsx({
                ('A', 1): 'Family name',
                ('B', 1): 'Given name',
                ('C', 1): 'Unique ID',
                ('D', 1): 'Username',
            }, filepath)

            result = finder.categorize_file(filepath)
            assert result == 'all_fields'

    def test_categorize_no_unique_id(self):
        """Test categorization of file with Family name, Given name, Username (no Unique ID)."""
        with tempfile.TemporaryDirectory() as tmpdir:
            finder = PATonlineFinder()
            filepath = os.path.join(tmpdir, "test.xlsx")
            create_test_xlsx({
                ('A', 1): 'Family name',
                ('B', 1): 'Given name',
                ('C', 1): 'Username',
            }, filepath)

            result = finder.categorize_file(filepath)
            assert result == 'no_unique_id'

    def test_categorize_only_unique_id(self):
        """Test categorization of file with Family name, Given name, Unique ID (no Username)."""
        with tempfile.TemporaryDirectory() as tmpdir:
            finder = PATonlineFinder()
            filepath = os.path.join(tmpdir, "test.xlsx")
            create_test_xlsx({
                ('A', 1): 'Family name',
                ('B', 1): 'Given name',
                ('C', 1): 'Unique ID',
            }, filepath)

            result = finder.categorize_file(filepath)
            assert result == 'only_unique_id'

    def test_categorize_unidentified(self):
        """Test categorization of file that doesn't match any category."""
        with tempfile.TemporaryDirectory() as tmpdir:
            finder = PATonlineFinder()
            filepath = os.path.join(tmpdir, "test.xlsx")
            create_test_xlsx({
                ('A', 1): 'Family name',
                # Missing required Given name
                ('C', 1): 'Unique ID',
            }, filepath)

            result = finder.categorize_file(filepath)
            assert result == 'unidentified'

    def test_categorize_empty_file(self):
        """Test categorization of empty file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            finder = PATonlineFinder()
            filepath = os.path.join(tmpdir, "test.xlsx")
            create_test_xlsx({}, filepath)

            result = finder.categorize_file(filepath)
            assert result == 'unidentified'


class TestFileProcessing:
    """Tests for file movement and directory creation."""

    def test_process_file_all_fields_creates_root_directory(self):
        """Test that all_fields files are moved to root output directory."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create test file
            input_dir = os.path.join(tmpdir, 'input')
            output_dir = os.path.join(tmpdir, 'output')
            os.makedirs(input_dir)
            os.makedirs(output_dir)

            finder = PATonlineFinder()
            filepath = os.path.join(input_dir, "test.xlsx")
            create_test_xlsx({
                ('A', 1): 'Family name',
                ('B', 1): 'Given name',
                ('C', 1): 'Unique ID',
                ('D', 1): 'Username',
            }, filepath)

            # Process file
            result = finder.process_file(filepath, output_dir)

            # Verify
            assert result is True
            assert not os.path.exists(filepath)  # Original file moved
            assert os.path.exists(os.path.join(output_dir, 'test.xlsx'))

    def test_process_file_no_unique_id_creates_subdirectory(self):
        """Test that no_unique_id files are moved to No_UniqueID/ subdirectory."""
        with tempfile.TemporaryDirectory() as tmpdir:
            input_dir = os.path.join(tmpdir, 'input')
            output_dir = os.path.join(tmpdir, 'output')
            os.makedirs(input_dir)
            os.makedirs(output_dir)

            finder = PATonlineFinder()
            filepath = os.path.join(input_dir, "test.xlsx")
            create_test_xlsx({
                ('A', 1): 'Family name',
                ('B', 1): 'Given name',
                ('C', 1): 'Username',
            }, filepath)

            result = finder.process_file(filepath, output_dir)

            assert result is True
            assert not os.path.exists(filepath)
            assert os.path.exists(os.path.join(output_dir, 'No_UniqueID', 'test.xlsx'))

    def test_process_file_only_unique_id_creates_subdirectory(self):
        """Test that only_unique_id files are moved to Only_UniqueID/ subdirectory."""
        with tempfile.TemporaryDirectory() as tmpdir:
            input_dir = os.path.join(tmpdir, 'input')
            output_dir = os.path.join(tmpdir, 'output')
            os.makedirs(input_dir)
            os.makedirs(output_dir)

            finder = PATonlineFinder()
            filepath = os.path.join(input_dir, "test.xlsx")
            create_test_xlsx({
                ('A', 1): 'Family name',
                ('B', 1): 'Given name',
                ('C', 1): 'Unique ID',
            }, filepath)

            result = finder.process_file(filepath, output_dir)

            assert result is True
            assert not os.path.exists(filepath)
            assert os.path.exists(os.path.join(output_dir, 'Only_UniqueID', 'test.xlsx'))

    def test_process_file_unidentified_not_moved(self):
        """Test that unidentified files are not moved."""
        with tempfile.TemporaryDirectory() as tmpdir:
            input_dir = os.path.join(tmpdir, 'input')
            output_dir = os.path.join(tmpdir, 'output')
            os.makedirs(input_dir)
            os.makedirs(output_dir)

            finder = PATonlineFinder()
            filepath = os.path.join(input_dir, "test.xlsx")
            create_test_xlsx({
                ('A', 1): 'Some Random Header',
            }, filepath)

            result = finder.process_file(filepath, output_dir)

            assert result is False
            assert os.path.exists(filepath)  # File still in original location
            assert not os.path.exists(os.path.join(output_dir, 'test.xlsx'))

    def test_process_multiple_files_mixed_categories(self):
        """Test processing multiple files with different categories."""
        with tempfile.TemporaryDirectory() as tmpdir:
            input_dir = os.path.join(tmpdir, 'input')
            output_dir = os.path.join(tmpdir, 'output')
            os.makedirs(input_dir)
            os.makedirs(output_dir)

            finder = PATonlineFinder()

            # Create three test files with different categories
            file1 = os.path.join(input_dir, "all_fields.xlsx")
            create_test_xlsx({
                ('A', 1): 'Family name',
                ('B', 1): 'Given name',
                ('C', 1): 'Unique ID',
                ('D', 1): 'Username',
            }, file1)

            file2 = os.path.join(input_dir, "no_uid.xlsx")
            create_test_xlsx({
                ('A', 1): 'Family name',
                ('B', 1): 'Given name',
                ('C', 1): 'Username',
            }, file2)

            file3 = os.path.join(input_dir, "only_uid.xlsx")
            create_test_xlsx({
                ('A', 1): 'Family name',
                ('B', 1): 'Given name',
                ('C', 1): 'Unique ID',
            }, file3)

            # Process all files
            finder.process_file(file1, output_dir)
            finder.process_file(file2, output_dir)
            finder.process_file(file3, output_dir)

            # Verify placement
            assert os.path.exists(os.path.join(output_dir, 'all_fields.xlsx'))
            assert os.path.exists(os.path.join(output_dir, 'No_UniqueID', 'no_uid.xlsx'))
            assert os.path.exists(os.path.join(output_dir, 'Only_UniqueID', 'only_uid.xlsx'))

    def test_subdirectories_created_on_demand(self):
        """Test that subdirectories are created automatically."""
        with tempfile.TemporaryDirectory() as tmpdir:
            input_dir = os.path.join(tmpdir, 'input')
            output_dir = os.path.join(tmpdir, 'output')
            os.makedirs(input_dir)
            os.makedirs(output_dir)

            finder = PATonlineFinder()
            filepath = os.path.join(input_dir, "test.xlsx")
            create_test_xlsx({
                ('A', 1): 'Family name',
                ('B', 1): 'Given name',
                ('C', 1): 'Username',
            }, filepath)

            # Verify subdirectory doesn't exist initially
            assert not os.path.exists(os.path.join(output_dir, 'No_UniqueID'))

            finder.process_file(filepath, output_dir)

            # Verify subdirectory was created
            assert os.path.exists(os.path.join(output_dir, 'No_UniqueID'))


class TestModuleImport:
    """Tests for module import and usability."""

    def test_import_class(self):
        """Test that PATonlineFinder can be imported."""
        from Finders.PATonline_FINDER import PATonlineFinder
        assert callable(PATonlineFinder)

    def test_import_function(self):
        """Test that find_headers function can be imported."""
        from Finders.PATonline_FINDER import find_headers
        assert callable(find_headers)

    def test_instantiate_finder(self):
        """Test that PATonlineFinder can be instantiated."""
        finder = PATonlineFinder()
        assert hasattr(finder, 'categorize_file')
        assert hasattr(finder, 'process_file')
        assert hasattr(finder, 'run')
