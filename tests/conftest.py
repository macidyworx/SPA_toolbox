"""
Shared pytest fixtures and configuration for SPA_toolbox tests.
"""

import pytest
from pathlib import Path
import sys
import csv
import tempfile
from io import StringIO

# Ensure project root is in path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))


@pytest.fixture
def project_root():
    """Return the project root directory."""
    return Path(__file__).parent.parent


@pytest.fixture
def tmp_xlsx(tmp_path):
    """Create a temporary Excel file for testing."""
    try:
        import openpyxl
        xlsx_path = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Header1", "Header2"])
        ws.append(["Value1", "Value2"])
        wb.save(xlsx_path)
        wb.close()
        return xlsx_path
    except ImportError:
        pytest.skip("openpyxl not available")


# ============================================================================
# IDswappers Test Fixtures (Phase 1: Test Infrastructure Setup)
# ============================================================================

@pytest.fixture
def create_test_excel():
    """
    Factory function to create test Excel files with custom headers and data.

    Usage:
        xlsx_path = create_test_excel(
            headers=["Surname", "First_Name", "Student_ID"],
            rows=[["Smith", "John", "S001"]],
            sheet_name="TestSheet"
        )
    """
    def _create(headers=None, rows=None, sheet_name="Sheet1", keep_vba=False):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not available")

        if headers is None:
            headers = ["Header1", "Header2"]
        if rows is None:
            rows = []

        tmp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        xlsx_path = tmp_file.name
        tmp_file.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        ws.append(headers)
        for row in rows:
            ws.append(row)

        wb.save(xlsx_path)
        wb.close()

        return xlsx_path

    return _create


@pytest.fixture
def create_test_xls():
    """
    Factory function to create test XLS files using xlwt.

    Usage:
        xls_path = create_test_xls(
            headers=["Surname", "First Name", "Student ID"],
            rows=[["Smith", "John", "S001"]]
        )
    """
    def _create(headers=None, rows=None, sheet_name="Sheet1"):
        try:
            import xlwt
        except ImportError:
            pytest.skip("xlwt not available")

        if headers is None:
            headers = ["Header1", "Header2"]
        if rows is None:
            rows = []

        tmp_file = tempfile.NamedTemporaryFile(suffix=".xls", delete=False)
        xls_path = tmp_file.name
        tmp_file.close()

        wb = xlwt.Workbook()
        ws = wb.add_sheet(sheet_name)

        for col_idx, header in enumerate(headers):
            ws.write(0, col_idx, header)

        for row_idx, row in enumerate(rows, start=1):
            for col_idx, value in enumerate(row):
                ws.write(row_idx, col_idx, value)

        wb.save(xls_path)
        return xls_path

    return _create


@pytest.fixture
def create_test_csv():
    """
    Factory function to create test CSV files.

    Usage:
        csv_path = create_test_csv(
            headers=["ID", "Name"],
            rows=[["1", "John Smith"]]
        )
    """
    def _create(headers=None, rows=None):
        if headers is None:
            headers = ["Header1", "Header2"]
        if rows is None:
            rows = []

        tmp_file = tempfile.NamedTemporaryFile(mode='w', suffix=".csv", delete=False, newline='')
        csv_path = tmp_file.name

        writer = csv.writer(tmp_file)
        writer.writerow(headers)
        writer.writerows(rows)
        tmp_file.close()

        return csv_path

    return _create


@pytest.fixture
def sif_lookup_file(create_test_excel):
    """Create a SIF lookup file with Firstname, Surname, StudentID columns."""
    sif_data = [
        ["John", "Smith", "S001"],
        ["Jane", "Doe", "S002"],
        ["James", "Wilson", "S003"],
    ]
    return create_test_excel(
        headers=["Firstname", "Surname", "StudentID"],
        rows=sif_data,
        sheet_name="SIF"
    )


@pytest.fixture
def ssot_lookup_file(create_test_excel):
    """Create an SSOT lookup file with Old_ID, New_ID columns."""
    ssot_data = [
        ["OLD001", "NEW001"],
        ["OLD002", "NEW002"],
        ["OLD003", "NEW003"],
    ]
    return create_test_excel(
        headers=["Old_ID", "New_ID"],
        rows=ssot_data,
        sheet_name="SSOT"
    )


@pytest.fixture
def magic_test_excel(create_test_excel):
    """Create test Excel file with MagicWords sheet."""
    data = [
        ["John", "Smith", "S001"],
        ["Jane", "Doe", "S002"],
        ["James", "Wilson", "S003"],
    ]
    return create_test_excel(
        headers=["First_Name", "Surname", "Student_ID"],
        rows=data,
        sheet_name="MagicWords"
    )


@pytest.fixture
def rol_test_excel(create_test_excel):
    """Create test Excel file with ROL Data sheet."""
    data = [
        ["Smith", "John", "S001"],
        ["Doe", "Jane", "S002"],
        ["Wilson", "James", "S003"],
    ]
    return create_test_excel(
        headers=["Surname", "First Name", "Student ID"],
        rows=data,
        sheet_name="ROL Data"
    )


@pytest.fixture
def obs_test_excel(create_test_excel):
    """Create test Excel file for OBS module (can have multiple sheets)."""
    data = [
        ["Smith", "John", "S001"],
        ["Doe", "Jane", "S002"],
    ]
    return create_test_excel(
        headers=["Surname", "First_Name", "Student ID"],
        rows=data,
        sheet_name="Observations"
    )


@pytest.fixture
def sssr_test_csv(create_test_csv):
    """Create test CSV file with SSSR format (Local Student ID, Student Name)."""
    data = [
        ["S001", "JOHN SMITH"],
        ["S002", "JANE MARIE DOE"],
        ["S003", "JAMES WILLIAM WILSON"],
    ]
    return create_test_csv(
        headers=["Local Student ID", "Student Name"],
        rows=data
    )


@pytest.fixture
def test_data_constants():
    """Shared test data constants."""
    return {
        "student_names": [
            ("John", "Smith"),
            ("Jane", "Doe"),
            ("James", "William", "Wilson"),  # middle name
        ],
        "student_ids": ["S001", "S002", "S003", "10001", "10002"],
        "sif_lookup": {
            ("john", "smith"): "S001",
            ("jane", "doe"): "S002",
            ("james", "wilson"): "S003",
        },
        "ssot_lookup": {
            "old001": "new001",
            "old002": "new002",
            "old003": "new003",
        },
        "malformed_headers": [
            "Surname  ",  # extra spaces
            "first_name",  # lowercase
            "STUDENT ID",  # all caps
            "First Name",  # different casing
        ]
    }
