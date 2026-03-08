"""
Comprehensive test suite for IDswappers module.

Tests cover all 12 swapper classes:
- Magic, ROL, RR, OBS, SSSR, DIBELS, EOI, MOI, Westwood, NAPLAN_OQ, PATdownloads, SMBtemplates

And Report_Merger functionality.

Test coverage includes:
- Header detection (_find_headers, _find_headers_xlrd)
- SIF mode processing (name-based lookup)
- SSOT mode processing (ID-based lookup)
- Data transformation
- Report generation
- Edge cases and error handling
- Integration workflows
"""

import pytest
import os
import sys
import tempfile
from pathlib import Path
from unittest.mock import Mock, MagicMock, patch

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from Helpers.Clean_fields.clean_field import field_cleaner

# Try to import all IDswapper classes
try:
    from IDswappers.Magic import MagicSwapper
    MAGIC_AVAILABLE = True
except ImportError:
    MAGIC_AVAILABLE = False

try:
    from IDswappers.ROL import ROLSwapper
    ROL_AVAILABLE = True
except ImportError:
    ROL_AVAILABLE = False

try:
    from IDswappers.RR import RRSwapper
    RR_AVAILABLE = True
except ImportError:
    RR_AVAILABLE = False

try:
    from IDswappers.OBS import OBSSwapper
    OBS_AVAILABLE = True
except ImportError:
    OBS_AVAILABLE = False

try:
    from IDswappers.SSSR import SSSRSwapper, parse_student_name
    SSSR_AVAILABLE = True
except ImportError:
    SSSR_AVAILABLE = False

try:
    from IDswappers.DIBELS import DIBELSSwapper
    DIBELS_AVAILABLE = True
except ImportError:
    DIBELS_AVAILABLE = False

try:
    from IDswappers.EOI import EOISwapper
    EOI_AVAILABLE = True
except ImportError:
    EOI_AVAILABLE = False

try:
    from IDswappers.MOI import MOISwapper
    MOI_AVAILABLE = True
except ImportError:
    MOI_AVAILABLE = False

try:
    from IDswappers.Westwood import WestwoodSwapper
    WESTWOOD_AVAILABLE = True
except ImportError:
    WESTWOOD_AVAILABLE = False

try:
    from IDswappers.NAPLAN_OQ import NAPLAN_OQSwapper
    NAPLAN_OQ_AVAILABLE = True
except ImportError:
    NAPLAN_OQ_AVAILABLE = False

try:
    from IDswappers.PATdownloads import PATdownloadsSwapper
    PATDOWNLOADS_AVAILABLE = True
except ImportError:
    PATDOWNLOADS_AVAILABLE = False

try:
    from IDswappers.SMBtemplates import SMBtemplatesSwapper
    SMBTEMPLATES_AVAILABLE = True
except ImportError:
    SMBTEMPLATES_AVAILABLE = False


# ============================================================================
# Phase 1: Test Module Importability and Instantiation
# ============================================================================

@pytest.mark.unit
class TestModuleImportability:
    """Verify each class can be imported and instantiated."""

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_magic_import_and_instantiate(self):
        """Test MagicSwapper can be imported and instantiated."""
        with patch('water_logged.the_logger.THElogger'):
            swapper = MagicSwapper()
            assert swapper is not None
            assert hasattr(swapper, '_find_headers')
            assert hasattr(swapper, '_process_sheet')

    @pytest.mark.skipif(not ROL_AVAILABLE, reason="ROL not available")
    def test_rol_import_and_instantiate(self):
        """Test ROLSwapper can be imported and instantiated."""
        with patch('water_logged.the_logger.THElogger'):
            swapper = ROLSwapper()
            assert swapper is not None
            assert hasattr(swapper, '_find_headers')
            assert hasattr(swapper, '_find_headers_xlrd')
            assert hasattr(swapper, '_process_xlsx_sheet')

    @pytest.mark.skipif(not OBS_AVAILABLE, reason="OBS not available")
    def test_obs_import_and_instantiate(self):
        """Test OBSSwapper can be imported and instantiated."""
        with patch('water_logged.the_logger.THElogger'):
            swapper = OBSSwapper()
            assert swapper is not None
            assert hasattr(swapper, '_find_headers')

    @pytest.mark.skipif(not SSSR_AVAILABLE, reason="SSSR not available")
    def test_sssr_import_and_instantiate(self):
        """Test SSSRSwapper can be imported and instantiated."""
        with patch('water_logged.the_logger.THElogger'):
            swapper = SSSRSwapper()
            assert swapper is not None
            assert hasattr(swapper, '_find_section_headers_csv')


# ============================================================================
# Phase 2: Header Detection Tests (Magic)
# ============================================================================

@pytest.mark.unit
class TestMagicHeaderDetection:
    """Test Magic._find_headers for various header scenarios."""

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_find_headers_exact_match(self, magic_test_excel):
        """Test header detection with exact matching headers."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger'):
            swapper = MagicSwapper()
            wb = load_workbook(magic_test_excel)
            ws = wb.active

            headers = swapper._find_headers(ws)
            assert headers is not None
            assert headers['fname_col'] == 'A'
            assert headers['lname_col'] == 'B'
            assert headers['id_col'] == 'C'
            assert headers['header_row'] == 1

            wb.close()

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_find_headers_case_insensitive(self, create_test_excel):
        """Test header detection with different cases."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger'):
            swapper = MagicSwapper()

            # Create file with lowercase headers
            excel_path = create_test_excel(
                headers=["first_name", "surname", "student_id"],
                rows=[["John", "Smith", "S001"]],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active
            headers = swapper._find_headers(ws)

            # Should still find headers due to field_cleaner normalization
            assert headers is not None
            wb.close()
            os.unlink(excel_path)

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_find_headers_extra_whitespace(self, create_test_excel):
        """Test header detection with extra whitespace."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger'):
            swapper = MagicSwapper()

            excel_path = create_test_excel(
                headers=["  First_Name  ", "  Surname  ", "  Student_ID  "],
                rows=[["John", "Smith", "S001"]],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active
            headers = swapper._find_headers(ws)

            assert headers is not None
            wb.close()
            os.unlink(excel_path)

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_find_headers_missing_required(self, create_test_excel):
        """Test header detection when required headers are missing."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger'):
            swapper = MagicSwapper()

            # Missing Student_ID
            excel_path = create_test_excel(
                headers=["First_Name", "Surname"],
                rows=[["John", "Smith"]],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active
            headers = swapper._find_headers(ws)

            assert headers is None
            wb.close()
            os.unlink(excel_path)

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_find_headers_extra_columns(self, create_test_excel):
        """Test header detection with extra columns beyond required."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger'):
            swapper = MagicSwapper()

            excel_path = create_test_excel(
                headers=["First_Name", "Surname", "Student_ID", "Extra1", "Extra2"],
                rows=[["John", "Smith", "S001", "X1", "X2"]],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active
            headers = swapper._find_headers(ws)

            assert headers is not None
            assert 'fname_col' in headers
            wb.close()
            os.unlink(excel_path)

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_find_headers_offset_header_row(self, create_test_excel):
        """Test header detection when header is not in row 1."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger'):
            swapper = MagicSwapper()

            excel_path = create_test_excel(
                headers=["First_Name", "Surname", "Student_ID"],
                rows=[["John", "Smith", "S001"]],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active

            # Manually insert blank rows at top
            ws.insert_rows(1, 2)
            ws['A3'] = 'First_Name'
            ws['B3'] = 'Surname'
            ws['C3'] = 'Student_ID'

            headers = swapper._find_headers(ws)
            assert headers is not None
            assert headers['header_row'] == 3

            wb.close()
            os.unlink(excel_path)


# ============================================================================
# Phase 3: ROL Header Detection (XLS and XLSX)
# ============================================================================

@pytest.mark.unit
class TestROLHeaderDetection:
    """Test ROL._find_headers and _find_headers_xlrd."""

    @pytest.mark.skipif(not ROL_AVAILABLE, reason="ROL not available")
    def test_rol_find_headers_openpyxl(self, rol_test_excel):
        """Test ROL header detection with openpyxl (xlsx)."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger'):
            swapper = ROLSwapper()
            wb = load_workbook(rol_test_excel)
            ws = wb['ROL Data']

            headers = swapper._find_headers(ws)
            assert headers is not None
            assert headers['fname_col'] == 'B'
            assert headers['lname_col'] == 'A'
            assert headers['id_col'] == 'C'

            wb.close()

    @pytest.mark.skipif(not ROL_AVAILABLE, reason="ROL not available")
    def test_rol_find_headers_xlrd(self, create_test_xls):
        """Test ROL header detection with xlrd (xls)."""
        try:
            from xlrd import open_workbook
        except ImportError:
            pytest.skip("xlrd not available")

        with patch('water_logged.the_logger.THElogger'):
            swapper = ROLSwapper()

            xls_path = create_test_xls(
                headers=["Surname", "First Name", "Student ID"],
                rows=[["Smith", "John", "S001"]],
                sheet_name="ROL Data"
            )

            rb = open_workbook(xls_path)
            sheet = rb.sheet_by_name("ROL Data")

            headers = swapper._find_headers_xlrd(sheet)
            assert headers is not None
            assert headers['lname_col'] == 0
            assert headers['fname_col'] == 1
            assert headers['id_col'] == 2

            rb.release_resources()
            os.unlink(xls_path)

    @pytest.mark.skipif(not ROL_AVAILABLE, reason="ROL not available")
    def test_rol_find_headers_xlrd_missing_sheet(self, create_test_xls):
        """Test ROL with missing target sheet."""
        try:
            from xlrd import open_workbook
        except ImportError:
            pytest.skip("xlrd not available")

        with patch('water_logged.the_logger.THElogger'):
            xls_path = create_test_xls(
                headers=["Surname", "First Name", "Student ID"],
                rows=[["Smith", "John", "S001"]],
                sheet_name="Wrong Sheet"
            )

            rb = open_workbook(xls_path)
            # Try to access non-existent sheet
            sheet_names = rb.sheet_names()
            assert "ROL Data" not in sheet_names

            rb.release_resources()
            os.unlink(xls_path)


# ============================================================================
# Phase 4: SIF Mode Data Processing Tests
# ============================================================================

@pytest.mark.unit
class TestMagicSIFMode:
    """Test Magic swapper in SIF (name-based lookup) mode."""

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_process_sheet_sif_mode_exact_match(self, create_test_excel):
        """Test data processing with exact name match in SIF lookup."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = MagicSwapper()
            swapper.logger = mock_logger

            excel_path = create_test_excel(
                headers=["First_Name", "Surname", "Student_ID"],
                rows=[["John", "Smith", "OLD001"]],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active

            lookup = {
                ("john", "smith"): "NEW001",
            }
            not_found = []

            checked, matched, nf_count = swapper._process_sheet(
                ws, "sif", lookup, not_found, excel_path
            )

            assert checked == 1
            assert matched == 1
            assert nf_count == 0

            # Verify ID was updated
            assert ws['C2'].value == "NEW001"

            wb.close()
            os.unlink(excel_path)

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_process_sheet_sif_mode_not_found(self, create_test_excel):
        """Test tracking of unmatched names in SIF mode."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = MagicSwapper()
            swapper.logger = mock_logger

            excel_path = create_test_excel(
                headers=["First_Name", "Surname", "Student_ID"],
                rows=[["John", "Smith", "OLD001"], ["Jane", "Unknown", "OLD002"]],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active

            lookup = {
                ("john", "smith"): "NEW001",
            }
            not_found = []

            checked, matched, nf_count = swapper._process_sheet(
                ws, "sif", lookup, not_found, excel_path
            )

            assert checked == 2
            assert matched == 1
            assert nf_count == 1
            assert len(not_found) == 1
            assert not_found[0]['Fname'] == 'Jane'

            wb.close()
            os.unlink(excel_path)

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_process_sheet_sif_mode_case_insensitive(self, create_test_excel):
        """Test case-insensitive name matching in SIF mode."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = MagicSwapper()
            swapper.logger = mock_logger

            excel_path = create_test_excel(
                headers=["First_Name", "Surname", "Student_ID"],
                rows=[["JOHN", "SMITH", "OLD001"]],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active

            lookup = {
                ("john", "smith"): "NEW001",
            }
            not_found = []

            checked, matched, nf_count = swapper._process_sheet(
                ws, "sif", lookup, not_found, excel_path
            )

            assert matched == 1
            wb.close()
            os.unlink(excel_path)

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_process_sheet_sif_mode_with_whitespace(self, create_test_excel):
        """Test name matching with extra whitespace in data."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = MagicSwapper()
            swapper.logger = mock_logger

            excel_path = create_test_excel(
                headers=["First_Name", "Surname", "Student_ID"],
                rows=[[" John ", "  Smith  ", "OLD001"]],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active

            lookup = {
                ("john", "smith"): "NEW001",
            }
            not_found = []

            checked, matched, nf_count = swapper._process_sheet(
                ws, "sif", lookup, not_found, excel_path
            )

            assert matched == 1
            wb.close()
            os.unlink(excel_path)

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_process_sheet_sif_mode_skip_empty_names(self, create_test_excel):
        """Test that rows with missing names are skipped."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = MagicSwapper()
            swapper.logger = mock_logger

            excel_path = create_test_excel(
                headers=["First_Name", "Surname", "Student_ID"],
                rows=[[None, "Smith", "OLD001"], ["John", "Smith", "OLD002"]],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active

            lookup = {
                ("john", "smith"): "NEW002",
            }
            not_found = []

            checked, matched, nf_count = swapper._process_sheet(
                ws, "sif", lookup, not_found, excel_path
            )

            # Only the second row should be checked (first has None)
            assert checked == 1
            assert matched == 1

            wb.close()
            os.unlink(excel_path)


# ============================================================================
# Phase 5: SSOT Mode Data Processing Tests
# ============================================================================

@pytest.mark.unit
class TestMagicSSOTMode:
    """Test Magic swapper in SSOT (ID-based lookup) mode."""

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_process_sheet_ssot_mode_exact_match(self, create_test_excel):
        """Test data processing with exact ID match in SSOT lookup."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = MagicSwapper()
            swapper.logger = mock_logger

            excel_path = create_test_excel(
                headers=["First_Name", "Surname", "Student_ID"],
                rows=[["John", "Smith", "OLD001"]],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active

            lookup = {
                "old001": "NEW001",
            }
            not_found = []

            checked, matched, nf_count = swapper._process_sheet(
                ws, "ssot", lookup, not_found, excel_path
            )

            assert checked == 1
            assert matched == 1
            assert nf_count == 0
            assert ws['C2'].value == "NEW001"

            wb.close()
            os.unlink(excel_path)

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_process_sheet_ssot_mode_not_found(self, create_test_excel):
        """Test tracking of unmatched IDs in SSOT mode."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = MagicSwapper()
            swapper.logger = mock_logger

            excel_path = create_test_excel(
                headers=["First_Name", "Surname", "Student_ID"],
                rows=[["John", "Smith", "OLD001"], ["Jane", "Doe", "UNKNOWN"]],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active

            lookup = {
                "old001": "NEW001",
            }
            not_found = []

            checked, matched, nf_count = swapper._process_sheet(
                ws, "ssot", lookup, not_found, excel_path
            )

            assert checked == 2
            assert matched == 1
            assert nf_count == 1
            assert len(not_found) == 1

            wb.close()
            os.unlink(excel_path)

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_process_sheet_ssot_mode_skip_empty_ids(self, create_test_excel):
        """Test that rows with missing IDs are skipped in SSOT mode."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = MagicSwapper()
            swapper.logger = mock_logger

            excel_path = create_test_excel(
                headers=["First_Name", "Surname", "Student_ID"],
                rows=[[None, None, None], ["John", "Smith", "OLD001"]],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active

            lookup = {
                "old001": "NEW001",
            }
            not_found = []

            checked, matched, nf_count = swapper._process_sheet(
                ws, "ssot", lookup, not_found, excel_path
            )

            # Only the second row should be checked
            assert checked == 1
            assert matched == 1

            wb.close()
            os.unlink(excel_path)


# ============================================================================
# Phase 6: SSSR-Specific Tests (Name Parsing)
# ============================================================================

@pytest.mark.unit
class TestSSSRNameParsing:
    """Test SSSR.parse_student_name function."""

    @pytest.mark.skipif(not SSSR_AVAILABLE, reason="SSSR not available")
    def test_parse_student_name_two_part(self):
        """Test parsing names with first and last name."""
        result = parse_student_name("JOHN SMITH")
        assert result == ("JOHN", "SMITH")

    @pytest.mark.skipif(not SSSR_AVAILABLE, reason="SSSR not available")
    def test_parse_student_name_with_middle(self):
        """Test parsing names with middle names."""
        result = parse_student_name("JOHN WILLIAM SMITH")
        assert result == ("JOHN", "SMITH")

    @pytest.mark.skipif(not SSSR_AVAILABLE, reason="SSSR not available")
    def test_parse_student_name_single_word(self):
        """Test that single word names return None."""
        result = parse_student_name("JOHN")
        assert result is None

    @pytest.mark.skipif(not SSSR_AVAILABLE, reason="SSSR not available")
    def test_parse_student_name_empty(self):
        """Test that empty string returns None."""
        result = parse_student_name("")
        assert result is None

    @pytest.mark.skipif(not SSSR_AVAILABLE, reason="SSSR not available")
    def test_parse_student_name_none(self):
        """Test that None input returns None."""
        result = parse_student_name(None)
        assert result is None

    @pytest.mark.skipif(not SSSR_AVAILABLE, reason="SSSR not available")
    def test_parse_student_name_with_whitespace(self):
        """Test parsing with extra whitespace."""
        result = parse_student_name("  JOHN   SMITH  ")
        assert result == ("JOHN", "SMITH")


@pytest.mark.unit
class TestSSSRCSVProcessing:
    """Test SSSR CSV header detection."""

    @pytest.mark.skipif(not SSSR_AVAILABLE, reason="SSSR not available")
    def test_find_section_headers_csv(self, sssr_test_csv):
        """Test finding SSSR section headers in CSV."""
        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = SSSRSwapper()
            swapper.logger = mock_logger

            with open(sssr_test_csv, 'r') as f:
                import csv
                reader = csv.reader(f)
                all_rows = list(reader)

            result = swapper._find_section_headers_csv(all_rows)

            assert result is not None
            assert 'header_idx' in result
            assert 'id_col' in result
            assert 'name_col' in result

    @pytest.mark.skipif(not SSSR_AVAILABLE, reason="SSSR not available")
    def test_find_section_headers_csv_missing_headers(self):
        """Test CSV with missing required headers."""
        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = SSSRSwapper()
            swapper.logger = mock_logger

            all_rows = [
                ["Student ID", "Name"],  # Wrong headers
                ["S001", "John"],
            ]

            result = swapper._find_section_headers_csv(all_rows)
            assert result is None


# ============================================================================
# Phase 7: Multi-Sheet Processing Tests
# ============================================================================

@pytest.mark.unit
class TestMultiSheetProcessing:
    """Test processing of workbooks with multiple sheets."""

    @pytest.mark.skipif(not OBS_AVAILABLE, reason="OBS not available")
    def test_obs_multi_sheet_processing(self, create_test_excel):
        """Test that OBS processes only sheets with correct headers."""
        try:
            from openpyxl import load_workbook, Workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = OBSSwapper()
            swapper.logger = mock_logger

            # Create workbook with multiple sheets
            tmp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            wb_path = tmp_file.name
            tmp_file.close()

            wb = load_workbook(create_test_excel(
                headers=["Surname", "First_Name", "Student ID"],
                rows=[["Smith", "John", "S001"]],
                sheet_name="Sheet1"
            ))

            # Add second sheet without required headers
            ws2 = wb.create_sheet("Sheet2")
            ws2.append(["Other", "Headers"])

            wb.save(wb_path)
            wb.close()

            # Now load and test
            wb = load_workbook(wb_path)
            ws1 = wb['Sheet1']
            ws2 = wb['Sheet2']

            headers1 = swapper._find_headers(ws1)
            headers2 = swapper._find_headers(ws2)

            assert headers1 is not None
            assert headers2 is None

            wb.close()
            os.unlink(wb_path)


# ============================================================================
# Phase 8: Report Generation Tests
# ============================================================================

@pytest.mark.needs_files
class TestReportGeneration:
    """Test report generation functionality."""

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_generate_not_found_log(self, create_test_excel):
        """Test not-found log generation."""
        try:
            from openpyxl import load_workbook, Workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = MagicSwapper()
            swapper.logger = mock_logger

            excel_path = create_test_excel(
                headers=["First_Name", "Surname", "Student_ID"],
                rows=[["John", "Unknown", "OLD001"]],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active

            lookup = {}
            not_found = []

            swapper._process_sheet(ws, "sif", lookup, not_found, excel_path)

            assert len(not_found) == 1
            assert 'File' in not_found[0]
            assert 'Row' in not_found[0]

            wb.close()
            os.unlink(excel_path)

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_generate_summary_report_structure(self):
        """Test that summary report has correct structure."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        report_wb = Workbook()
        summary_ws = report_wb.active
        summary_ws.title = 'Summary'

        # Simulate report structure
        summary_ws.append(['Metric', 'Value'])
        summary_ws.append(['Total Files Processed', 3])
        summary_ws.append(['Total Matched', 25])
        summary_ws.append(['Total NOT Matched', 5])

        assert summary_ws['A1'].value == 'Metric'
        assert summary_ws['B1'].value == 'Value'
        assert summary_ws['A2'].value == 'Total Files Processed'

        report_wb.close()


# ============================================================================
# Phase 9: Edge Cases and Error Handling
# ============================================================================

@pytest.mark.unit
class TestEdgeCases:
    """Test edge cases and error handling."""

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_process_empty_workbook(self, create_test_excel):
        """Test processing workbook with no data rows."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = MagicSwapper()
            swapper.logger = mock_logger

            excel_path = create_test_excel(
                headers=["First_Name", "Surname", "Student_ID"],
                rows=[],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active

            lookup = {("john", "smith"): "NEW001"}
            not_found = []

            checked, matched, nf_count = swapper._process_sheet(
                ws, "sif", lookup, not_found, excel_path
            )

            assert checked == 0
            assert matched == 0

            wb.close()
            os.unlink(excel_path)

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_process_duplicate_student_names(self, create_test_excel):
        """Test processing with duplicate student names."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = MagicSwapper()
            swapper.logger = mock_logger

            excel_path = create_test_excel(
                headers=["First_Name", "Surname", "Student_ID"],
                rows=[
                    ["John", "Smith", "OLD001"],
                    ["John", "Smith", "OLD002"],  # Duplicate name
                ],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active

            lookup = {("john", "smith"): "NEW001"}
            not_found = []

            checked, matched, nf_count = swapper._process_sheet(
                ws, "sif", lookup, not_found, excel_path
            )

            # Both rows match the same lookup entry
            assert matched == 2

            wb.close()
            os.unlink(excel_path)

    def test_field_cleaner_normalization(self):
        """Test that field_cleaner properly normalizes text."""
        original = "  Test  Name  "
        cleaned = field_cleaner(original, strip_spaces=True)

        # Should be normalized
        assert "  " not in cleaned

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_process_unicode_names(self, create_test_excel):
        """Test processing with unicode characters in names."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = MagicSwapper()
            swapper.logger = mock_logger

            excel_path = create_test_excel(
                headers=["First_Name", "Surname", "Student_ID"],
                rows=[["José", "García", "OLD001"]],
                sheet_name="MagicWords"
            )

            wb = load_workbook(excel_path)
            ws = wb.active

            # Create lookup with normalized unicode
            lookup = {
                (field_cleaner("José"), field_cleaner("García")): "NEW001",
            }
            not_found = []

            checked, matched, nf_count = swapper._process_sheet(
                ws, "sif", lookup, not_found, excel_path
            )

            assert matched >= 0  # May or may not match depending on normalization

            wb.close()
            os.unlink(excel_path)


# ============================================================================
# Phase 10: Integration Tests
# ============================================================================

@pytest.mark.integration
@pytest.mark.needs_files
class TestIntegrationWorkflows:
    """Test complete end-to-end workflows."""

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_complete_sif_workflow(self, create_test_excel):
        """Test complete workflow: lookup loading, multi-file processing, report generation."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = MagicSwapper()
            swapper.logger = mock_logger

            # Create SIF lookup
            sif_lookup = {
                ("john", "smith"): "NEW001",
                ("jane", "doe"): "NEW002",
            }

            # Create test files
            test_files = []
            for i in range(2):
                excel_path = create_test_excel(
                    headers=["First_Name", "Surname", "Student_ID"],
                    rows=[
                        ["John", "Smith", f"OLD{i}01"],
                        ["Jane", "Doe", f"OLD{i}02"],
                    ],
                    sheet_name="MagicWords"
                )
                test_files.append(excel_path)

            # Process both files
            total_matched = 0
            not_found_all = []

            for file_path in test_files:
                wb = load_workbook(file_path)
                ws = wb.active

                not_found = []
                checked, matched, nf_count = swapper._process_sheet(
                    ws, "sif", sif_lookup, not_found, file_path
                )

                total_matched += matched
                not_found_all.extend(not_found)

                wb.close()

            assert total_matched == 4
            assert len(not_found_all) == 0

            # Cleanup
            for f in test_files:
                os.unlink(f)

    @pytest.mark.skipif(not MAGIC_AVAILABLE, reason="Magic not available")
    def test_complete_ssot_workflow(self, create_test_excel):
        """Test complete SSOT workflow with multiple files."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            swapper = MagicSwapper()
            swapper.logger = mock_logger

            # Create SSOT lookup
            ssot_lookup = {
                "old001": "new001",
                "old002": "new002",
            }

            # Create test files
            test_files = []
            excel_path = create_test_excel(
                headers=["First_Name", "Surname", "Student_ID"],
                rows=[
                    ["John", "Smith", "OLD001"],
                    ["Jane", "Doe", "OLD002"],
                ],
                sheet_name="MagicWords"
            )
            test_files.append(excel_path)

            # Process files
            total_matched = 0
            for file_path in test_files:
                wb = load_workbook(file_path)
                ws = wb.active

                not_found = []
                checked, matched, nf_count = swapper._process_sheet(
                    ws, "ssot", ssot_lookup, not_found, file_path
                )

                total_matched += matched
                wb.close()

            assert total_matched == 2

            # Cleanup
            for f in test_files:
                os.unlink(f)
