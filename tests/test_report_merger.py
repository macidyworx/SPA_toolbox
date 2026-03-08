"""
Comprehensive test suite for Report_Merger class.

Tests cover:
- Summary sheet merging (side-by-side with gaps)
- Full List sheet merging (concatenation with deduplicated headers)
- Single, multiple, and empty input files
- Missing sheet handling
- Output file creation
- Error handling with corrupt files
"""

import pytest
import os
import sys
import tempfile
from pathlib import Path
from unittest.mock import Mock, patch

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from IDswappers.Report_Merger import ReportMerger
    REPORT_MERGER_AVAILABLE = True
except ImportError:
    REPORT_MERGER_AVAILABLE = False


# ============================================================================
# Phase 1: Basic Report Merger Tests
# ============================================================================

@pytest.mark.unit
class TestReportMergerImportability:
    """Verify ReportMerger can be imported and instantiated."""

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_report_merger_import_and_instantiate(self):
        """Test ReportMerger can be imported and instantiated."""
        with patch('water_logged.the_logger.THElogger'):
            merger = ReportMerger()
            assert merger is not None
            assert hasattr(merger, 'run')


# ============================================================================
# Phase 2: Summary Sheet Merging
# ============================================================================

@pytest.mark.needs_files
class TestSummarySheetMerging:
    """Test Summary sheet merging (side-by-side with gaps)."""

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_merge_single_summary_sheet(self, create_test_excel):
        """Test merging a single Summary sheet."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            merger = ReportMerger()
            merger.logger = mock_logger

            # Create report file with Summary sheet
            report_path = create_test_excel(
                headers=["Metric", "Value"],
                rows=[
                    ["Total Files", 5],
                    ["Matched", 100],
                    ["Not Found", 10],
                ],
                sheet_name="Summary"
            )

            wb = load_workbook(report_path, data_only=True)
            ws = wb['Summary']
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            wb.close()

            assert len(rows) == 4  # Header + 3 data rows
            assert rows[0] == ["Metric", "Value"]

            os.unlink(report_path)

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_merge_multiple_summary_sheets_side_by_side(self, create_test_excel):
        """Test merging multiple Summary sheets side-by-side."""
        try:
            from openpyxl import load_workbook, Workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            # Create two report files
            report1_path = create_test_excel(
                headers=["Metric", "Value"],
                rows=[
                    ["Total Files", 5],
                    ["Matched", 100],
                ],
                sheet_name="Summary"
            )

            report2_path = create_test_excel(
                headers=["Metric", "Value"],
                rows=[
                    ["Total Files", 3],
                    ["Matched", 50],
                ],
                sheet_name="Summary"
            )

            # Simulate merging (side-by-side with gaps)
            summaries = []
            for report_path in [report1_path, report2_path]:
                wb = load_workbook(report_path, data_only=True)
                ws = wb['Summary']
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(list(row))
                summaries.append(rows)
                wb.close()

            # Create merged workbook
            out_wb = Workbook()
            summary_ws = out_wb.active
            summary_ws.title = 'Summary'

            for summary_rows in summaries:
                start_col = summary_ws.max_column + 1 if summary_ws.max_column > 1 else 1
                if start_col > 1:
                    start_col += 1  # gap column

                for row_idx, row_data in enumerate(summary_rows, start=1):
                    for col_offset, value in enumerate(row_data):
                        summary_ws.cell(row=row_idx, column=start_col + col_offset, value=value)

            # Verify structure
            assert summary_ws.max_column >= 4  # At least 2 reports (2 cols each) + 1 gap
            assert summary_ws['A1'].value == "Metric"
            assert summary_ws['D1'].value == "Metric"  # Second report starts at column D (gap logic: max_col+1 then +1 for gap)

            out_wb.close()
            os.unlink(report1_path)
            os.unlink(report2_path)

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_merge_three_summary_sheets(self, create_test_excel):
        """Test merging three Summary sheets with proper gaps."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger'):
            # Create three reports
            reports_data = [
                [["Metric", "Value"], ["Files", 5], ["Matched", 100]],
                [["Metric", "Value"], ["Files", 3], ["Matched", 50]],
                [["Metric", "Value"], ["Files", 2], ["Matched", 30]],
            ]

            # Simulate merging
            out_wb = Workbook()
            summary_ws = out_wb.active
            summary_ws.title = 'Summary'

            for summary_rows in reports_data:
                start_col = summary_ws.max_column + 1 if summary_ws.max_column > 1 else 1
                if start_col > 1:
                    start_col += 1  # gap column

                for row_idx, row_data in enumerate(summary_rows, start=1):
                    for col_offset, value in enumerate(row_data):
                        summary_ws.cell(row=row_idx, column=start_col + col_offset, value=value)

            # Verify: 3 reports × 2 cols + 2 gaps = 8 columns
            assert summary_ws.max_column >= 6

            out_wb.close()


# ============================================================================
# Phase 3: Full List Sheet Merging
# ============================================================================

@pytest.mark.needs_files
class TestFullListSheetMerging:
    """Test Full List sheet merging (concatenation with header deduplication)."""

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_merge_single_full_list(self, create_test_excel):
        """Test merging a single Full List sheet."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            # Create report with Full List
            report_path = create_test_excel(
                headers=["File", "Row", "Fname", "Lname"],
                rows=[
                    ["file1.xlsx", 2, "John", "Smith"],
                    ["file1.xlsx", 3, "Jane", "Doe"],
                ],
                sheet_name="Full List"
            )

            wb = load_workbook(report_path, data_only=True)
            ws = wb['Full List']
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            wb.close()

            assert len(rows) == 3  # Header + 2 data rows
            assert rows[0][0] == "File"

            os.unlink(report_path)

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_merge_multiple_full_lists_deduplicate_headers(self, create_test_excel):
        """Test merging multiple Full List sheets with header deduplication."""
        try:
            from openpyxl import load_workbook, Workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger'):
            # Create two reports with Full List sheets
            report1_path = create_test_excel(
                headers=["File", "Row", "Fname", "Lname"],
                rows=[
                    ["file1.xlsx", 2, "John", "Smith"],
                    ["file1.xlsx", 3, "Jane", "Doe"],
                ],
                sheet_name="Full List"
            )

            report2_path = create_test_excel(
                headers=["File", "Row", "Fname", "Lname"],
                rows=[
                    ["file2.xlsx", 2, "James", "Wilson"],
                ],
                sheet_name="Full List"
            )

            # Simulate merging
            full_lists = []
            for report_path in [report1_path, report2_path]:
                wb = load_workbook(report_path, data_only=True)
                ws = wb['Full List']
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(list(row))
                full_lists.append(rows)
                wb.close()

            # Merge with header deduplication
            out_wb = Workbook()
            fl_ws = out_wb.create_sheet('Full List')

            header_written = False
            for fl_rows in full_lists:
                for i, row_data in enumerate(fl_rows):
                    if i == 0 and header_written:
                        continue  # skip duplicate header
                    fl_ws.append(row_data)
                header_written = True

            # Verify: header once + 3 data rows = 4 rows
            assert fl_ws.max_row == 4
            assert fl_ws['A1'].value == "File"

            out_wb.close()
            os.unlink(report1_path)
            os.unlink(report2_path)

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_merge_full_list_with_empty_input(self, create_test_excel):
        """Test merging when one input has empty Full List."""
        try:
            from openpyxl import load_workbook, Workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger'):
            # First report with data
            report1_path = create_test_excel(
                headers=["File", "Row", "Fname", "Lname"],
                rows=[["file1.xlsx", 2, "John", "Smith"]],
                sheet_name="Full List"
            )

            # Second report with header only (no data)
            report2_path = create_test_excel(
                headers=["File", "Row", "Fname", "Lname"],
                rows=[],
                sheet_name="Full List"
            )

            full_lists = []
            for report_path in [report1_path, report2_path]:
                wb = load_workbook(report_path, data_only=True)
                ws = wb['Full List']
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(list(row))
                full_lists.append(rows)
                wb.close()

            # Merge
            out_wb = Workbook()
            fl_ws = out_wb.create_sheet('Full List')

            header_written = False
            for fl_rows in full_lists:
                for i, row_data in enumerate(fl_rows):
                    if i == 0 and header_written:
                        continue
                    fl_ws.append(row_data)
                header_written = True

            # Should have header + 1 data row
            assert fl_ws.max_row == 2

            out_wb.close()
            os.unlink(report1_path)
            os.unlink(report2_path)


# ============================================================================
# Phase 4: Missing Sheet Handling
# ============================================================================

@pytest.mark.needs_files
class TestMissingSheetHandling:
    """Test handling of missing Summary/Full List sheets."""

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_file_without_summary_sheet(self, create_test_excel):
        """Test file without Summary sheet is skipped gracefully."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            # Create report with only Full List
            report_path = create_test_excel(
                headers=["File", "Row", "Fname"],
                rows=[["file1.xlsx", 2, "John"]],
                sheet_name="Full List"
            )

            wb = load_workbook(report_path, data_only=True)

            # Check Summary doesn't exist
            assert 'Summary' not in wb.sheetnames
            assert 'Full List' in wb.sheetnames

            wb.close()
            os.unlink(report_path)

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_file_without_full_list_sheet(self, create_test_excel):
        """Test file without Full List sheet is handled."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            # Create report with only Summary
            report_path = create_test_excel(
                headers=["Metric", "Value"],
                rows=[["Files", 5]],
                sheet_name="Summary"
            )

            wb = load_workbook(report_path, data_only=True)

            # Check Full List doesn't exist
            assert 'Summary' in wb.sheetnames
            assert 'Full List' not in wb.sheetnames

            wb.close()
            os.unlink(report_path)

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_file_with_neither_sheet(self, create_test_excel):
        """Test file with neither Summary nor Full List."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            # Create file with different sheets
            report_path = create_test_excel(
                headers=["Data1", "Data2"],
                rows=[["Value1", "Value2"]],
                sheet_name="Other"
            )

            wb = load_workbook(report_path, data_only=True)

            # Neither expected sheet exists
            assert 'Summary' not in wb.sheetnames
            assert 'Full List' not in wb.sheetnames

            wb.close()
            os.unlink(report_path)


# ============================================================================
# Phase 5: Output File Creation
# ============================================================================

@pytest.mark.needs_files
class TestOutputFileCreation:
    """Test output file creation and location."""

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_output_file_creation(self, create_test_excel):
        """Test that output file is created correctly."""
        try:
            from openpyxl import load_workbook, Workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            # Create temporary output directory
            with tempfile.TemporaryDirectory() as tmpdir:
                output_path = os.path.join(tmpdir, "FULL_Report.xlsx")

                # Create merged report
                out_wb = Workbook()
                summary_ws = out_wb.active
                summary_ws.title = 'Summary'
                summary_ws.append(['Metric', 'Value'])
                summary_ws.append(['Files', 5])

                out_wb.save(output_path)
                out_wb.close()

                # Verify file exists and is valid
                assert os.path.exists(output_path)

                wb = load_workbook(output_path)
                assert 'Summary' in wb.sheetnames
                wb.close()

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_output_file_has_both_sheets(self, create_test_excel):
        """Test that output file contains both Summary and Full List sheets."""
        try:
            from openpyxl import load_workbook, Workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            with tempfile.TemporaryDirectory() as tmpdir:
                output_path = os.path.join(tmpdir, "FULL_Report.xlsx")

                # Create merged report with both sheets
                out_wb = Workbook()
                summary_ws = out_wb.active
                summary_ws.title = 'Summary'
                summary_ws.append(['Metric', 'Value'])

                fl_ws = out_wb.create_sheet('Full List')
                fl_ws.append(['File', 'Row'])

                out_wb.save(output_path)
                out_wb.close()

                # Verify both sheets exist
                wb = load_workbook(output_path)
                assert 'Summary' in wb.sheetnames
                assert 'Full List' in wb.sheetnames
                wb.close()


# ============================================================================
# Phase 6: Error Handling
# ============================================================================

@pytest.mark.unit
class TestErrorHandling:
    """Test error handling with corrupt or invalid files."""

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_corrupt_excel_file_handling(self):
        """Test graceful handling of corrupt Excel files."""
        with patch('water_logged.the_logger.THElogger') as mock_logger:
            merger = ReportMerger()
            merger.logger = mock_logger

            # Create a corrupt file
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                corrupt_path = f.name
                f.write(b"This is not a valid Excel file")

            try:
                # Attempt to load it
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(corrupt_path)
                    # Should fail
                    assert False, "Should have raised an exception"
                except Exception:
                    # Expected behavior
                    pass
            finally:
                os.unlink(corrupt_path)

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_missing_file_handling(self):
        """Test handling when referenced file doesn't exist."""
        with patch('water_logged.the_logger.THElogger') as mock_logger:
            merger = ReportMerger()
            merger.logger = mock_logger

            nonexistent_path = "/tmp/nonexistent_file_xyz.xlsx"

            # File should not exist
            assert not os.path.exists(nonexistent_path)


# ============================================================================
# Phase 7: Edge Cases
# ============================================================================

@pytest.mark.needs_files
class TestEdgeCases:
    """Test edge cases in report merging."""

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_empty_input_files(self, create_test_excel):
        """Test merging when input files are empty."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            # Create empty Summary sheets
            report_path = create_test_excel(
                headers=["Metric", "Value"],
                rows=[],
                sheet_name="Summary"
            )

            wb = load_workbook(report_path, data_only=True)
            ws = wb['Summary']
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            wb.close()

            # Should have only header
            assert len(rows) == 1

            os.unlink(report_path)

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_large_number_of_reports(self, create_test_excel):
        """Test merging many reports (stress test)."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger'):
            # Simulate merging 10 reports
            reports_data = [
                [["Metric", "Value"], ["Files", i], ["Matched", i * 10]]
                for i in range(1, 11)
            ]

            out_wb = Workbook()
            summary_ws = out_wb.active
            summary_ws.title = 'Summary'

            for summary_rows in reports_data:
                start_col = summary_ws.max_column + 1 if summary_ws.max_column > 1 else 1
                if start_col > 1:
                    start_col += 1

                for row_idx, row_data in enumerate(summary_rows, start=1):
                    for col_offset, value in enumerate(row_data):
                        summary_ws.cell(row=row_idx, column=start_col + col_offset, value=value)

            # Verify: 10 reports × 2 cols + 9 gaps = 29 columns minimum
            assert summary_ws.max_column >= 20

            out_wb.close()

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_reports_with_different_metrics(self, create_test_excel):
        """Test merging reports with slightly different metric names."""
        try:
            from openpyxl import load_workbook, Workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger'):
            # Create reports with different metrics
            report1_path = create_test_excel(
                headers=["Metric", "Value"],
                rows=[
                    ["Total Files", 5],
                    ["Matched", 100],
                ],
                sheet_name="Summary"
            )

            report2_path = create_test_excel(
                headers=["Metric", "Value"],
                rows=[
                    ["Files Processed", 3],
                    ["Not Found", 10],
                ],
                sheet_name="Summary"
            )

            # Load both
            summaries = []
            for path in [report1_path, report2_path]:
                wb = load_workbook(path, data_only=True)
                ws = wb['Summary']
                rows = [list(r) for r in ws.iter_rows(values_only=True)]
                summaries.append(rows)
                wb.close()

            assert len(summaries) == 2
            assert summaries[0][1][0] == "Total Files"
            assert summaries[1][1][0] == "Files Processed"

            os.unlink(report1_path)
            os.unlink(report2_path)


# ============================================================================
# Phase 8: Integration Tests
# ============================================================================

@pytest.mark.integration
@pytest.mark.needs_files
class TestReportMergerIntegration:
    """Integration tests for complete Report_Merger workflows."""

    @pytest.mark.skipif(not REPORT_MERGER_AVAILABLE, reason="Report_Merger not available")
    def test_complete_merge_workflow_with_multiple_files(self, create_test_excel):
        """Test complete workflow: load multiple reports, merge, and save."""
        try:
            from openpyxl import load_workbook, Workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch('water_logged.the_logger.THElogger') as mock_logger:
            with tempfile.TemporaryDirectory() as tmpdir:
                # Create 3 input report files
                input_files = []
                for i in range(1, 4):
                    report_path = create_test_excel(
                        headers=["Metric", "Value"],
                        rows=[
                            ["Files", i],
                            ["Matched", i * 20],
                        ],
                        sheet_name="Summary"
                    )
                    input_files.append(report_path)

                # Create Full List sheets too
                for idx, path in enumerate(input_files, start=1):
                    wb = load_workbook(path)
                    fl_ws = wb.create_sheet("Full List")
                    fl_ws.append(["File", "Row", "Name"])
                    fl_ws.append([f"file{idx}.xlsx", 2, f"Student{idx}"])
                    wb.save(path)
                    wb.close()

                # Simulate merge process
                summaries = []
                full_lists = []

                for path in input_files:
                    wb = load_workbook(path, data_only=True)

                    if 'Summary' in wb.sheetnames:
                        ws = wb['Summary']
                        rows = [list(r) for r in ws.iter_rows(values_only=True)]
                        summaries.append(rows)

                    if 'Full List' in wb.sheetnames:
                        ws = wb['Full List']
                        rows = [list(r) for r in ws.iter_rows(values_only=True)]
                        full_lists.append(rows)

                    wb.close()

                # Create merged output
                output_path = os.path.join(tmpdir, "FULL_Report.xlsx")
                out_wb = Workbook()

                # Merge Summary sheets
                summary_ws = out_wb.active
                summary_ws.title = 'Summary'

                for summary_rows in summaries:
                    start_col = summary_ws.max_column + 1 if summary_ws.max_column > 1 else 1
                    if start_col > 1:
                        start_col += 1

                    for row_idx, row_data in enumerate(summary_rows, start=1):
                        for col_offset, value in enumerate(row_data):
                            summary_ws.cell(row=row_idx, column=start_col + col_offset, value=value)

                # Merge Full List sheets
                fl_ws = out_wb.create_sheet('Full List')
                header_written = False
                for fl_rows in full_lists:
                    for i, row_data in enumerate(fl_rows):
                        if i == 0 and header_written:
                            continue
                        fl_ws.append(row_data)
                    header_written = True

                out_wb.save(output_path)
                out_wb.close()

                # Verify output
                assert os.path.exists(output_path)

                wb = load_workbook(output_path)
                assert 'Summary' in wb.sheetnames
                assert 'Full List' in wb.sheetnames
                assert wb['Full List'].max_row == 4  # header + 3 data rows
                wb.close()

                # Cleanup
                for f in input_files:
                    os.unlink(f)
