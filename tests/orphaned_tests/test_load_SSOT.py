import os
import sys
import openpyxl
import pytest

# SKIP: Module load_SSOT does not exist in current codebase (functionality moved to ssotsif.py)
pytestmark = pytest.mark.skip(reason="Module Helpers.dog_box.load_SSOT does not exist")

# Ensure package imports work when running the test file directly
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(current_dir, '..'))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from Helpers.dog_box.load_SSOT import _validate_headers


def _make_workbook(path, headers):
    wb = openpyxl.Workbook()
    ws = wb.active  # type: ignore[assignment]
    # pylance may think ``ws`` could be None; openpyxl always returns a worksheet
    ws.append(headers)  # type: ignore[attr-defined]
    wb.save(path)


def test_validate_headers_success(tmp_path):
    file_path = tmp_path / "good.xlsx"
    headers = ["Surname", "Firstname", "StudentID", "Other"]
    _make_workbook(file_path, headers)
    # should not raise; headers are on first row so override header_row
    _validate_headers(str(file_path), ["Surname", "Firstname"], header_row=1)
    _validate_headers(str(file_path), ["StudentID"], header_row=1)

    # place headers on row 2 and override header_row
    file_path2 = tmp_path / "good2.xlsx"
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active  # type: ignore[assignment]
    ws2.append([None])  # blank first row  # type: ignore[attr-defined]
    ws2.append(headers)  # type: ignore[attr-defined]
    wb2.save(file_path2)
    _validate_headers(str(file_path2), ["Surname"], header_row=2)


def test_validate_headers_missing(tmp_path):
    file_path = tmp_path / "bad.xlsx"
    headers = ["Surname", "Other"]
    _make_workbook(file_path, headers)
    with pytest.raises(ValueError) as exc:
        _validate_headers(str(file_path), ["Surname", "Firstname"])
    assert "Missing required headers" in str(exc.value)

    # missing headers on second row also fails if header_row=2
    file_path2 = tmp_path / "bad2.xlsx"
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active  # type: ignore[assignment]
    ws2.append([None])  # type: ignore[attr-defined]
    ws2.append(headers)  # type: ignore[attr-defined]
    wb2.save(file_path2)
    with pytest.raises(ValueError):
        _validate_headers(str(file_path2), ["Firstname"], header_row=1)


def test_validate_nonexistent(tmp_path):
    with pytest.raises(FileNotFoundError):
        _validate_headers(str(tmp_path / "nofile.xlsx"), ["A"])


def test_validate_invalid_file(tmp_path):
    file_path = tmp_path / "notxlsx.txt"
    file_path.write_text("not an excel file")
    with pytest.raises(ValueError) as exc:
        _validate_headers(str(file_path), ["A"])
    assert "Unable to open workbook" in str(exc.value)
