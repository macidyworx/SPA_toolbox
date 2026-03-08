"""
Tests for format-specific readers.
"""

import csv
import os
import pytest

from openpyxl import Workbook

from Finders.File_sorter.readers.xlsx_reader import XlsxReader
from Finders.File_sorter.readers.xlsm_reader import XlsmReader
from Finders.File_sorter.readers.csv_reader import CsvReader
from Finders.File_sorter.readers import READERS


# === FIXTURES ===

@pytest.fixture
def xlsx_file(tmp_path):
    """Create a test .xlsx file with known values."""
    path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "PAT Reading 5th Edition"
    ws["B2"] = "Running Record"
    ws["C1"] = "  Extra Spaces  "
    ws["A3"] = ""
    ws["D5"] = "Deep Cell"
    wb.save(str(path))
    wb.close()
    return str(path)


@pytest.fixture
def xlsx_named_sheet(tmp_path):
    """Create a .xlsx file with a named sheet."""
    path = tmp_path / "named.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "MySheet"
    ws["A1"] = "Named Sheet Value"
    wb.save(str(path))
    wb.close()
    return str(path)


@pytest.fixture
def xlsm_file(tmp_path):
    """Create a test .xlsm file with known values."""
    path = tmp_path / "test.xlsm"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "XLSM Test Value"
    ws["B2"] = "Second Cell"
    wb.save(str(path))
    wb.close()
    return str(path)


@pytest.fixture
def csv_file(tmp_path):
    """Create a test .csv file with known values."""
    path = tmp_path / "test.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["FirstName", "Surname", "StudentNumber"])
        writer.writerow(["Jane", "Doe", "12345"])
        writer.writerow(["John", "Smith", "67890"])
    return str(path)


@pytest.fixture
def csv_bom_file(tmp_path):
    """Create a .csv file with BOM encoding."""
    path = tmp_path / "bom.csv"
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["BOM Header", "Value"])
    return str(path)


# === TESTS: READERS registry ===

class TestReadersRegistry:
    """Tests for the READERS format mapping."""

    def test_xlsx_registered(self):
        assert ".xlsx" in READERS

    def test_xlsm_registered(self):
        assert ".xlsm" in READERS

    def test_xls_registered(self):
        assert ".xls" in READERS

    def test_csv_registered(self):
        assert ".csv" in READERS


# === TESTS: XlsxReader ===

class TestXlsxReader:
    """Tests for XlsxReader."""

    def test_read_cell_a1(self, xlsx_file):
        reader = XlsxReader()
        result = reader.read_cell(xlsx_file, 0, "A1")
        assert result == "patreading5thedition"

    def test_read_cell_b2(self, xlsx_file):
        reader = XlsxReader()
        result = reader.read_cell(xlsx_file, 0, "B2")
        assert result == "runningrecord"

    def test_read_cell_normalizes_spaces(self, xlsx_file):
        reader = XlsxReader()
        result = reader.read_cell(xlsx_file, 0, "C1")
        assert result == "extraspaces"

    def test_read_cell_empty(self, xlsx_file):
        reader = XlsxReader()
        result = reader.read_cell(xlsx_file, 0, "A3")
        assert result == ""

    def test_read_cell_beyond_data(self, xlsx_file):
        reader = XlsxReader()
        result = reader.read_cell(xlsx_file, 0, "Z99")
        assert result == ""

    def test_read_cell_by_sheet_name(self, xlsx_named_sheet):
        reader = XlsxReader()
        result = reader.read_cell(xlsx_named_sheet, "MySheet", "A1")
        assert result == "namedsheetvalue"

    def test_scan_area(self, xlsx_file):
        reader = XlsxReader()
        results = reader.scan_area(xlsx_file, 0, max_rows=10, max_cols=10)
        assert "patreading5thedition" in results
        assert "runningrecord" in results
        assert "deepcell" in results

    def test_scan_area_respects_limits(self, xlsx_file):
        reader = XlsxReader()
        results = reader.scan_area(xlsx_file, 0, max_rows=2, max_cols=2)
        assert "patreading5thedition" in results
        assert "runningrecord" in results
        assert "deepcell" not in results


# === TESTS: XlsmReader ===

class TestXlsmReader:
    """Tests for XlsmReader (inherits from XlsxReader)."""

    def test_read_cell(self, xlsm_file):
        reader = XlsmReader()
        result = reader.read_cell(xlsm_file, 0, "A1")
        assert result == "xlsmtestvalue"

    def test_scan_area(self, xlsm_file):
        reader = XlsmReader()
        results = reader.scan_area(xlsm_file, 0)
        assert "xlsmtestvalue" in results
        assert "secondcell" in results

    def test_is_subclass_of_xlsx(self):
        assert issubclass(XlsmReader, XlsxReader)


# === TESTS: CsvReader ===

class TestCsvReader:
    """Tests for CsvReader."""

    def test_read_cell_a1(self, csv_file):
        reader = CsvReader()
        result = reader.read_cell(csv_file, 0, "A1")
        assert result == "firstname"

    def test_read_cell_b1(self, csv_file):
        reader = CsvReader()
        result = reader.read_cell(csv_file, 0, "B1")
        assert result == "surname"

    def test_read_cell_c1(self, csv_file):
        reader = CsvReader()
        result = reader.read_cell(csv_file, 0, "C1")
        assert result == "studentnumber"

    def test_read_cell_a2(self, csv_file):
        reader = CsvReader()
        result = reader.read_cell(csv_file, 0, "A2")
        assert result == "jane"

    def test_read_cell_beyond_rows(self, csv_file):
        reader = CsvReader()
        result = reader.read_cell(csv_file, 0, "A99")
        assert result == ""

    def test_read_cell_beyond_cols(self, csv_file):
        reader = CsvReader()
        result = reader.read_cell(csv_file, 0, "Z1")
        assert result == ""

    def test_read_cell_bom_handled(self, csv_bom_file):
        reader = CsvReader()
        result = reader.read_cell(csv_bom_file, 0, "A1")
        assert result == "bomheader"

    def test_scan_area(self, csv_file):
        reader = CsvReader()
        results = reader.scan_area(csv_file, 0, max_rows=10, max_cols=10)
        assert "firstname" in results
        assert "surname" in results
        assert "jane" in results
        assert "12345" in results

    def test_scan_area_respects_row_limit(self, csv_file):
        reader = CsvReader()
        results = reader.scan_area(csv_file, 0, max_rows=1, max_cols=10)
        assert "firstname" in results
        assert "jane" not in results

    def test_scan_area_respects_col_limit(self, csv_file):
        reader = CsvReader()
        results = reader.scan_area(csv_file, 0, max_rows=10, max_cols=1)
        assert "firstname" in results
        assert "surname" not in results


# === TESTS: Normalization consistency ===

class TestNormalization:
    """Verify all readers normalize text consistently via field_cleaner."""

    def test_xlsx_normalizes(self, xlsx_file):
        """XlsxReader strips spaces and lowercases."""
        reader = XlsxReader()
        result = reader.read_cell(xlsx_file, 0, "A1")
        assert result == "patreading5thedition"
        assert " " not in result
        assert result == result.lower()

    def test_csv_normalizes(self, csv_file):
        """CsvReader strips spaces and lowercases."""
        reader = CsvReader()
        result = reader.read_cell(csv_file, 0, "A1")
        assert result == "firstname"
        assert " " not in result
        assert result == result.lower()
