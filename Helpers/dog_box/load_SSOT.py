# === CODE OVERVIEW DOCSTRING STARTS HERE ===
"""
Simple helper for selecting a single "single source of truth" Excel file
(usually an XLSX workbook).  The caller provides a list of headers that must
exist in the first row of the spreadsheet; the module verifies the selection
before returning.

Usage pattern:

    from tool_box.Helpers.dog_box.load_SSOT import select_ssot_file

    required = ["Surname", "Firstname", "StudentID"]
    path = select_ssot_file(required)
    if path is None:
        # user cancelled or file invalid
        return
    # proceed with valid workbook path

The returned path will always refer to an existing `.xlsx` file that contains
all of the requested headers (comparison is case‑insensitive and normalized via
`field_cleaner`).  By default headers are expected on row 2; the caller may
pass a different ``header_row`` value when invoking :func:`select_ssot_file`.
"""
# === CODE OVERVIEW DOCSTRING STOPS HERE ===

# === IMPORTS STARTS HERE ===
import os
import sys
from typing import List, Optional

# add project root to Python path for tool_box imports
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(current_dir, '..'))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

import wx
import openpyxl

from tool_box.Helpers.Clean_fields.clean_field import field_cleaner
from tool_box.Helpers.water_logged.the_logger import THElogger
# === IMPORTS STOPS HERE ===

# === GLOBAL CONSTANTS STARTS HERE ===
LOG_FOLDER = "./logs"
TO_CONSOLE = True
TO_LOGFILE = True
DEBUG_ON = False
# === GLOBAL CONSTANTS STOPS HERE ===

# === HELPER FUNCTIONS STARTS HERE ===

def _validate_headers(file_path: str, required_columns: List[str], header_row: int = 2) -> None:
    """Ensure ``file_path`` contains the requested headers on the designated row.

    By default the header row is 2 (to allow a title or description in row 1).
    The caller may override this by passing ``header_row`` (e.g. 1 for the very
    first row).

    Args:
        file_path: Path to an existing `.xlsx` workbook.
        required_columns: List of header names that must appear (original
            casing retained).

    Raises:
        FileNotFoundError: If ``file_path`` does not exist.
        ValueError: If the workbook is not a valid `.xlsx` file or if any of the
            required headers are missing.
    """
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:  # openpyxl will raise for invalid files
        raise ValueError(f"Unable to open workbook: {e}")

    # explicitly annotate as Worksheet so pylance knows methods exist
    from openpyxl.worksheet.worksheet import Worksheet
    ws: Worksheet = wb.active  # type: ignore[assignment]
    # iter_rows is correctly available on a Worksheet; pylance sometimes
    # thinks ``ws`` could be None.  Silence with type ignore.
    row = next(
        ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True),
        (),
    )  # type: ignore[attr-defined]
    headers = [field_cleaner(cell) for cell in row if cell is not None]

    missing = [hdr for hdr in required_columns if field_cleaner(hdr) not in headers]
    if missing:
        raise ValueError(f"Missing required headers: {missing}")


# === HELPER FUNCTIONS STOPS HERE ===

# === MAIN FUNCTIONS/CLASSES STARTS HERE ===

def select_ssot_file(required_columns: List[str], header_row: int = 2) -> Optional[str]:
    """Prompt user for a single `.xlsx` file and validate its headers.

    The dialog restricts selection to one existing Excel workbook.  After the
    user picks a file, the workbook is opened and the top row is compared against
    ``required_columns`` using `field_cleaner` normalization.  If validation
    fails the user is shown an error message and ``None`` is returned.

    Args:
        required_columns: Headers that must exist in the worksheet.
        header_row: Row number where the headers are located (default 2).

    Returns:
        Absolute path to the selected file, or ``None`` if the user cancelled or
        chose an invalid workbook.
    """
    logger = THElogger(
        script_name="load_SSOT",
        log_folder=LOG_FOLDER,
        to_console=TO_CONSOLE,
        to_logfile=TO_LOGFILE,
        debug_on=DEBUG_ON,
    )

    app = wx.App(False)
    style = wx.FD_OPEN | wx.FD_FILE_MUST_EXIST
    wildcard = "Excel files (*.xlsx)|*.xlsx"
    dlg = wx.FileDialog(None, message="Select SSOT workbook", wildcard=wildcard, style=style)
    dlg.CentreOnScreen()
    result = dlg.ShowModal()
    if result != wx.ID_OK:
        dlg.Destroy()
        logger.info("User cancelled SSOT file selection.")
        logger.finalize_report()
        return None

    path = dlg.GetPath()
    dlg.Destroy()
    logger.info(f"User selected file: {path}")

    try:
        _validate_headers(path, required_columns, header_row)
        logger.info("Header validation succeeded.")
        logger.finalize_report()
        return path
    except Exception as exc:
        msg = f"Selected file is invalid:\n{exc}"
        logger.error(msg)
        wx.MessageDialog(None, msg, "Validation error", wx.OK | wx.ICON_ERROR).ShowModal()
        logger.finalize_report()
        return None


# === MAIN FUNCTIONS/CLASSES STOPS HERE ===

# === STANDALONE EXECUTION STARTS HERE ===
if __name__ == "__main__":
    # quick demo: require some common headers
    required = ["Surname", "Firstname", "StudentID"]
    select_ssot_file(required)
# === STANDALONE EXECUTION STOPS HERE ===
