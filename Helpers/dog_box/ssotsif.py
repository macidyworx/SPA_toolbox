"""
ssotsif.py
Helpers for selecting and validating a single Excel file as either a
SIF (Student Import File) or SSOT (Single Source of Truth).

SIF:  Headers are fixed in row 2 — CalendarYear, YearLevel, Surname,
      Firstname, StudentID in columns A-E.  Validation checks these exist.

SSOT: User is prompted for header row, oldID column, and newID column
      via a wx dialog.  No fixed header expectations.
"""

import os
import wx
import openpyxl
from openpyxl.utils import column_index_from_string
from typing import Optional


SIF_HEADERS = {
    1: "CalendarYear",
    2: "YearLevel",
    3: "Surname",
    4: "Firstname",
    5: "StudentID",
}


def _ensure_app():
    """Return the existing wx.App or create one if none exists."""
    app = wx.GetApp()
    if app is None:
        app = wx.App(False)
    return app


def _pick_file(title="Select Excel file"):
    """Open a file dialog for a single .xlsx file. Returns path or None."""
    _ensure_app()
    wildcard = "Excel files (*.xlsx)|*.xlsx"
    style = wx.FD_OPEN | wx.FD_FILE_MUST_EXIST
    dlg = wx.FileDialog(None, message=title, wildcard=wildcard, style=style)
    dlg.CentreOnScreen()
    try:
        if dlg.ShowModal() == wx.ID_OK:
            return dlg.GetPath()
        return None
    finally:
        dlg.Destroy()


def _validate_sif(file_path):
    """Check that row 2 contains the required SIF headers in columns A-E.

    Returns:
        None on success.

    Raises:
        ValueError with details of any mismatches.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        problems = []
        for col_num, expected in SIF_HEADERS.items():
            cell_value = ws.cell(row=2, column=col_num).value
            actual = str(cell_value).strip() if cell_value is not None else ""
            if actual.lower() != expected.lower():
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(col_num)
                problems.append(
                    f"  Column {col_letter}: expected '{expected}', got '{actual or '(empty)'}'"
                )
        if problems:
            raise ValueError(
                "SIF header validation failed (row 2):\n" + "\n".join(problems)
            )
    finally:
        wb.close()


def select_sif():
    """Prompt user to select a SIF Excel file and validate its headers.

    Row 2 must contain: CalendarYear (A), YearLevel (B), Surname (C),
    Firstname (D), StudentID (E).

    Returns:
        str: Path to the validated file, or None if cancelled/invalid.
    """
    _ensure_app()
    path = _pick_file("Select SIF (Student Import File)")
    if path is None:
        return None

    try:
        _validate_sif(path)
        return path
    except ValueError as exc:
        wx.MessageBox(str(exc), "SIF Validation Error", wx.OK | wx.ICON_ERROR)
        return None


class _SSOTDialog(wx.Dialog):
    """Dialog prompting the user for SSOT parameters: header row, oldID and newID columns."""

    def __init__(self, parent=None):
        super().__init__(parent, title="SSOT Settings", style=wx.DEFAULT_DIALOG_STYLE)

        sizer = wx.FlexGridSizer(rows=3, cols=2, vgap=8, hgap=12)

        sizer.Add(wx.StaticText(self, label="Header row number:"), 0, wx.ALIGN_CENTER_VERTICAL)
        self.header_row_ctrl = wx.SpinCtrl(self, value="1", min=1, max=200)
        sizer.Add(self.header_row_ctrl, 0, wx.EXPAND)

        sizer.Add(wx.StaticText(self, label="Old ID column (e.g. A):"), 0, wx.ALIGN_CENTER_VERTICAL)
        self.old_id_ctrl = wx.TextCtrl(self, value="")
        sizer.Add(self.old_id_ctrl, 0, wx.EXPAND)

        sizer.Add(wx.StaticText(self, label="New ID column (e.g. B):"), 0, wx.ALIGN_CENTER_VERTICAL)
        self.new_id_ctrl = wx.TextCtrl(self, value="")
        sizer.Add(self.new_id_ctrl, 0, wx.EXPAND)

        btn_sizer = self.CreateStdDialogButtonSizer(wx.OK | wx.CANCEL)

        outer = wx.BoxSizer(wx.VERTICAL)
        outer.Add(sizer, 0, wx.ALL | wx.EXPAND, 12)
        outer.Add(btn_sizer, 0, wx.ALL | wx.EXPAND, 8)
        self.SetSizerAndFit(outer)
        self.CentreOnScreen()

    def get_values(self):
        """Return (header_row, old_id_col, new_id_col) or raise ValueError."""
        header_row = self.header_row_ctrl.GetValue()

        old_id = self.old_id_ctrl.GetValue().strip().upper()
        new_id = self.new_id_ctrl.GetValue().strip().upper()

        if not old_id or not new_id:
            raise ValueError("Both Old ID and New ID columns are required.")

        # Validate column letters
        try:
            column_index_from_string(old_id)
        except ValueError:
            raise ValueError(f"'{old_id}' is not a valid column letter.")
        try:
            column_index_from_string(new_id)
        except ValueError:
            raise ValueError(f"'{new_id}' is not a valid column letter.")

        if old_id == new_id:
            raise ValueError("Old ID and New ID columns must be different.")

        return header_row, old_id, new_id


def select_ssot():
    """Prompt user to select an SSOT file and provide header row / ID columns.

    Returns:
        dict with keys 'path', 'header_row', 'old_id_col', 'new_id_col',
        or None if the user cancelled.
    """
    _ensure_app()
    path = _pick_file("Select SSOT (Single Source of Truth)")
    if path is None:
        return None

    dlg = _SSOTDialog()
    try:
        if dlg.ShowModal() != wx.ID_OK:
            return None
        header_row, old_id, new_id = dlg.get_values()
    except ValueError as exc:
        wx.MessageBox(str(exc), "SSOT Input Error", wx.OK | wx.ICON_ERROR)
        return None
    finally:
        dlg.Destroy()

    return {
        "path": path,
        "header_row": header_row,
        "old_id_col": old_id,
        "new_id_col": new_id,
    }


def select_single_file(mode="sif"):
    """Convenience wrapper — pick mode 'sif', 'ssot', or 'choose' to let the user decide.

    Args:
        mode: 'sif' for Student Import File, 'ssot' for Single Source of Truth,
              'choose' to let the user pick via a dialog.

    Returns:
        For 'sif': str path or None.
        For 'ssot': dict with path/header_row/old_id_col/new_id_col, or None.
        For 'choose': either of the above depending on user choice, or None.
    """
    mode = mode.lower()
    if mode == "sif":
        return select_sif()
    if mode == "ssot":
        return select_ssot()
    if mode == "choose":
        _ensure_app()
        dlg = wx.SingleChoiceDialog(
            None, "What type of file are you loading?", "File Type",
            ["SIF (Student Import File)", "SSOT (Single Source of Truth)"],
        )
        dlg.CentreOnScreen()
        try:
            if dlg.ShowModal() != wx.ID_OK:
                return None
            choice = dlg.GetSelection()
        finally:
            dlg.Destroy()
        return select_sif() if choice == 0 else select_ssot()

    raise ValueError(f"Unknown mode '{mode}'. Use 'sif', 'ssot', or 'choose'.")
