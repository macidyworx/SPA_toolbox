import pytest
import os
from openpyxl import Workbook
from Helpers.Last_row_finder.real_last_row import get_last_row, ws_last_row


@pytest.fixture
def sample_xlsx(tmp_path):
    """Create a test xlsx with known data and some empty trailing rows."""
    path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Column A: data in rows 1-5
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=f"A{i}")

    # Column B: data in rows 1-3
    for i in range(1, 4):
        ws.cell(row=i, column=2, value=f"B{i}")

    # Column C: empty (but touch row 10 to inflate max_row)
    ws.cell(row=10, column=3, value=None)

    # Column D: only whitespace in row 2
    ws.cell(row=1, column=4, value="D1")
    ws.cell(row=2, column=4, value="   ")

    wb.save(path)
    wb.close()
    return str(path)


@pytest.fixture
def multi_sheet_xlsx(tmp_path):
    """Create a test xlsx with multiple sheets."""
    path = tmp_path / "multi.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "First"
    ws1.cell(row=1, column=1, value="hello")

    ws2 = wb.create_sheet("Second")
    ws2.cell(row=1, column=1, value="a")
    ws2.cell(row=2, column=1, value="b")
    ws2.cell(row=3, column=1, value="c")

    wb.save(path)
    wb.close()
    return str(path)


# --- get_last_row tests ---

def test_last_row_column_a(sample_xlsx):
    assert get_last_row(sample_xlsx, "Sheet1", "A") == 5

def test_last_row_column_b(sample_xlsx):
    assert get_last_row(sample_xlsx, "Sheet1", "B") == 3

def test_last_row_empty_column(sample_xlsx):
    assert get_last_row(sample_xlsx, "Sheet1", "C") is None

def test_last_row_whitespace_only(sample_xlsx):
    """Whitespace-only cells should not count as data."""
    assert get_last_row(sample_xlsx, "Sheet1", "D") == 1

def test_last_row_column_number(sample_xlsx):
    assert get_last_row(sample_xlsx, "Sheet1", 1) == 5

def test_last_row_sheet_by_index(sample_xlsx):
    assert get_last_row(sample_xlsx, 0, "A") == 5

def test_last_row_sheet_name_not_found(sample_xlsx):
    with pytest.raises(ValueError, match="not found"):
        get_last_row(sample_xlsx, "NoSuchSheet", "A")

def test_last_row_sheet_index_out_of_range(sample_xlsx):
    with pytest.raises(ValueError, match="out of range"):
        get_last_row(sample_xlsx, 99, "A")

def test_last_row_multi_sheet_by_name(multi_sheet_xlsx):
    assert get_last_row(multi_sheet_xlsx, "Second", "A") == 3

def test_last_row_multi_sheet_by_index(multi_sheet_xlsx):
    assert get_last_row(multi_sheet_xlsx, 1, "A") == 3


# --- ws_last_row tests ---

def test_ws_last_row_direct(sample_xlsx):
    """ws_last_row works with an already-open worksheet."""
    from openpyxl import load_workbook
    wb = load_workbook(sample_xlsx, data_only=True)
    ws = wb["Sheet1"]
    assert ws_last_row(ws, "A") == 5
    assert ws_last_row(ws, "B") == 3
    assert ws_last_row(ws, "C") is None
    wb.close()
