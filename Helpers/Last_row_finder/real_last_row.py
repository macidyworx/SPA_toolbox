"""
real_last_row.py
Find the actual last row with data in an Excel column.
Works around openpyxl's max_row reporting phantom rows from activated-but-empty cells.
"""

# === IMPORTS ===
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


# === MAIN FUNCTIONS/CLASSES ===
def ws_last_row(ws, column):
    """
    Find the last row with data in a specific column of an already-open worksheet.

    Args:
        ws: An openpyxl Worksheet object.
        column (str or int): Column letter (e.g., 'A') or number (e.g., 1).

    Returns:
        int or None: The last row number with data, or None if the column is empty.
    """
    col_num = column_index_from_string(column) if isinstance(column, str) else column

    for row in range(ws.max_row, 0, -1):
        cell = ws.cell(row=row, column=col_num)
        if cell.value is not None and str(cell.value).strip() != "":
            return row

    return None


def get_last_row(file_path, sheet, column):
    """
    Open an Excel file and find the last row with data in a specific column.

    Convenience wrapper around ws_last_row for one-off lookups.
    If you need multiple columns from the same file, open the workbook yourself
    and call ws_last_row directly to avoid re-reading the file each time.

    Args:
        file_path (str): Path to the Excel file.
        sheet (str or int): Sheet name (str) or 0-based index (int).
        column (str or int): Column letter (e.g., 'A') or number (e.g., 1).

    Returns:
        int or None: The last row number with data, or None if the column is empty.
    """
    wb = load_workbook(file_path, data_only=True)
    try:
        if isinstance(sheet, int):
            if sheet < 0 or sheet >= len(wb.worksheets):
                raise ValueError(f"Sheet index {sheet} is out of range. Valid indices: 0 to {len(wb.worksheets)-1}")
            ws = wb.worksheets[sheet]
        else:
            if sheet not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in the workbook.")
            ws = wb[sheet]

        return ws_last_row(ws, column)
    finally:
        wb.close()
