"""
test_PATUniqueID_Checker.py

Comprehensive test suite for PATUniqueID_Checker module.
Tests format validators, column detection, file validation, file moves, and edge cases.
"""

import os
import tempfile
import shutil
import sys
from pathlib import Path
from unittest import mock

import pytest
import wx
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Add project root to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))

from RandomTools.PAT.PATUniqueID_Checker import (
    PATUniqueIDChecker,
    validate_unique_id_format,
    find_unique_id_column,
)


class TestFormatValidators:
    """Test format validators for each Unique ID pattern."""

    def test_validate_abc0001_valid(self):
        """ABC0001 format: accept valid values."""
        checker = PATUniqueIDChecker()
        assert checker.validate_unique_id_format("ABC0001", "ABC0001") is True
        assert checker.validate_unique_id_format("xyz1234", "ABC0001") is True
        assert checker.validate_unique_id_format("AAA0000", "ABC0001") is True
        assert checker.validate_unique_id_format("ZZZ9999", "ABC0001") is True

    def test_validate_abc0001_invalid(self):
        """ABC0001 format: reject invalid values."""
        checker = PATUniqueIDChecker()
        assert checker.validate_unique_id_format("ab1234", "ABC0001") is False  # Only 2 letters
        assert checker.validate_unique_id_format("ABC00001", "ABC0001") is False  # 5 digits
        assert checker.validate_unique_id_format("ABCD0001", "ABC0001") is False  # 4 letters
        assert checker.validate_unique_id_format("", "ABC0001") is False
        assert checker.validate_unique_id_format(None, "ABC0001") is False

    def test_validate_abc0001_whitespace(self):
        """ABC0001 format: strip whitespace before validation."""
        checker = PATUniqueIDChecker()
        assert checker.validate_unique_id_format("  ABC0001  ", "ABC0001") is True
        assert checker.validate_unique_id_format("\tABC0001\n", "ABC0001") is True

    def test_validate_182815548_valid(self):
        """182815548 format: accept valid 9-digit values."""
        checker = PATUniqueIDChecker()
        assert checker.validate_unique_id_format("123456789", "182815548") is True
        assert checker.validate_unique_id_format("000000000", "182815548") is True
        assert checker.validate_unique_id_format("999999999", "182815548") is True

    def test_validate_182815548_invalid(self):
        """182815548 format: reject invalid values."""
        checker = PATUniqueIDChecker()
        assert checker.validate_unique_id_format("12345678", "182815548") is False  # 8 digits
        assert checker.validate_unique_id_format("1234567890", "182815548") is False  # 10 digits
        assert checker.validate_unique_id_format("12345678a", "182815548") is False  # Contains letter
        assert checker.validate_unique_id_format("", "182815548") is False
        assert checker.validate_unique_id_format(None, "182815548") is False

    def test_validate_182815548_whitespace(self):
        """182815548 format: strip whitespace before validation."""
        checker = PATUniqueIDChecker()
        assert checker.validate_unique_id_format("  123456789  ", "182815548") is True

    def test_validate_5548_valid(self):
        """5548 format: accept valid 4-digit values."""
        checker = PATUniqueIDChecker()
        assert checker.validate_unique_id_format("1234", "5548") is True
        assert checker.validate_unique_id_format("9999", "5548") is True
        assert checker.validate_unique_id_format("0000", "5548") is True

    def test_validate_5548_invalid(self):
        """5548 format: reject invalid values."""
        checker = PATUniqueIDChecker()
        assert checker.validate_unique_id_format("123", "5548") is False  # 3 digits
        assert checker.validate_unique_id_format("12345", "5548") is False  # 5 digits
        assert checker.validate_unique_id_format("123a", "5548") is False  # Contains letter
        assert checker.validate_unique_id_format("", "5548") is False
        assert checker.validate_unique_id_format(None, "5548") is False

    def test_validate_5548_whitespace(self):
        """5548 format: strip whitespace before validation."""
        checker = PATUniqueIDChecker()
        assert checker.validate_unique_id_format("  1234  ", "5548") is True


class TestColumnDetection:
    """Test Unique ID column detection with various header placements."""

    def test_detect_unique_id_column_a1(self):
        """Header in A1 -> correctly identified."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Unique ID"
        ws['A2'] = "ABC0001"

        checker = PATUniqueIDChecker()
        result = checker.find_unique_id_column(ws)
        assert result == 'A'

    def test_detect_unique_id_column_c3(self):
        """Header in C3 -> correctly identified."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Family name"
        ws['B1'] = "Given name"
        ws['C3'] = "Unique ID"

        checker = PATUniqueIDChecker()
        result = checker.find_unique_id_column(ws)
        assert result == 'C'

    def test_detect_unique_id_column_with_whitespace(self):
        """Header with whitespace variations -> found and normalized."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "  Unique ID  "

        checker = PATUniqueIDChecker()
        result = checker.find_unique_id_column(ws)
        assert result == 'A'

    def test_detect_unique_id_column_with_case_variations(self):
        """Header with case variations -> found via field_cleaner."""
        wb = Workbook()
        ws = wb.active
        ws['B1'] = "unique id"
        ws['B2'] = "ABC0001"

        checker = PATUniqueIDChecker()
        result = checker.find_unique_id_column(ws)
        assert result == 'B'

    def test_detect_unique_id_column_not_found(self):
        """Header missing -> returns None."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Family name"
        ws['B1'] = "Given name"

        checker = PATUniqueIDChecker()
        result = checker.find_unique_id_column(ws)
        assert result is None

    def test_detect_unique_id_column_beyond_scan_range(self):
        """Header beyond scan range (outside A-M, rows 1-20) -> returns None."""
        wb = Workbook()
        ws = wb.active
        ws['N1'] = "Unique ID"  # Column N is outside scan range

        checker = PATUniqueIDChecker()
        result = checker.find_unique_id_column(ws)
        assert result is None


class TestFileValidation:
    """Test file validation logic with various scenarios."""

    def test_validate_file_abc0001_valid(self):
        """File with all valid ABC0001 format values -> returns (True, ...)."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = "ABC0001"
            ws['A3'] = "XYZ5678"
            ws['A4'] = "DEF9999"
            wb.save(file_path)

            checker = PATUniqueIDChecker()
            is_valid, reason = checker.validate_file(file_path, "ABC0001")
            assert is_valid is True
            assert "match format" in reason.lower()

    def test_validate_file_182815548_valid(self):
        """File with all valid 9-digit values -> returns (True, ...)."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = "123456789"
            ws['A3'] = "987654321"
            wb.save(file_path)

            checker = PATUniqueIDChecker()
            is_valid, reason = checker.validate_file(file_path, "182815548")
            assert is_valid is True

    def test_validate_file_5548_valid(self):
        """File with all valid 4-digit values -> returns (True, ...)."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = "1234"
            ws['A3'] = "5678"
            wb.save(file_path)

            checker = PATUniqueIDChecker()
            is_valid, reason = checker.validate_file(file_path, "5548")
            assert is_valid is True

    def test_validate_file_empty_cells(self):
        """File with empty cells in Unique ID column -> returns (False, "Empty...")."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = "ABC0001"
            ws['A3'] = None  # Empty cell
            wb.save(file_path)

            checker = PATUniqueIDChecker()
            is_valid, reason = checker.validate_file(file_path, "ABC0001")
            assert is_valid is False
            assert "Empty" in reason

    def test_validate_file_format_mismatch(self):
        """File with format mismatch -> returns (False, "Invalid format...")."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = "ABC00001"  # 5 digits instead of 4
            wb.save(file_path)

            checker = PATUniqueIDChecker()
            is_valid, reason = checker.validate_file(file_path, "ABC0001")
            assert is_valid is False
            assert "Invalid format" in reason

    def test_validate_file_no_unique_id_column(self):
        """File without Unique ID column -> returns (False, "Unique ID column not found")."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Family name"
            ws['B1'] = "Given name"
            wb.save(file_path)

            checker = PATUniqueIDChecker()
            is_valid, reason = checker.validate_file(file_path, "ABC0001")
            assert is_valid is False
            assert "not found" in reason.lower()

    def test_validate_file_whitespace_in_values(self):
        """Values with whitespace -> stripped and validated correctly."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = "  ABC0001  "
            ws['A3'] = "\tDEF5678\n"
            wb.save(file_path)

            checker = PATUniqueIDChecker()
            is_valid, reason = checker.validate_file(file_path, "ABC0001")
            assert is_valid is True


class TestFileMoveWorkflow:
    """Test file move workflow with conflict handling."""

    def test_valid_file_stays_in_place(self):
        """Valid file stays in source folder."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = "ABC0001"
            wb.save(file_path)

            output_dir = os.path.join(tmpdir, "output")
            os.makedirs(output_dir)

            checker = PATUniqueIDChecker()
            result = checker.process_file(file_path, output_dir, "ABC0001")
            assert result is True
            assert os.path.exists(file_path)  # File still exists in original location

    def test_invalid_file_moved_to_invalid_folder(self):
        """Invalid file moved to Invalid_UniqueID/ subfolder."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = "ABC00001"  # Invalid: 5 digits
            wb.save(file_path)

            output_dir = os.path.join(tmpdir, "output")
            os.makedirs(output_dir)

            checker = PATUniqueIDChecker()
            result = checker.process_file(file_path, output_dir, "ABC0001")
            assert result is True
            assert not os.path.exists(file_path)  # Original file moved
            assert os.path.exists(os.path.join(output_dir, "Invalid_UniqueID", "test.xlsx"))

    def test_invalid_folder_created_if_needed(self):
        """Invalid_UniqueID/ folder created if not exists."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = None  # Empty cell
            wb.save(file_path)

            output_dir = os.path.join(tmpdir, "output")
            os.makedirs(output_dir)

            checker = PATUniqueIDChecker()
            checker.process_file(file_path, output_dir, "ABC0001")
            assert os.path.exists(os.path.join(output_dir, "Invalid_UniqueID"))

    @mock.patch('RandomTools.PAT.PATUniqueID_Checker.PATUniqueIDChecker._handle_file_conflict')
    def test_conflict_dialog_called_on_existing_file(self, mock_conflict):
        """Conflict dialog called when file exists at destination."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create source file
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = None  # Invalid
            wb.save(file_path)

            # Create output directory and pre-existing target file
            output_dir = os.path.join(tmpdir, "output")
            invalid_dir = os.path.join(output_dir, "Invalid_UniqueID")
            os.makedirs(invalid_dir)
            existing_file = os.path.join(invalid_dir, "test.xlsx")
            with open(existing_file, 'w') as f:
                f.write("existing")

            mock_conflict.return_value = "skip"

            checker = PATUniqueIDChecker()
            result = checker.process_file(file_path, output_dir, "ABC0001")
            assert mock_conflict.called

    @mock.patch('RandomTools.PAT.PATUniqueID_Checker.PATUniqueIDChecker._handle_file_conflict')
    def test_overwrite_action_removes_old_file(self, mock_conflict):
        """Overwrite action removes old file and moves new one."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create source file
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = None  # Invalid
            wb.save(file_path)

            # Create output directory and pre-existing target file
            output_dir = os.path.join(tmpdir, "output")
            invalid_dir = os.path.join(output_dir, "Invalid_UniqueID")
            os.makedirs(invalid_dir)
            existing_file = os.path.join(invalid_dir, "test.xlsx")
            with open(existing_file, 'w') as f:
                f.write("old content")

            mock_conflict.return_value = "overwrite"

            checker = PATUniqueIDChecker()
            result = checker.process_file(file_path, output_dir, "ABC0001")
            assert result is True
            assert os.path.exists(existing_file)  # File exists with new content
            assert not os.path.exists(file_path)  # Source moved

    @mock.patch('RandomTools.PAT.PATUniqueID_Checker.PATUniqueIDChecker._handle_file_conflict')
    def test_skip_action_leaves_both_files(self, mock_conflict):
        """Skip action leaves both files in place."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create source file
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = None  # Invalid
            wb.save(file_path)

            # Create output directory and pre-existing target file
            output_dir = os.path.join(tmpdir, "output")
            invalid_dir = os.path.join(output_dir, "Invalid_UniqueID")
            os.makedirs(invalid_dir)
            existing_file = os.path.join(invalid_dir, "test.xlsx")
            with open(existing_file, 'w') as f:
                f.write("old content")

            mock_conflict.return_value = "skip"

            checker = PATUniqueIDChecker()
            result = checker.process_file(file_path, output_dir, "ABC0001")
            assert result is False
            assert os.path.exists(file_path)  # Source still exists
            assert os.path.exists(existing_file)  # Existing target still exists

    @mock.patch('RandomTools.PAT.PATUniqueID_Checker.PATUniqueIDChecker._handle_file_conflict')
    def test_rename_action_appends_dup_suffix(self, mock_conflict):
        """Rename action appends _dup suffix."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create source file
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = None  # Invalid
            wb.save(file_path)

            # Create output directory and pre-existing target file
            output_dir = os.path.join(tmpdir, "output")
            invalid_dir = os.path.join(output_dir, "Invalid_UniqueID")
            os.makedirs(invalid_dir)
            existing_file = os.path.join(invalid_dir, "test.xlsx")
            with open(existing_file, 'w') as f:
                f.write("old content")

            mock_conflict.return_value = "rename"

            checker = PATUniqueIDChecker()
            result = checker.process_file(file_path, output_dir, "ABC0001")
            assert result is True
            assert not os.path.exists(file_path)  # Source moved
            assert os.path.exists(existing_file)  # Original target still exists
            assert os.path.exists(os.path.join(invalid_dir, "test_dup.xlsx"))  # Renamed file exists


class TestEdgeCases:
    """Test edge cases and special scenarios."""

    def test_zero_files_selected(self):
        """Zero files selected -> graceful exit."""
        with mock.patch('RandomTools.PAT.PATUniqueID_Checker.select_work_files', return_value=None):
            checker = PATUniqueIDChecker()
            # Should exit gracefully without error
            checker.run("ABC0001")

    def test_corrupted_file(self):
        """Corrupted/unreadable file -> caught and logged, left in place."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a non-Excel file
            file_path = os.path.join(tmpdir, "corrupted.xlsx")
            with open(file_path, 'w') as f:
                f.write("not an excel file")

            output_dir = os.path.join(tmpdir, "output")
            os.makedirs(output_dir)

            checker = PATUniqueIDChecker()
            is_valid, reason = checker.validate_file(file_path, "ABC0001")
            assert is_valid is False
            # Original file should still exist (not moved)
            assert os.path.exists(file_path)

    def test_output_folder_same_as_source(self):
        """Output folder same as source -> Invalid_UniqueID/ subfolder created within source."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create source file in tmpdir
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = None  # Invalid
            wb.save(file_path)

            checker = PATUniqueIDChecker()
            result = checker.process_file(file_path, tmpdir, "ABC0001")
            assert result is True
            # File moved to Invalid_UniqueID subfolder within tmpdir
            assert os.path.exists(os.path.join(tmpdir, "Invalid_UniqueID", "test.xlsx"))

    def test_large_filename(self):
        """Large filename -> handled correctly."""
        with tempfile.TemporaryDirectory() as tmpdir:
            long_name = "a" * 200 + ".xlsx"
            file_path = os.path.join(tmpdir, long_name)
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = None  # Invalid
            wb.save(file_path)

            output_dir = os.path.join(tmpdir, "output")
            os.makedirs(output_dir)

            checker = PATUniqueIDChecker()
            result = checker.process_file(file_path, output_dir, "ABC0001")
            assert result is True
            # File should be moved despite long name
            assert os.path.exists(os.path.join(output_dir, "Invalid_UniqueID", long_name))


class TestProgressCallback:
    """Test progress callback functionality."""

    def test_callback_called_with_correct_args(self):
        """Callback called with (index, total, filename)."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create test files
            files = []
            for i in range(3):
                file_path = os.path.join(tmpdir, f"test{i}.xlsx")
                wb = Workbook()
                ws = wb.active
                ws['A1'] = "Unique ID"
                ws['A2'] = "ABC0001"
                wb.save(file_path)
                files.append(file_path)

            output_dir = os.path.join(tmpdir, "output")
            os.makedirs(output_dir)

            calls = []

            def track_callback(index, total, filename):
                calls.append((index, total, filename))
                return True

            checker = PATUniqueIDChecker()

            with mock.patch('RandomTools.PAT.PATUniqueID_Checker.select_work_files', return_value=files):
                with mock.patch('RandomTools.PAT.PATUniqueID_Checker.select_output_folder', return_value=output_dir):
                    checker.run("ABC0001", progress_callback=track_callback)

            # Verify callback was called for each file
            assert len(calls) == 3
            assert calls[0] == (1, 3, "test0.xlsx")
            assert calls[1] == (2, 3, "test1.xlsx")
            assert calls[2] == (3, 3, "test2.xlsx")

    def test_callback_returning_false_cancels_processing(self):
        """Returning False from callback cancels processing."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create test files
            files = []
            for i in range(3):
                file_path = os.path.join(tmpdir, f"test{i}.xlsx")
                wb = Workbook()
                ws = wb.active
                ws['A1'] = "Unique ID"
                ws['A2'] = "ABC0001"
                wb.save(file_path)
                files.append(file_path)

            output_dir = os.path.join(tmpdir, "output")
            os.makedirs(output_dir)

            calls = []

            def cancel_after_one(index, total, filename):
                calls.append((index, total, filename))
                return index < 1  # Cancel after first call

            checker = PATUniqueIDChecker()

            with mock.patch('RandomTools.PAT.PATUniqueID_Checker.select_work_files', return_value=files):
                with mock.patch('RandomTools.PAT.PATUniqueID_Checker.select_output_folder', return_value=output_dir):
                    checker.run("ABC0001", progress_callback=cancel_after_one)

            # Verify callback was only called once
            assert len(calls) == 1

    def test_callback_exception_logged_and_continues(self):
        """Exception in callback -> logged as warning, processing continues."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create test files
            files = []
            for i in range(2):
                file_path = os.path.join(tmpdir, f"test{i}.xlsx")
                wb = Workbook()
                ws = wb.active
                ws['A1'] = "Unique ID"
                ws['A2'] = "ABC0001"
                wb.save(file_path)
                files.append(file_path)

            output_dir = os.path.join(tmpdir, "output")
            os.makedirs(output_dir)

            def failing_callback(index, total, filename):
                if index == 1:
                    raise ValueError("Test error")
                return True

            checker = PATUniqueIDChecker()

            with mock.patch('RandomTools.PAT.PATUniqueID_Checker.select_work_files', return_value=files):
                with mock.patch('RandomTools.PAT.PATUniqueID_Checker.select_output_folder', return_value=output_dir):
                    # Should not raise, should continue despite callback error
                    checker.run("ABC0001", progress_callback=failing_callback)


class TestTypeValidation:
    """Test type validation and error handling."""

    def test_progress_callback_type_validation(self):
        """Non-callable progress_callback raises TypeError."""
        checker = PATUniqueIDChecker()
        with pytest.raises(TypeError):
            checker.run("ABC0001", progress_callback="not_callable")

    def test_invalid_format_type(self):
        """Invalid format_type handled gracefully."""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Unique ID"
            ws['A2'] = "ABC0001"
            wb.save(file_path)

            checker = PATUniqueIDChecker()
            # Invalid format should be handled gracefully
            is_valid, reason = checker.validate_file(file_path, "INVALID_FORMAT")
            assert is_valid is False
