"""
PAT_Username_Checker.py

Auto-detects the username column in PATonline Excel files, validates all usernames
against a user-selected format (alphanumeric, long numeric, or short numeric),
and organizes files into categorized output folders based on validation results.

Dual-mode: Can be used as a standalone script or imported as a module.
"""

import os
import re
import shutil
import sys
import wx
from pathlib import Path
from typing import List, Optional

# Add project root to path so Helpers can be imported
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from Helpers.Clean_fields.clean_field import field_cleaner
from Helpers.Last_row_finder.real_last_row import ws_last_row
from Helpers.dog_box.work_files import select_work_files, select_output_folder
from water_logged.the_logger import THElogger


# === SCAN RANGE (matches PATonline_FINDER convention) ===
SCAN_COLUMNS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
SCAN_ROWS = range(1, 21)

# === STATUS CODES ===
STATUS_EXPECTED_ID = "EXPECTED_ID"
STATUS_FILES_TO_CHECK = "FILES_TO_CHECK"
STATUS_EMPTY_OR_UNREADABLE = "EMPTY_OR_UNREADABLE"


# === FORMAT VALIDATORS ===

def validate_alphanumeric(value: str) -> bool:
    """
    Validate alphanumeric format: ABC0001 or ABC-0001 (2 letters, then 1 letter or dash, then 4 digits).
    Case-insensitive. Normalizes input using field_cleaner.

    Args:
        value: The value to validate (string or convertible to string).

    Returns:
        True if matches pattern, False otherwise.
    """
    if value is None:
        return False

    # Normalize: remove spaces, convert to lowercase
    normalized = field_cleaner(str(value), lowercase=True, strip_spaces=True)

    # Pattern: exactly 2 letters, then 1 letter or dash, then 4 digits (total 7-8 chars)
    pattern = r'^[a-z]{2}[a-z\-]\d{4}$'
    return bool(re.match(pattern, normalized))


def validate_long_numeric(value: str) -> bool:
    """
    Validate long numeric format: 4-12 digits.
    Handles int or string input. Normalizes using field_cleaner.

    Args:
        value: The value to validate (string, int, or convertible).

    Returns:
        True if 4-12 digits, False otherwise.
    """
    if value is None:
        return False

    # Normalize: convert to string, strip whitespace
    normalized = field_cleaner(str(value), lowercase=False, strip_spaces=True, strip_bom=True)

    # Check: 4-12 digits only
    return bool(re.match(r'^\d{4,12}$', normalized))


def validate_short_numeric(value: str) -> bool:
    """
    Validate short numeric format: 2-8 digits.
    Handles int or string input. Normalizes using field_cleaner.

    Args:
        value: The value to validate (string, int, or convertible).

    Returns:
        True if 2-8 digits, False otherwise.
    """
    if value is None:
        return False

    # Normalize: convert to string, strip whitespace
    normalized = field_cleaner(str(value), lowercase=False, strip_spaces=True, strip_bom=True)

    # Check: 2-8 digits only
    return bool(re.match(r'^\d{2,8}$', normalized))


# === FORMAT SELECTION ===

def select_format() -> Optional[str]:
    """
    Display wxPython dialog to select username format.

    Returns:
        Format key ('alphanumeric', 'long_numeric', 'short_numeric'), or None if cancelled.
    """
    app = wx.GetApp()
    if app is None:
        app = wx.App(False)

    options = [
        "ABC0001 (alphanumeric: 2 letters, 1 letter/dash, 4 digits)",
        "182815548 (long numeric: 4-12 digits)",
        "5548 (short numeric: 2-8 digits)"
    ]

    dlg = wx.SingleChoiceDialog(
        None,
        "Select the username format:",
        "Username Format",
        options
    )
    dlg.CentreOnScreen()

    try:
        if dlg.ShowModal() != wx.ID_OK:
            return None
        choice = dlg.GetSelection()
        format_keys = ['alphanumeric', 'long_numeric', 'short_numeric']
        return format_keys[choice]
    finally:
        dlg.Destroy()


# === COLUMN DETECTION ===

def detect_username_column(file_path: str) -> Optional[tuple]:
    """
    Auto-detect the username column by scanning rows 1-20, columns A-M for header variants.

    Searches for: Username, User ID, UserID (case-insensitive, whitespace-tolerant).
    Uses field_cleaner for robust matching. Scans the same range as PATonline_FINDER.

    Args:
        file_path: Path to the Excel file to inspect.

    Returns:
        Tuple of (column_letter, header_row) if found, None if not detected.
    """
    # Header variants to search for (readable format, normalized at match time)
    SEARCH_HEADERS = {
        "Username": "username",
        "User ID": "userid",
        "UserID": "userid",
    }
    normalized_search = {field_cleaner(k, strip_spaces=True): v for k, v in SEARCH_HEADERS.items()}

    try:
        wb = load_workbook(file_path, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return None

            # Scan rows 1-20, columns A-M (matches PATonline_FINDER convention)
            for row in SCAN_ROWS:
                for col in SCAN_COLUMNS:
                    cell = ws[f'{col}{row}']
                    if cell.value is None:
                        continue

                    cell_text = field_cleaner(str(cell.value), strip_spaces=True)
                    if cell_text in normalized_search:
                        return (col, row)

            return None
        finally:
            wb.close()
    except Exception:
        return None


# === MAIN CLASS ===

class PAT_Username_Checker:
    """
    Validates usernames in PATonline Excel files against selected format
    and organizes files into categorized output folders.
    """

    def __init__(self, output_folder: str, selected_format: str):
        """
        Initialize the checker with output folder and selected format.

        Args:
            output_folder: Path to the directory where output subfolders will be created.
            selected_format: Format key ('alphanumeric', 'long_numeric', 'short_numeric').
        """
        self.output_folder = Path(output_folder)
        self.selected_format = selected_format

        # Initialize logger
        script_dir = os.path.dirname(os.path.abspath(__file__))
        config_path = os.path.join(script_dir, "logging.ini")
        if not os.path.exists(config_path):
            config_path = os.path.join(os.path.dirname(os.path.dirname(script_dir)), "IDswappers", "logging.ini")
        self.logger = THElogger(script_name="PAT_Username_Checker", config_file=config_path)

        # Map format to validator function
        self.validators = {
            'alphanumeric': validate_alphanumeric,
            'long_numeric': validate_long_numeric,
            'short_numeric': validate_short_numeric,
        }

        # Statistics
        self.stats = {
            'total_processed': 0,
            'expected_id': 0,
            'files_to_check': 0,
            'empty_or_unreadable': 0,
            'errors': 0,
        }

    def _ensure_output_structure(self) -> bool:
        """
        Create output subdirectories if they don't exist.

        Returns:
            True if successful, False if there's an error.
        """
        try:
            subdirs = ['Expected_ID', 'Files_to_check', 'Empty_or_unreadable']
            for subdir in subdirs:
                path = self.output_folder / subdir
                path.mkdir(parents=True, exist_ok=True)
            return True
        except Exception as e:
            self.logger.warning(f"Failed to create output structure: {e}")
            return False

    def _get_unique_path(self, dest_path: Path) -> Path:
        """
        If destination file exists, append _(1), _(2), etc.

        Args:
            dest_path: The desired destination path.

        Returns:
            A unique path (either the original or with _(N) appended).
        """
        if not dest_path.exists():
            return dest_path

        stem = dest_path.stem
        suffix = dest_path.suffix
        parent = dest_path.parent
        counter = 1

        while True:
            new_name = f"{stem}_{counter}{suffix}"
            new_path = parent / new_name
            if not new_path.exists():
                return new_path
            counter += 1

    def _validate_file(self, file_path: str) -> str:
        """
        Validate all usernames in a file against the selected format.

        Args:
            file_path: Path to the Excel file to validate.

        Returns:
            Status code: STATUS_EXPECTED_ID, STATUS_FILES_TO_CHECK, or STATUS_EMPTY_OR_UNREADABLE.
        """
        try:
            wb = load_workbook(file_path, data_only=True)
            try:
                ws = wb.active
                if ws is None:
                    return STATUS_EMPTY_OR_UNREADABLE

                # Auto-detect username column (returns (col_letter, header_row) or None)
                result = detect_username_column(file_path)
                if result is None:
                    self.logger.warning(f"Could not auto-detect username column in {os.path.basename(file_path)}")
                    return STATUS_EMPTY_OR_UNREADABLE

                username_col, header_row = result
                self.logger.debug(f"Found username column '{username_col}' at row {header_row} in {os.path.basename(file_path)}")

                # Find last row with data in username column
                last_row = ws_last_row(ws, username_col)
                if last_row is None or last_row <= header_row:
                    # No data rows (header only or empty)
                    return STATUS_EMPTY_OR_UNREADABLE

                # Get the validator for this format
                validator = self.validators.get(self.selected_format)
                if validator is None:
                    return STATUS_EMPTY_OR_UNREADABLE

                # Validate all values in the username column (data starts after header row)
                all_match = True
                col_num = column_index_from_string(username_col)

                for row in range(header_row + 1, last_row + 1):
                    cell = ws.cell(row=row, column=col_num)
                    if cell.value is None:
                        # Treat None as non-matching
                        all_match = False
                        break

                    if not validator(cell.value):
                        all_match = False
                        break

                return STATUS_EXPECTED_ID if all_match else STATUS_FILES_TO_CHECK

            finally:
                wb.close()

        except Exception as e:
            self.logger.warning(f"Exception while validating {os.path.basename(file_path)}: {e}")
            return STATUS_EMPTY_OR_UNREADABLE

    def run(self, files: List[str]) -> dict:
        """
        Process a list of files: validate and move to appropriate output folders.

        Args:
            files: List of file paths to process.

        Returns:
            Dictionary with statistics: total_processed, expected_id, files_to_check,
            empty_or_unreadable, errors.
        """
        # Ensure output structure exists
        if not self._ensure_output_structure():
            return self.stats

        self.logger.info(f"PAT_Username_Checker started - Format: {self.selected_format}")
        self.logger.info(f"Processing {len(files)} file(s) to {self.output_folder}")

        for file_path in files:
            self.stats['total_processed'] += 1
            filename = Path(file_path).name

            try:
                # Validate the file
                status = self._validate_file(file_path)

                # Determine destination folder
                if status == STATUS_EXPECTED_ID:
                    dest_folder = self.output_folder / 'Expected_ID'
                    self.stats['expected_id'] += 1
                elif status == STATUS_FILES_TO_CHECK:
                    dest_folder = self.output_folder / 'Files_to_check'
                    self.stats['files_to_check'] += 1
                else:  # STATUS_EMPTY_OR_UNREADABLE
                    dest_folder = self.output_folder / 'Empty_or_unreadable'
                    self.stats['empty_or_unreadable'] += 1

                # Move file to destination
                dest_path = dest_folder / filename
                dest_path = self._get_unique_path(dest_path)

                shutil.move(file_path, str(dest_path))
                self.logger.info(f"Processed {filename}: Status={status}, Destination={dest_folder}")

            except Exception as e:
                self.stats['errors'] += 1
                self.logger.warning(f"Failed to process {filename}: {e}")

        self.logger.info(f"Processing complete. {self.stats['expected_id']} expected, "
                         f"{self.stats['files_to_check']} to check, "
                         f"{self.stats['empty_or_unreadable']} empty/unreadable, "
                         f"{self.stats['errors']} errors.")
        self.logger.finalize_report()
        return self.stats


# === STANDALONE EXECUTION ===

def main():
    """
    Standalone entry point. Prompts user for files, format, output folder,
    then processes and prints summary.
    """
    # Create wx app (required for dialogs)
    app = wx.App(False)

    # Step 1: Select files
    files = select_work_files(['.xlsx', '.xls'])
    if files is None or len(files) == 0:
        print("No files selected. Exiting.")
        return

    # Step 2: Select format
    selected_format = select_format()
    if selected_format is None:
        print("Format selection cancelled. Exiting.")
        return

    # Step 3: Select output folder
    output_folder = select_output_folder("Select output folder for organized files")
    if output_folder is None:
        print("Output folder selection cancelled. Exiting.")
        return

    # Step 4: Create checker and run
    checker = PAT_Username_Checker(output_folder, selected_format)
    stats = checker.run(files)

    # Step 5: Print summary
    print("\n" + "="*60)
    print("USERNAME VALIDATION SUMMARY")
    print("="*60)
    print(f"Format Selected: {selected_format}")
    print(f"Output Folder: {output_folder}")
    print(f"Total Files Processed: {stats['total_processed']}")
    print(f"  - Expected ID: {stats['expected_id']}")
    print(f"  - Files to Check: {stats['files_to_check']}")
    print(f"  - Empty or Unreadable: {stats['empty_or_unreadable']}")
    print(f"  - Errors: {stats['errors']}")
    print("="*60)

    app.Destroy()


if __name__ == "__main__":
    main()
