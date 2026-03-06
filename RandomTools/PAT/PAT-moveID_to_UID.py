"""
PAT-moveID_to_UID.py - Copy expected ID values from Username column to Unique ID column.

Scans PATonline Excel files, finds rows where the Username column contains a value
matching the expected ID format, and copies those values into the Unique ID column.
Files with copied IDs are saved to [output]/IDsMOVED/, files with no matches are
moved to [output]/NO_moves/.

Can be run standalone or imported as a module.
"""

# === IMPORTS ===
import os
import sys
import shutil

# Add project root to path so Helpers can be imported
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))

import wx
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from Helpers.Clean_fields.clean_field import field_cleaner
from Helpers.Last_row_finder.real_last_row import ws_last_row
from Helpers.dog_box.work_files import select_work_files, select_output_folder
from water_logged.the_logger import THElogger

from PAT_Username_Checker import (
    SCAN_COLUMNS, SCAN_ROWS,
    select_format,
    validate_alphanumeric, validate_long_numeric, validate_short_numeric,
)


# === COLUMN DETECTION ===
# Header variants (readable format, normalized at match time)
USERNAME_HEADERS = {
    "Username": "username",
    "User ID": "userid",
    "UserID": "userid",
}

UNIQUEID_HEADERS = {
    "Unique ID": "uniqueid",
    "UniqueID": "uniqueid",
    "Unique_ID": "uniqueid",
}


def detect_column(file_path, header_map):
    """
    Detect a column by scanning rows 1-20, columns A-M for header variants.

    Uses field_cleaner for robust matching. Normalizes both sides of the comparison.

    Args:
        file_path: Path to the Excel file.
        header_map: Dict of readable header names to normalized keys.

    Returns:
        Tuple of (column_letter, header_row) if found, None if not detected.
    """
    normalized_search = {field_cleaner(k, strip_spaces=True): v for k, v in header_map.items()}

    try:
        wb = load_workbook(file_path, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return None

            for row in SCAN_ROWS:
                for col in SCAN_COLUMNS:
                    cell = ws[f'{col}{row}']
                    if cell.value is None:
                        continue
                    cell_text = field_cleaner(str(cell.value), strip_spaces=True)
                    if cell_text in normalized_search:
                        return (col, row)
            return None
        finally:
            wb.close()
    except Exception:
        return None


# === MAIN CLASS ===
class PATMoveIDtoUID:
    """Copies expected ID values from Username column to Unique ID column in PATonline files."""

    FORMAT_VALIDATORS = {
        'alphanumeric': validate_alphanumeric,
        'long_numeric': validate_long_numeric,
        'short_numeric': validate_short_numeric,
    }

    def __init__(self):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        config_path = os.path.join(script_dir, "logging.ini")
        if not os.path.exists(config_path):
            config_path = os.path.join(os.path.dirname(os.path.dirname(script_dir)), "IDswappers", "logging.ini")
        self.logger = THElogger(script_name="PAT-moveID_to_UID", config_file=config_path)

    def process_file(self, file_path, selected_format, output_folder):
        """
        Process a single file: copy matching IDs from Username to Unique ID column.

        Args:
            file_path: Path to the Excel file.
            selected_format: Format key ('alphanumeric', 'long_numeric', 'short_numeric').
            output_folder: Root output directory (Path).

        Returns:
            'moved' if IDs were copied, 'no_moves' if none matched, 'error' on failure.
        """
        filename = os.path.basename(file_path)
        validator = self.FORMAT_VALIDATORS[selected_format]

        # Detect both columns
        username_result = detect_column(file_path, USERNAME_HEADERS)
        if username_result is None:
            self.logger.warning(f"No Username column found in {filename}")
            return 'no_moves'

        uniqueid_result = detect_column(file_path, UNIQUEID_HEADERS)
        if uniqueid_result is None:
            self.logger.warning(f"No Unique ID column found in {filename}")
            return 'no_moves'

        username_col, username_hdr_row = username_result
        uniqueid_col, uniqueid_hdr_row = uniqueid_result

        # Data starts after the later header row
        data_start = max(username_hdr_row, uniqueid_hdr_row) + 1

        try:
            # Open for editing (not data_only so we can write back)
            wb = load_workbook(file_path)
            ws = wb.active

            last_row = ws_last_row(ws, username_col)
            if last_row is None or last_row < data_start:
                wb.close()
                self.logger.info(f"No data rows in {filename}")
                return 'no_moves'

            username_col_num = column_index_from_string(username_col)
            uniqueid_col_num = column_index_from_string(uniqueid_col)

            # Copy matching IDs from Username to Unique ID
            ids_copied = 0
            for row in range(data_start, last_row + 1):
                username_cell = ws.cell(row=row, column=username_col_num)
                if username_cell.value is None:
                    continue

                if validator(username_cell.value):
                    ws.cell(row=row, column=uniqueid_col_num).value = username_cell.value
                    ids_copied += 1

            if ids_copied > 0:
                # Save modified file to IDsMOVED
                dest_dir = output_folder / 'IDsMOVED'
                dest_dir.mkdir(parents=True, exist_ok=True)
                dest_path = self._get_unique_path(dest_dir / filename)
                wb.save(str(dest_path))
                wb.close()
                self.logger.info(f"{filename}: Copied {ids_copied} ID(s) to Unique ID -> IDsMOVED")
                return 'moved'
            else:
                wb.close()
                # Move original file to NO_moves
                dest_dir = output_folder / 'NO_moves'
                dest_dir.mkdir(parents=True, exist_ok=True)
                dest_path = self._get_unique_path(dest_dir / filename)
                shutil.move(file_path, str(dest_path))
                self.logger.info(f"{filename}: No matching IDs -> NO_moves")
                return 'no_moves'

        except Exception as e:
            self.logger.warning(f"Error processing {filename}: {e}")
            return 'error'

    def _get_unique_path(self, dest_path):
        """If destination exists, append _(1), _(2), etc."""
        dest_path = Path(dest_path)
        if not dest_path.exists():
            return dest_path
        stem = dest_path.stem
        suffix = dest_path.suffix
        parent = dest_path.parent
        counter = 1
        while True:
            new_path = parent / f"{stem}_{counter}{suffix}"
            if not new_path.exists():
                return new_path
            counter += 1

    def run(self, files, selected_format, output_folder, progress_callback=None):
        """
        Process files: copy matching IDs from Username to Unique ID, organize output.

        Args:
            files: List of file paths.
            selected_format: Format key ('alphanumeric', 'long_numeric', 'short_numeric').
            output_folder: Root output directory (str or Path).
            progress_callback: Optional callback(current_index, total_count, filename) -> bool.
                Returns True to continue, False to cancel.

        Returns:
            Dict with stats: total, ids_moved, no_moves, errors.
        """
        if progress_callback is not None and not callable(progress_callback):
            raise TypeError("progress_callback must be callable or None")

        output_folder = Path(output_folder)
        stats = {'total': 0, 'ids_moved': 0, 'no_moves': 0, 'errors': 0}

        self.logger.info(f"PAT-moveID_to_UID started - Format: {selected_format}")
        self.logger.info(f"Processing {len(files)} file(s) to {output_folder}")

        total_files = len(files)
        for index, file_path in enumerate(files, start=1):
            filename = os.path.basename(file_path)
            stats['total'] += 1

            # Progress callback
            if progress_callback is not None:
                try:
                    if not progress_callback(index, total_files, filename):
                        self.logger.info(f"Processing cancelled by user after {stats['total'] - 1} file(s)")
                        self.logger.finalize_report()
                        return stats
                except Exception as e:
                    self.logger.warning(f"Error in progress callback: {e}")

            result = self.process_file(file_path, selected_format, output_folder)
            if result == 'moved':
                stats['ids_moved'] += 1
            elif result == 'no_moves':
                stats['no_moves'] += 1
            else:
                stats['errors'] += 1

        self.logger.info(f"Complete. {stats['ids_moved']} with IDs copied, "
                         f"{stats['no_moves']} no matches, {stats['errors']} error(s).")
        self.logger.finalize_report()
        return stats


# === STANDALONE EXECUTION ===
def main():
    """Entry point for standalone execution."""
    app = wx.App(False)
    mover = PATMoveIDtoUID()

    # Select files
    files = select_work_files(['.xlsx', '.xls'])
    if not files:
        mover.logger.info("User cancelled file selection.")
        mover.logger.finalize_report()
        app.Destroy()
        return

    # Select expected ID format
    selected_format = select_format()
    if selected_format is None:
        mover.logger.info("User cancelled format selection.")
        mover.logger.finalize_report()
        app.Destroy()
        return

    # Select output folder
    output_folder = select_output_folder("Select output folder for processed files")
    if output_folder is None:
        mover.logger.info("User cancelled output folder selection.")
        mover.logger.finalize_report()
        app.Destroy()
        return

    stats = mover.run(files, selected_format, output_folder)

    print("\n" + "=" * 60)
    print("MOVE ID TO UID SUMMARY")
    print("=" * 60)
    print(f"Format Selected: {selected_format}")
    print(f"Total Files: {stats['total']}")
    print(f"  - IDs Copied: {stats['ids_moved']}")
    print(f"  - No Matches: {stats['no_moves']}")
    print(f"  - Errors: {stats['errors']}")
    print("=" * 60)

    app.Destroy()


if __name__ == "__main__":
    main()
