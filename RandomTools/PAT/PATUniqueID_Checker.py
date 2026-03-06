"""
PATUniqueID_Checker.py - Validate Unique ID column values in Excel files.

Auto-detects the Unique ID column in PATonline Excel files and validates all values
against a user-selected format (ABC0001 = 3 letters + 4 digits, 182815548 = 9 digits,
or 5548 = 4 digits). Files with invalid or empty Unique ID values are moved to an
Invalid_UniqueID/ subfolder; valid files remain in place.

Dual-mode: Can be run standalone with progress dialog or imported as a module with
optional progress_callback support.

Example module usage with progress callback:
    def my_progress_handler(idx, total, filename):
        print(f"Processing: {idx}/{total} - {filename}")
        return True  # continue processing

    checker = PATUniqueIDChecker()
    checker.run(format_type='ABC0001', progress_callback=my_progress_handler)
"""

# === IMPORTS ===
import os
import re
import shutil
import sys
from pathlib import Path
from typing import Optional, Callable, Tuple

# Add project root to path so Helpers can be imported
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))

import wx
from openpyxl import load_workbook
from xlrd import open_workbook

from Helpers.Clean_fields.clean_field import field_cleaner
from Helpers.dog_box.work_files import select_work_files, select_output_folder
from water_logged.the_logger import THElogger

# === SCAN RANGE (matches PATonline_FINDER convention) ===
SCAN_COLUMNS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
SCAN_ROWS = range(1, 21)

# === SUPPORTED FORMATS ===
FORMATS = {
    'ABC0001': {
        'pattern': r'^[A-Za-z]{3}[0-9]{4}$',
        'description': 'ABC0001 (3 letters + 4 digits)',
    },
    '182815548': {
        'pattern': r'^[0-9]{9}$',
        'description': '182815548 (9 digits)',
    },
    '5548': {
        'pattern': r'^[0-9]{4}$',
        'description': '5548 (4 digits)',
    },
}


# === PROGRESS DIALOG ===
class ProgressDialog(wx.Dialog):
    """Modal progress dialog with file counter, filename display, and cancel button."""

    def __init__(self, total_files, parent=None):
        """
        Initialize progress dialog.

        Args:
            total_files (int): Total number of files to process
            parent: Parent window (can be None for standalone mode)
        """
        super().__init__(parent, title="Processing Unique ID Validation", style=wx.DEFAULT_DIALOG_STYLE)
        self.total_files = total_files
        self._cancelled = False

        # Layout
        sizer = wx.BoxSizer(wx.VERTICAL)

        # Counter label
        self.counter_label = wx.StaticText(self, label="File 0 of 0")
        sizer.Add(self.counter_label, 0, wx.ALL | wx.EXPAND, 10)

        # Filename label
        self.filename_label = wx.StaticText(self, label="")
        sizer.Add(self.filename_label, 0, wx.ALL | wx.EXPAND, 10)

        # Progress bar
        self.progress_bar = wx.Gauge(self, range=max(1, total_files))
        sizer.Add(self.progress_bar, 0, wx.ALL | wx.EXPAND, 10)

        # Cancel button
        cancel_btn = wx.Button(self, wx.ID_CANCEL, "Cancel")
        cancel_btn.Bind(wx.EVT_BUTTON, self.on_cancel)
        sizer.Add(cancel_btn, 0, wx.ALL | wx.CENTER, 10)

        self.SetSizer(sizer)
        self.SetSize(400, 200)
        self.Centre()

    def update(self, current_index, filename):
        """
        Update progress display.

        Args:
            current_index (int): 1-based index of current file (1 to total_files)
            filename (str): Basename of file being processed
        """
        self.counter_label.SetLabel(f"File {current_index} of {self.total_files}")
        self.filename_label.SetLabel(f"Processing: {filename}")
        self.progress_bar.SetValue(current_index)
        self.Refresh()
        wx.SafeYield()

    def is_cancelled(self):
        """Check if user clicked cancel button."""
        return self._cancelled

    def on_cancel(self, event):
        """Handle cancel button click."""
        self._cancelled = True
        self.EndModal(wx.ID_CANCEL)


# === FORMAT SELECTION ===
def select_format() -> Optional[str]:
    """
    Display wxPython dialog to select Unique ID format.

    Returns:
        Format key ('ABC0001', '182815548', '5548'), or None if cancelled.
    """
    app = wx.GetApp()
    if app is None:
        app = wx.App(False)

    options = [
        "ABC0001 (3 letters + 4 digits)",
        "182815548 (9 digits)",
        "5548 (4 digits)",
    ]

    dlg = wx.SingleChoiceDialog(
        None,
        "Select the expected Unique ID format:",
        "Unique ID Format",
        options
    )
    dlg.CentreOnScreen()

    try:
        if dlg.ShowModal() != wx.ID_OK:
            return None
        choice = dlg.GetSelection()
        format_keys = ['ABC0001', '182815548', '5548']
        return format_keys[choice]
    finally:
        dlg.Destroy()


# === MAIN CLASS ===
class PATUniqueIDChecker:
    """Validates Unique ID column values in PATonline Excel files."""

    def __init__(self):
        """Initialize logger."""
        script_dir = os.path.dirname(os.path.abspath(__file__))
        # Look for logging.ini in parent directories
        config_path = os.path.join(script_dir, "logging.ini")
        if not os.path.exists(config_path):
            config_path = os.path.join(
                os.path.dirname(os.path.dirname(script_dir)), "IDswappers", "logging.ini"
            )
        if not os.path.exists(config_path):
            config_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(script_dir))),
                "IDswappers",
                "logging.ini",
            )

        self.logger = THElogger(script_name="PATUniqueID_Checker", config_file=config_path)

    @staticmethod
    def validate_unique_id_format(value: str, format_type: str) -> bool:
        """
        Validate a Unique ID value against the specified format.

        Args:
            value: The value to validate (string or convertible to string)
            format_type: Format key ('ABC0001', '182815548', '5548')

        Returns:
            True if value matches format, False otherwise
        """
        if value is None:
            return False

        # Strip whitespace using field_cleaner for consistency
        normalized = field_cleaner(str(value), strip_spaces=True)

        if not normalized:
            return False

        if format_type not in FORMATS:
            return False

        pattern = FORMATS[format_type]['pattern']
        return bool(re.match(pattern, normalized))

    @staticmethod
    def find_unique_id_column(worksheet) -> Optional[str]:
        """
        Scan worksheet cells A1-M20 for Unique ID header.

        Uses field_cleaner for robust text normalization to handle real-world data
        variations (unicode, extra whitespace, mixed case).

        Args:
            worksheet: openpyxl worksheet object

        Returns:
            Column letter if found, None otherwise
        """
        # Lookup dict: map normalized header to column marker
        search_headers = {"Unique ID": "unique_id"}
        normalized_search = {field_cleaner(k, strip_spaces=True): v for k, v in search_headers.items()}

        for row in SCAN_ROWS:
            for col in SCAN_COLUMNS:
                cell = worksheet[f'{col}{row}']
                if cell.value is None:
                    continue

                cell_text = field_cleaner(str(cell.value), strip_spaces=True)
                if cell_text in normalized_search:
                    return col

        return None

    def validate_file(self, file_path: str, format_type: str) -> Tuple[bool, str]:
        """
        Validate all Unique ID values in a file against the format.

        Args:
            file_path: Path to the Excel file
            format_type: Format key ('ABC0001', '182815548', '5548')

        Returns:
            Tuple of (is_valid, reason_message)
        """
        try:
            # Try openpyxl first for .xlsx files
            try:
                wb = load_workbook(file_path, data_only=True)
                try:
                    ws = wb.active
                    if ws is None:
                        return (False, "No active worksheet found")

                    # Find Unique ID column
                    unique_id_col = self.find_unique_id_column(ws)
                    if unique_id_col is None:
                        return (False, "Unique ID column not found")

                    # Find last row with data in this column
                    from Helpers.Last_row_finder.real_last_row import ws_last_row
                    last_row = ws_last_row(ws, unique_id_col)

                    if last_row is None:
                        return (False, "No data rows found in Unique ID column")

                    # Scan for header row (find the actual header)
                    header_row = None
                    for row in SCAN_ROWS:
                        cell = ws[f'{unique_id_col}{row}']
                        if cell.value is not None:
                            cell_text = field_cleaner(str(cell.value), strip_spaces=True)
                            if cell_text == field_cleaner("Unique ID", strip_spaces=True):
                                header_row = row
                                break

                    if header_row is None:
                        return (False, "Unique ID header not properly located")

                    # Validate all data rows
                    for row in range(header_row + 1, last_row + 1):
                        cell = ws[f'{unique_id_col}{row}']

                        # Empty cell is invalid
                        if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                            return (False, f"Empty Unique ID found at {unique_id_col}{row}")

                        # Validate format
                        if not self.validate_unique_id_format(cell.value, format_type):
                            return (
                                False,
                                f"Invalid format at {unique_id_col}{row}: {cell.value}",
                            )

                    return (True, "All Unique ID values match format")

                finally:
                    wb.close()

            except Exception as openpyxl_error:
                # Fallback to xlrd for .xls files
                try:
                    wb = open_workbook(file_path, on_demand=True)
                    ws = wb.sheet_by_index(0)

                    if ws is None:
                        return (False, "No active worksheet found")

                    # Find Unique ID column manually for xlrd
                    unique_id_col_idx = None
                    header_row = None
                    normalized_search = {field_cleaner("Unique ID", strip_spaces=True): True}

                    for row_idx in range(20):
                        for col_idx in range(13):  # A-M = 0-12
                            try:
                                cell_value = ws.cell_value(row_idx, col_idx)
                                if cell_value:
                                    cell_text = field_cleaner(str(cell_value), strip_spaces=True)
                                    if cell_text in normalized_search:
                                        # Convert col_idx to letter
                                        unique_id_col_idx = col_idx
                                        header_row = row_idx + 1  # xlrd is 0-indexed, convert to 1-indexed
                                        break
                            except (IndexError, TypeError):
                                pass
                        if unique_id_col_idx is not None:
                            break

                    if unique_id_col_idx is None:
                        return (False, "Unique ID column not found")

                    # Find last row with data
                    last_row = ws.nrows

                    if last_row is None or last_row <= header_row:
                        return (False, "No data rows found in Unique ID column")

                    # Validate all data rows
                    for row_idx in range(header_row, last_row):
                        try:
                            cell_value = ws.cell_value(row_idx, unique_id_col_idx)

                            # Empty cell is invalid
                            if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ""):
                                return (False, f"Empty Unique ID found at row {row_idx + 1}")

                            # Validate format
                            if not self.validate_unique_id_format(cell_value, format_type):
                                return (
                                    False,
                                    f"Invalid format at row {row_idx + 1}: {cell_value}",
                                )

                        except (IndexError, TypeError):
                            continue

                    return (True, "All Unique ID values match format")

                except Exception as xlrd_error:
                    return (False, f"Failed to read file: {xlrd_error}")

        except Exception as e:
            return (False, f"Validation error: {e}")

    def _handle_file_conflict(self, source_path: str, target_path: str) -> str:
        """
        Show dialog when target file already exists.

        Args:
            source_path: Path to the source file
            target_path: Path to the target file

        Returns:
            "overwrite", "skip", or "rename"
        """
        filename = os.path.basename(target_path)
        dlg = wx.SingleChoiceDialog(
            None,
            f"{filename} already exists in Invalid_UniqueID/.\n\nWhat do you want to do?",
            "File Exists",
            ["Overwrite", "Skip", "Rename"],
        )
        dlg.SetSelection(2)  # Default to Rename

        try:
            if dlg.ShowModal() != wx.ID_OK:
                return "skip"
            choice = dlg.GetSelection()
            return ["overwrite", "skip", "rename"][choice]
        finally:
            dlg.Destroy()

    def process_file(self, file_path: str, output_dir: str, format_type: str) -> bool:
        """
        Validate file and move to Invalid_UniqueID/ if invalid.

        Args:
            file_path: Path to the file
            output_dir: Root output directory
            format_type: Format to validate against

        Returns:
            True if processed successfully, False otherwise
        """
        is_valid, reason = self.validate_file(file_path, format_type)
        filename = os.path.basename(file_path)

        if is_valid:
            # Valid file stays in place
            self.logger.info(f"Valid: {filename} - {reason}")
            return True

        # Invalid file: move to Invalid_UniqueID/ subfolder
        invalid_dir = os.path.join(output_dir, "Invalid_UniqueID")
        os.makedirs(invalid_dir, exist_ok=True)

        target_path = os.path.join(invalid_dir, filename)

        # Handle existing file
        if os.path.exists(target_path):
            action = self._handle_file_conflict(file_path, target_path)

            if action == "overwrite":
                os.remove(target_path)
            elif action == "skip":
                self.logger.info(f"Skipped existing file: {target_path}")
                return False
            elif action == "rename":
                # Generate unique name with _dup suffix
                stem = Path(target_path).stem
                suffix = Path(target_path).suffix
                parent = os.path.dirname(target_path)
                counter = 1

                while True:
                    new_name = f"{stem}_dup{counter}{suffix}" if counter > 1 else f"{stem}_dup{suffix}"
                    target_path = os.path.join(parent, new_name)
                    if not os.path.exists(target_path):
                        break
                    counter += 1

        # Move file
        try:
            shutil.move(file_path, target_path)
            self.logger.info(f"Invalid: {filename} moved to Invalid_UniqueID/ - {reason}")
            return True
        except Exception as e:
            self.logger.warning(f"Failed to move {filename}: {e}")
            return False

    def run(self, format_type: str, progress_callback: Optional[Callable] = None) -> None:
        """
        Main workflow: select files, output folder, and process.

        Args:
            format_type: Format to validate against ('ABC0001', '182815548', '5548')
            progress_callback: Optional callback for progress updates.
                Signature: def callback(current_index, total_count, filename) -> bool
                    current_index: 1-based index of current file (1 to total_count)
                    total_count: Total number of files to process
                    filename: Basename of file being processed
                Returns: True to continue processing, False to cancel

        Raises:
            TypeError: If progress_callback is provided but not callable
        """
        self.logger.info("PATUniqueID_Checker started")

        # Validate callback parameter
        if progress_callback is not None and not callable(progress_callback):
            raise TypeError("progress_callback must be callable or None")

        # Prompt user for input files
        files = select_work_files([".xlsx", ".xls"])
        if not files:
            self.logger.info("User cancelled file selection.")
            self.logger.finalize_report()
            return

        # Prompt user for output folder
        output_dir = select_output_folder("Select output folder for validated Unique ID files")
        if not output_dir:
            self.logger.info("User cancelled output folder selection.")
            self.logger.finalize_report()
            return

        self.logger.info(f"Processing {len(files)} file(s) with format={format_type}")
        self.logger.info(f"Output folder: {output_dir}")

        # Process each file
        processed_count = 0
        moved_count = 0
        total_files = len(files)

        for index, file_path in enumerate(files, start=1):
            # Check for cancellation
            try:
                if progress_callback is not None:
                    filename = os.path.basename(file_path)
                    should_continue = progress_callback(index, total_files, filename)
                    if not should_continue:
                        self.logger.info(f"Processing cancelled by user after {processed_count} file(s) processed.")
                        self.logger.finalize_report()
                        return
            except Exception as e:
                self.logger.warning(f"Error in progress callback: {e}")
                # Continue processing despite callback error

            is_valid, reason = self.validate_file(file_path, format_type)
            if not is_valid:
                if self.process_file(file_path, output_dir, format_type):
                    moved_count += 1
                    processed_count += 1
            else:
                processed_count += 1

        self.logger.info(
            f"Processing complete. {processed_count} file(s) processed, "
            f"{moved_count} moved to Invalid_UniqueID/, "
            f"{processed_count - moved_count} valid files remain in place."
        )
        self.logger.finalize_report()


# === STANDALONE EXECUTION ===
def main():
    """Entry point for standalone execution."""
    print(r"""
================================================================================
  ____  ___  ______          _    _     _     ____  _______ __________
 / __ \/   |/_  __/         | |  | |   | |   / __ \/  ____//_  __/ ___/
/ /_/ / /| | / / _____ ___  | |  | |   | |  / /  \/ /____   / /  \__ \
\____/ ___ |/ / / ___/ / _ \ | |  | |   | | / /    / ____/  / /  /   / /
     / /  |/ / / /__  /  __/ \ \__/ \__/ / / /____/ /_____  / /  /   /
    /_/  |_/_/\___/   \___/   \___/\____/  \_____/_____/  /_/  /   /

================================================================================
""")

    app = wx.App(False)

    # Prompt for format selection
    format_type = select_format()
    if format_type is None:
        print("Format selection cancelled. Exiting.")
        app.Destroy()
        return

    # Create checker and run
    checker = PATUniqueIDChecker()
    checker.run(format_type, progress_callback=_create_progress_callback())

    app.Destroy()


def _create_progress_callback():
    """
    Create a progress callback that manages a wxPython progress dialog.

    Returns:
        Callable that creates and manages the progress dialog
    """
    progress_dialog = None

    def progress_handler(current_index, total_count, filename):
        """Handle progress updates and manage dialog lifecycle."""
        nonlocal progress_dialog

        # Create dialog on first call
        if progress_dialog is None:
            progress_dialog = ProgressDialog(total_count, parent=None)
            progress_dialog.Show()

        # Update dialog
        progress_dialog.update(current_index, filename)

        # Check for cancellation
        if progress_dialog.is_cancelled():
            # Clean up dialog
            if progress_dialog:
                progress_dialog.Destroy()
                progress_dialog = None
            return False  # Signal to cancel processing

        return True  # Continue processing

    return progress_handler


if __name__ == "__main__":
    main()
