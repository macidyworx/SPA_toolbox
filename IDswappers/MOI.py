"""
MOI.py - Swaps student IDs in MOI (Mathematics Online Interview) Excel files based on SIF or SSOT lookup.
"""

# === IMPORTS ===
import os
import sys

# Add project root to path so Helpers can be imported
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import shutil
import wx
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string
from xlrd import open_workbook
from xlutils.copy import copy
from Helpers.Clean_fields.clean_field import field_cleaner
from Helpers.dog_box import select_single_file, select_work_files, select_output_folder
from water_logged.the_logger import THElogger


# === CONSTANTS ===
FILE_STUDENT_HEADER = "Student"  # Stored as "Surname, Firstname"
FILE_ID_HEADER = "ID"
FILE_DATE_HEADER = "Date"
SIF_SURNAME = "Surname"
SIF_FIRSTNAME = "Firstname"
SIF_STUDENTID = "StudentID"


# === MAIN CLASS ===
class MOISwapper:
    """
    MOI (Mathematics Online Interview Insight Platform) Swapper.

    Processes Excel files (.xls and .xlsx) to swap student IDs based on a
    Student Information File (SIF).
    """

    def __init__(self, logger):
        """Initialize MOI Swapper with logger."""
        self.logger = logger
        self.mode = None
        self.sif_lookup = None
        self.ssot_lookup = {}
        self.moi_folder = None
        self.skipped_folder = None
        self.not_found = []
        self.total_checked = 0
        self.total_matched = 0
        self.files_checked = []
        self.files_skipped = []

    def run(self):
        """Execute the MOI swapping process."""
        try:
            # Step 1: Get user inputs using dog_box helpers
            lookup_result = select_single_file(mode="choose")
            if not lookup_result:
                self.logger.info("User cancelled file selection.")
                self.logger.finalize_report()
                return

            if isinstance(lookup_result, str):
                self.mode = "sif"
                sif_path = lookup_result
            else:
                self.mode = "ssot"
                ssot_info = lookup_result

            files = select_work_files([".xlsx", ".xls", ".xlsm"])
            if not files:
                self.logger.info("No files selected.")
                self.logger.finalize_report()
                return

            # Filter out temporary Excel files
            files = [f for f in files if not os.path.basename(f).startswith('~$')]

            output_dir = select_output_folder("Select output folder for MOI")
            if output_dir is None:
                self.logger.info("User cancelled output folder selection.")
                self.logger.finalize_report()
                return

            lookup_path = sif_path if self.mode == "sif" else ssot_info['path']
            self.logger.info(f"Using values: Mode={self.mode.upper()}, Lookup={lookup_path}, Files={len(files)}, Output={output_dir}")

            # Step 2: Setup output folder
            output_subfolder = os.path.join(output_dir, "MOIswapped")
            if os.path.exists(output_subfolder):
                result = wx.MessageBox(
                    f"{output_subfolder} already exists.\nRemove it and continue?",
                    "Output Folder Exists", wx.YES_NO | wx.ICON_WARNING)
                if result == wx.YES:
                    shutil.rmtree(output_subfolder)
                else:
                    self.logger.info("User cancelled due to existing output folder.")
                    self.logger.finalize_report()
                    return

            self.moi_folder = output_subfolder
            os.mkdir(self.moi_folder)

            self.skipped_folder = os.path.join(self.moi_folder, "SKIPPED")
            os.makedirs(self.skipped_folder, exist_ok=True)

            # Step 3: Load lookup (SIF or SSOT)
            if self.mode == "sif":
                sif_wb = load_workbook(sif_path, read_only=True, data_only=True)
                sif_ws = sif_wb.active
                self.sif_lookup = {}
                for row in sif_ws.iter_rows(min_row=3, values_only=True):
                    if row[3] and row[2] and row[4]:  # Firstname, Surname, StudentID
                        fname = field_cleaner(str(row[3]))
                        lname = field_cleaner(str(row[2]))
                        key = (fname, lname)
                        self.sif_lookup[key] = row[4]
                sif_wb.close()
            else:
                ssot_wb = load_workbook(ssot_info['path'], read_only=True, data_only=True)
                ssot_ws = ssot_wb.active
                hr = ssot_info['header_row']
                old_col = column_index_from_string(ssot_info['old_id_col'])
                new_col = column_index_from_string(ssot_info['new_id_col'])
                for row in ssot_ws.iter_rows(min_row=hr + 1):
                    old_val = row[old_col - 1].value
                    new_val = row[new_col - 1].value
                    if old_val and new_val:
                        self.ssot_lookup[field_cleaner(str(old_val))] = new_val
                ssot_wb.close()
                self.logger.info(f"Loaded {len(self.ssot_lookup)} ID mappings from SSOT")

            self.logger.info(f"Total files to process: {len(files)}")

            # Step 4-7: Process each file
            for file_count, file in enumerate(files, 1):
                self.logger.info(f"Processing file {file_count}/{len(files)} > {file}")
                self._process_file(file)

            # Step 8: Save report
            self._save_report()

            self.logger.info(f"Total Students Checked --> {self.total_checked}")
            self.logger.info(f"Total Students Matched --> {self.total_matched}")
            self.logger.info(f"Total NOT Found --> {len(self.not_found)}")
            self.logger.info(f"Processing complete. Files saved in {self.moi_folder} folder.")

        except Exception as e:
            self.logger.error(f"Error during MOI processing: {e}")
        finally:
            self.logger.finalize_report()

    def _process_file(self, file):
        """Process a single file (xlsx/xlsm or xls)."""
        # Per-file counters
        file_checked = 0
        file_matched = 0
        file_not_found = 0

        if '.xlsx' in file.lower() or '.xlsm' in file.lower():
            self._process_xlsx(file, file_checked, file_matched, file_not_found)
        elif '.xls' in file.lower():
            self._process_xls(file, file_checked, file_matched, file_not_found)
        else:
            self.logger.info(f"Unsupported file format: {file}. Skipping.")
            self.files_skipped.append(os.path.basename(file))
            shutil.copy(file, os.path.join(self.skipped_folder, os.path.basename(file)))

    def _process_xlsx(self, file, file_checked, file_matched, file_not_found):
        """Process .xlsx/.xlsm file."""
        wb = load_workbook(file)
        ws = wb.active  # Assume first sheet

        # Find "Student" cell using field_cleaner for robust matching
        student_col = None
        header_row = None
        student_normalized = field_cleaner(FILE_STUDENT_HEADER, strip_spaces=True)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_normalized = field_cleaner(str(cell.value), strip_spaces=True)
                if cell_normalized == student_normalized:
                    student_col = cell.column_letter
                    header_row = cell.row
                    break
            if student_col:
                break

        if not student_col:
            self.logger.info(f"Warning: '{FILE_STUDENT_HEADER}' header not found in {file}. Skipping.")
            self.files_skipped.append(os.path.basename(file))
            shutil.copy(file, os.path.join(self.skipped_folder, os.path.basename(file)))
            return

        # Find "ID" and "Date" in the header row using field_cleaner
        id_col = None
        date_col = None
        id_normalized = field_cleaner(FILE_ID_HEADER, strip_spaces=True)
        date_normalized = field_cleaner(FILE_DATE_HEADER, strip_spaces=True)
        for cell in ws[header_row]:
            if cell.value is None:
                continue
            cell_normalized = field_cleaner(str(cell.value), strip_spaces=True)
            if cell_normalized == id_normalized and id_col is None:
                id_col = cell.column_letter
            elif cell_normalized == date_normalized and date_col is None:
                date_col = cell.column_letter

        if not id_col:
            self.logger.info(f"Warning: '{FILE_ID_HEADER}' header not found in {file}. Skipping.")
            self.files_skipped.append(os.path.basename(file))
            shutil.copy(file, os.path.join(self.skipped_folder, os.path.basename(file)))
            return

        self.files_checked.append(os.path.basename(file))

        # Process each student row
        for row in range(header_row + 1, ws.max_row + 1):
            id_cell = ws[f"{id_col}{row}"]

            if self.mode == "sif":
                name_cell = ws[f"{student_col}{row}"]
                name = name_cell.value
                if name and isinstance(name, str):
                    self.total_checked += 1
                    file_checked += 1
                    # Parse name: "Brown, Florence" -> Lname=Brown, Fname=Florence
                    if ", " in name:
                        lname, fname = name.split(", ", 1)
                        fname = field_cleaner(fname)
                        lname = field_cleaner(lname)
                    else:
                        # If not in expected format, skip or handle
                        self.logger.info(f"Warning: Name format not recognized in {file} row {row}: {name}")
                        continue

                    # Get date for year
                    date_value = ws[f"{date_col}{row}"].value if date_col else None
                    year = None
                    if date_value:
                        date_str = str(date_value)
                        if '/' in date_str:
                            year = date_str.split('/')[-1].split()[0]
                        elif '-' in date_str:
                            parts = date_str.split('-')
                            if len(parts) >= 3:
                                if len(parts[0]) == 4:  # yyyy-mm-dd format
                                    year = parts[0]
                                else:  # dd-mm-yyyy format
                                    year = parts[2].split()[0]

                    # Find in SIF
                    new_id = self.sif_lookup.get((fname, lname))
                    if new_id is not None:
                        id_cell.value = new_id
                        self.total_matched += 1
                        file_matched += 1
                    else:
                        self.not_found.append({'File': file, 'Row': row, 'Name': name, 'Fname': fname, 'Lname': lname, 'Year': year})
                        file_not_found += 1
                        self.logger.debug(f"NOT FOUND in SIF: {fname} {lname}")
            else:  # ssot
                old_id = id_cell.value
                if old_id:
                    self.total_checked += 1
                    file_checked += 1
                    cleaned_id = field_cleaner(str(old_id))
                    new_id = self.ssot_lookup.get(cleaned_id)
                    if new_id is not None:
                        id_cell.value = new_id
                        self.total_matched += 1
                        file_matched += 1
                    else:
                        self.not_found.append({'File': file, 'Row': row, 'Old ID': str(old_id)})
                        file_not_found += 1
                        self.logger.debug(f"NOT FOUND in SSOT: {old_id}")

        # Save to MOIswapped
        output_path = os.path.join(self.moi_folder, os.path.basename(file))
        try:
            wb.save(output_path)
        except Exception as e:
            self.logger.error(f"Error saving {output_path}: {e}")
            try:
                wb.save(output_path)
            except Exception as e2:
                self.logger.error(f"Failed again: {e2}. Skipping save for {file}.")

        self.logger.info(f"Students Checked: {file_checked}")
        self.logger.info(f"Students Matched: {file_matched}")
        self.logger.info(f"Students NOT Found: {file_not_found}")

    def _process_xls(self, file, file_checked, file_matched, file_not_found):
        """Process .xls file."""
        rb = open_workbook(file, formatting_info=True)
        wb = copy(rb)
        ws = wb.get_sheet(0)

        # Find header row and columns (0-based indices)
        header_row = None
        student_col = None
        id_col = None
        date_col = None
        sheet = rb.sheet_by_index(0)
        # Normalize header lookup keys using field_cleaner
        student_normalized = field_cleaner(FILE_STUDENT_HEADER, strip_spaces=True)
        id_normalized = field_cleaner(FILE_ID_HEADER, strip_spaces=True)
        date_normalized = field_cleaner(FILE_DATE_HEADER, strip_spaces=True)
        for row_idx in range(sheet.nrows):
            row = sheet.row(row_idx)
            for col_idx, cell in enumerate(row):
                if cell.value is None:
                    continue
                cell_normalized = field_cleaner(str(cell.value), strip_spaces=True)
                if cell_normalized == student_normalized and student_col is None:
                    student_col = col_idx
                    header_row = row_idx
                elif cell_normalized == id_normalized and id_col is None:
                    id_col = col_idx
                elif cell_normalized == date_normalized and date_col is None:
                    date_col = col_idx
            if student_col is not None and id_col is not None:
                break

        if student_col is None or id_col is None:
            self.logger.info(f"Warning: Required headers not found in {file}. Skipping.")
            self.files_skipped.append(os.path.basename(file))
            shutil.copy(file, os.path.join(self.skipped_folder, os.path.basename(file)))
            return

        self.files_checked.append(os.path.basename(file))

        # Process each student row
        for row_idx in range(header_row + 1, sheet.nrows):
            if self.mode == "sif":
                name = sheet.cell_value(row_idx, student_col)
                if name:
                    name = str(name).strip()
                    self.total_checked += 1
                    file_checked += 1
                    # Parse name: "Brown, Florence" -> Lname=Brown, Fname=Florence
                    if ", " in name:
                        lname, fname = name.split(", ", 1)
                        fname = field_cleaner(fname)
                        lname = field_cleaner(lname)
                    else:
                        # If not in expected format, skip or handle
                        self.logger.info(f"Warning: Name format not recognized in {file} row {row_idx + 1}: {name}")
                        continue

                    # Get date for year
                    date_value = sheet.cell_value(row_idx, date_col) if date_col is not None else None
                    year = None
                    if date_value:
                        if isinstance(date_value, float):
                            # xlrd returns dates as floats
                            import xlrd
                            try:
                                date_tuple = xlrd.xldate_as_tuple(date_value, rb.datemode)
                                year = str(date_tuple[0])  # Year is first element
                            except:
                                year = None
                        else:
                            date_str = str(date_value)
                            if '/' in date_str:
                                year = date_str.split('/')[-1].split()[0]
                            elif '-' in date_str:
                                parts = date_str.split('-')
                                if len(parts) >= 3:
                                    if len(parts[0]) == 4:  # yyyy-mm-dd format
                                        year = parts[0]
                                    else:  # dd-mm-yyyy format
                                        year = parts[2].split()[0]

                    # Find in SIF
                    new_id = self.sif_lookup.get((fname, lname))
                    if new_id is not None:
                        ws.write(row_idx, id_col, new_id)
                        self.total_matched += 1
                        file_matched += 1
                    else:
                        self.not_found.append({'File': file, 'Row': row_idx + 1, 'Name': name, 'Fname': fname, 'Lname': lname, 'Year': year})
                        file_not_found += 1
                        self.logger.debug(f"NOT FOUND in SIF: {fname} {lname}")
            else:  # ssot
                old_id = sheet.cell_value(row_idx, id_col)
                if old_id:
                    self.total_checked += 1
                    file_checked += 1
                    cleaned_id = field_cleaner(str(old_id))
                    new_id = self.ssot_lookup.get(cleaned_id)
                    if new_id is not None:
                        ws.write(row_idx, id_col, new_id)
                        self.total_matched += 1
                        file_matched += 1
                    else:
                        self.not_found.append({'File': file, 'Row': row_idx + 1, 'Old ID': str(old_id)})
                        file_not_found += 1
                        self.logger.debug(f"NOT FOUND in SSOT: {old_id}")

        # Save to MOIswapped
        output_path = os.path.join(self.moi_folder, os.path.basename(file))
        try:
            wb.save(output_path)
        except Exception as e:
            self.logger.error(f"Error saving {output_path}: {e}")
            try:
                wb.save(output_path)
            except Exception as e2:
                self.logger.error(f"Failed again: {e2}. Skipping save for {file}.")

        self.logger.info(f"Students Checked: {file_checked}")
        self.logger.info(f"Students Matched: {file_matched}")
        self.logger.info(f"Students NOT Found: {file_not_found}")

    def _save_report(self):
        """Save processing report to Excel."""
        if not self.not_found and not self.files_checked and not self.files_skipped:
            return

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Summary sheet
        summary_ws = wb.create_sheet("Summary")
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(bold=True)

        # Headers
        summary_ws.cell(row=1, column=1).value = "Metric"
        summary_ws.cell(row=1, column=2).value = "Value"
        summary_ws.cell(row=1, column=1).fill = header_fill
        summary_ws.cell(row=1, column=1).font = header_font
        summary_ws.cell(row=1, column=2).fill = header_fill
        summary_ws.cell(row=1, column=2).font = header_font

        # Summary data
        summary_data = [
            ('Total Files Processed', len(self.files_checked)),
            ('Total Matched', self.total_matched),
            ('Total NOT Matched', len(self.not_found)),
            ('Note', 'Numbers will be exaggerated, because students may be checked multiple times if they are in multiple files.'),
        ]
        for row_idx, (metric, value) in enumerate(summary_data, start=2):
            summary_ws.cell(row=row_idx, column=1).value = metric
            summary_ws.cell(row=row_idx, column=2).value = value

        # Files Checked
        summary_ws.cell(row=7, column=1).value = "Files Checked"
        summary_ws.cell(row=7, column=1).font = header_font
        for i, file in enumerate(self.files_checked, start=8):
            summary_ws.cell(row=i, column=1).value = file

        # Files Skipped
        summary_ws.cell(row=7, column=2).value = "Files Skipped"
        summary_ws.cell(row=7, column=2).font = header_font
        for i, file in enumerate(self.files_skipped, start=8):
            summary_ws.cell(row=i, column=2).value = file

        # Full List sheet
        full_list_ws = wb.create_sheet("Full List")
        if self.not_found:
            # Headers
            headers = ['File', 'Row', 'Name', 'Fname', 'Lname', 'Year']
            for col_idx, header in enumerate(headers, start=1):
                cell = full_list_ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font

            # Data rows
            for row_idx, entry in enumerate(self.not_found, start=2):
                full_list_ws.cell(row=row_idx, column=1).value = entry.get('File', '')
                full_list_ws.cell(row=row_idx, column=2).value = entry.get('Row', '')
                full_list_ws.cell(row=row_idx, column=3).value = entry.get('Name', '')
                full_list_ws.cell(row=row_idx, column=4).value = entry.get('Fname', '')
                full_list_ws.cell(row=row_idx, column=5).value = entry.get('Lname', '')
                full_list_ws.cell(row=row_idx, column=6).value = entry.get('Year', '')

        # Save workbook
        report_path = os.path.join(self.moi_folder, "MOI_report.xlsx")
        wb.save(report_path)


# === STANDALONE EXECUTION ===
def main():
    """Main entry point for MOI Swapper."""
    print(r"""
===================================================================================================
___  ________ _____            ___________
|  \/  |  _  |_   _|          |_   _|  _  \
| .  . | | | | | |    ______    | | | | | |_____      ____ _ _ __  _ __   ___ _ __
| |\/| | | | | | |   |______|   | | | | / __\ \ /\ / / _` | '_ \| '_ \ / _ \ '__|
| |  | \ \_/ /_| |_            _| |_| |/ /\__ \\ V  V / (_| | |_) | |_) |  __/ |
\_|  |_/\___/ \___/            \___/|___/ |___/ \_/\_/ \__,_| .__/| .__/ \___|_|
                                                            | |   | |
                                                            |_|   |_|
===================================================================================================
""")

    # Initialize logger
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, "logging.ini")
    logger = THElogger(script_name="MOI", config_file=config_path)

    # Run MOI Swapper
    swapper = MOISwapper(logger)
    swapper.run()


if __name__ == "__main__":
    main()
