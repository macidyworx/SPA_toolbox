import os
import sys

# Add project root to path so Helpers can be imported
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import shutil
import wx
from openpyxl import load_workbook, Workbook
from xlrd import open_workbook
from xlutils.copy import copy
from Helpers.Clean_fields.clean_field import field_cleaner
from Helpers.dog_box import select_single_file, select_work_files, select_output_folder
from openpyxl.utils import column_index_from_string
from water_logged.the_logger import THElogger

# Global constants for headers
FILE_FNAME = "First Name"
FILE_LNAME = "Last Name"
FILE_ID_HEADER = "ID"
FILE_DATE_HEADER = "Result Date"
SIF_SURNAME = "Surname"
SIF_FIRSTNAME = "Firstname"
SIF_STUDENTID = "StudentID"


class EOISwapper:
    """Processes EOI Template Excel files and swaps student IDs based on SIF."""

    def __init__(self, logger):
        self.logger = logger
        self.not_found = []
        self.total_checked = 0
        self.total_matched = 0
        self.files_checked = []
        self.files_skipped = []
        self.sif_lookup = {}
        self.mode = None
        self.ssot_lookup = {}

    def run(self):
        """Main execution method."""
        try:
            # Get lookup file (SIF or SSOT)
            lookup_result = select_single_file(mode="choose")
            if not lookup_result:
                self.logger.info("User cancelled file selection.")
                return

            if isinstance(lookup_result, str):
                self.mode = "sif"
                sif_path = lookup_result
            else:
                self.mode = "ssot"
                ssot_info = lookup_result

            # Get working files
            work_files = select_work_files([".xlsx", ".xls"])
            if not work_files:
                self.logger.info("No work files selected. Exiting.")
                return

            # Filter out temp files
            work_files = [f for f in work_files if not os.path.basename(f).startswith('~$')]

            if not work_files:
                self.logger.info("No valid work files after filtering. Exiting.")
                return

            # Get output directory
            output_dir = select_output_folder("Select output folder for EOI")
            if output_dir is None:
                self.logger.info("No output folder selected. Exiting.")
                return

            # Load lookup data
            if self.mode == "sif":
                sif_wb = load_workbook(sif_path, read_only=True, data_only=True)
                sif_ws = sif_wb.active
                self.sif_lookup = {}
                for row in sif_ws.iter_rows(min_row=3, values_only=True):
                    if row[3] and row[2] and row[4]:  # Firstname, Surname, StudentID
                        key = (field_cleaner(str(row[3])), field_cleaner(str(row[2])))
                        self.sif_lookup[key] = row[4]
                sif_wb.close()
                self.logger.info(f"Loaded SIF file: {sif_path}")
            else:
                ssot_wb = load_workbook(ssot_info['path'], read_only=True, data_only=True)
                ssot_ws = ssot_wb.active
                hr = ssot_info['header_row']
                old_col = column_index_from_string(ssot_info['old_id_col'])
                new_col = column_index_from_string(ssot_info['new_id_col'])
                for row in ssot_ws.iter_rows(min_row=hr + 1):
                    old_val = row[old_col - 1].value
                    new_val = row[new_col - 1].value
                    if old_val and new_val:
                        self.ssot_lookup[field_cleaner(str(old_val))] = new_val
                ssot_wb.close()
                self.logger.info(f"Loaded {len(self.ssot_lookup)} ID mappings from SSOT")

            # Setup output folder
            output_subfolder = os.path.join(output_dir, "EOIswapped")
            if os.path.exists(output_subfolder):
                result = wx.MessageBox(
                    f"{output_subfolder} already exists.\nRemove it and continue?",
                    "Output Folder Exists", wx.YES_NO | wx.ICON_WARNING)
                if result == wx.YES:
                    shutil.rmtree(output_subfolder)
                    self.logger.info(f"Removed existing folder: {output_subfolder}")
                else:
                    self.logger.info("User cancelled due to existing output folder.")
                    return

            os.mkdir(output_subfolder)
            skipped_folder = os.path.join(output_subfolder, "SKIPPED")
            os.makedirs(skipped_folder, exist_ok=True)
            self.logger.info(f"Created output folder: {output_subfolder}")

            self.logger.info(f"Total files to process: {len(work_files)}")

            # Process each file
            file_count = 0
            for file in work_files:
                file_count += 1
                self.logger.info(f"Processing file {file_count}/{len(work_files)}: {file}")

                file_checked = 0
                file_matched = 0
                file_not_found = 0

                if '.xlsx' in file.lower():
                    file_checked, file_matched, file_not_found = self._process_xlsx(
                        file, output_subfolder, skipped_folder)
                elif '.xls' in file.lower():
                    file_checked, file_matched, file_not_found = self._process_xls(
                        file, output_subfolder, skipped_folder)
                else:
                    self.logger.info(f"Unsupported file format: {file}. Skipping.")
                    self.files_skipped.append(os.path.basename(file))
                    shutil.copy(file, os.path.join(skipped_folder, os.path.basename(file)))

                self.logger.info(f"Students Checked: {file_checked}")
                self.logger.info(f"Students Matched: {file_matched}")
                self.logger.info(f"Students NOT Found: {file_not_found}")

            # Generate report
            self._generate_report(output_subfolder)

            self.logger.info(f"Total Students Checked: {self.total_checked}")
            self.logger.info(f"Total Students Matched: {self.total_matched}")
            self.logger.info(f"Total NOT Found: {len(self.not_found)}")
            self.logger.info(f"Processing complete. Files saved in {output_subfolder} folder.")

        except Exception as e:
            self.logger.error(f"Error during processing: {e}")
            raise

    def _process_xlsx(self, file, output_subfolder, skipped_folder):
        """Process .xlsx file."""
        file_checked = 0
        file_matched = 0
        file_not_found = 0

        try:
            wb = load_workbook(file)
            ws = wb.active  # Assume first sheet

            # Find header row and columns
            header_row = None
            fname_col = None
            lname_col = None
            id_col = None
            date_col = None
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == FILE_FNAME:
                        fname_col = cell.column_letter
                        header_row = cell.row
                    elif cell.value == FILE_LNAME:
                        lname_col = cell.column_letter
                    elif cell.value == FILE_ID_HEADER:
                        id_col = cell.column_letter
                    elif cell.value == FILE_DATE_HEADER:
                        date_col = cell.column_letter
                if fname_col and lname_col and id_col:
                    break

            if not fname_col or not lname_col or not id_col:
                self.logger.info(f"Warning: Required headers not found in {file}. Skipping.")
                self.files_skipped.append(os.path.basename(file))
                shutil.copy(file, os.path.join(skipped_folder, os.path.basename(file)))
                return file_checked, file_matched, file_not_found

            self.files_checked.append(os.path.basename(file))

            # Process each student row
            for row in range(header_row + 1, ws.max_row + 1):
                fname_cell = ws[f"{fname_col}{row}"]
                lname_cell = ws[f"{lname_col}{row}"]
                id_cell = ws[f"{id_col}{row}"]
                fname = fname_cell.value
                lname = lname_cell.value
                if self.mode == "sif":
                    if fname and lname and isinstance(fname, str) and isinstance(lname, str):
                        fname = field_cleaner(fname)
                        lname = field_cleaner(lname)
                        self.total_checked += 1
                        file_checked += 1
                        # Find in SIF
                        new_id = self.sif_lookup.get((fname, lname))
                        if new_id is not None:
                            id_cell.value = new_id
                            self.total_matched += 1
                            file_matched += 1
                        else:
                            date_value = ws[f"{date_col}{row}"].value if date_col else None
                            year = None
                            if date_value:
                                date_str = str(date_value)
                                if '/' in date_str:
                                    year = date_str.split('/')[-1].split()[0]
                                elif '-' in date_str:
                                    parts = date_str.split('-')
                                    if len(parts) >= 3:
                                        if len(parts[0]) == 4:  # yyyy-mm-dd format
                                            year = parts[0]
                                        else:  # dd-mm-yyyy format
                                            year = parts[2].split()[0]
                            self.not_found.append({'File': file, 'Row': row, 'Fname': fname, 'Lname': lname, 'Year': year})
                            file_not_found += 1
                            self.logger.debug(f"NOT FOUND in SIF: {fname} {lname}")
                else:  # ssot
                    old_id = id_cell.value
                    if old_id:
                        self.total_checked += 1
                        file_checked += 1
                        cleaned_id = field_cleaner(str(old_id))
                        new_id = self.ssot_lookup.get(cleaned_id)
                        if new_id is not None:
                            id_cell.value = new_id
                            self.total_matched += 1
                            file_matched += 1
                        else:
                            self.not_found.append({'File': file, 'Row': row, 'Old ID': str(old_id)})
                            file_not_found += 1
                            self.logger.debug(f"NOT FOUND in SSOT: {old_id}")

            # Save to EOIswapped
            output_path = os.path.join(output_subfolder, os.path.basename(file))
            try:
                wb.save(output_path)
            except Exception as e:
                self.logger.error(f"Error saving {output_path}: {e}")
                input("Please close the file in Excel and press Enter to retry.")
                try:
                    wb.save(output_path)
                except Exception as e2:
                    self.logger.error(f"Failed again: {e2}. Skipping save for {file}.")

        except Exception as e:
            self.logger.error(f"Error processing .xlsx file {file}: {e}")
            self.files_skipped.append(os.path.basename(file))
            shutil.copy(file, os.path.join(skipped_folder, os.path.basename(file)))

        return file_checked, file_matched, file_not_found

    def _process_xls(self, file, output_subfolder, skipped_folder):
        """Process .xls file."""
        file_checked = 0
        file_matched = 0
        file_not_found = 0

        try:
            rb = open_workbook(file, formatting_info=True)
            wb = copy(rb)
            ws = wb.get_sheet(0)

            # Find header row and columns (0-based indices)
            header_row = None
            fname_col = None
            lname_col = None
            id_col = None
            date_col = None
            sheet = rb.sheet_by_index(0)
            for row_idx in range(sheet.nrows):
                row = sheet.row(row_idx)
                for col_idx, cell in enumerate(row):
                    if cell.value == FILE_FNAME:
                        fname_col = col_idx
                        header_row = row_idx
                    elif cell.value == FILE_LNAME:
                        lname_col = col_idx
                    elif cell.value == FILE_ID_HEADER:
                        id_col = col_idx
                    elif cell.value == FILE_DATE_HEADER:
                        date_col = col_idx
                if fname_col is not None and lname_col is not None and id_col is not None:
                    break

            if fname_col is None or lname_col is None or id_col is None:
                self.logger.info(f"Warning: Required headers not found in {file}. Skipping.")
                self.files_skipped.append(os.path.basename(file))
                shutil.copy(file, os.path.join(skipped_folder, os.path.basename(file)))
                return file_checked, file_matched, file_not_found

            self.files_checked.append(os.path.basename(file))

            # Process each student row
            for row_idx in range(header_row + 1, sheet.nrows):
                fname = sheet.cell_value(row_idx, fname_col)
                lname = sheet.cell_value(row_idx, lname_col)
                if self.mode == "sif":
                    if fname and lname:
                        fname = field_cleaner(fname)
                        lname = field_cleaner(lname)
                        self.total_checked += 1
                        file_checked += 1
                        # Find in SIF
                        new_id = self.sif_lookup.get((fname, lname))
                        if new_id is not None:
                            ws.write(row_idx, id_col, new_id)
                            self.total_matched += 1
                            file_matched += 1
                        else:
                            date_value = sheet.cell_value(row_idx, date_col) if date_col is not None else None
                            year = None
                            if date_value:
                                if isinstance(date_value, float):
                                    # xlrd returns dates as floats
                                    import xlrd
                                    try:
                                        date_tuple = xlrd.xldate_as_tuple(date_value, rb.datemode)
                                        year = str(date_tuple[0])  # Year is first element
                                    except:
                                        year = None
                                else:
                                    date_str = str(date_value)
                                    if '/' in date_str:
                                        year = date_str.split('/')[-1].split()[0]
                                    elif '-' in date_str:
                                        parts = date_str.split('-')
                                        if len(parts) >= 3:
                                            if len(parts[0]) == 4:  # yyyy-mm-dd format
                                                year = parts[0]
                                            else:  # dd-mm-yyyy format
                                                year = parts[2].split()[0]
                            self.not_found.append({'File': file, 'Row': row_idx + 1, 'Fname': fname, 'Lname': lname, 'Year': year})
                            file_not_found += 1
                            self.logger.debug(f"NOT FOUND in SIF: {fname} {lname}")
                else:  # ssot
                    old_id = sheet.cell_value(row_idx, id_col)
                    if old_id:
                        self.total_checked += 1
                        file_checked += 1
                        cleaned_id = field_cleaner(str(old_id))
                        new_id = self.ssot_lookup.get(cleaned_id)
                        if new_id is not None:
                            ws.write(row_idx, id_col, new_id)
                            self.total_matched += 1
                            file_matched += 1
                        else:
                            self.not_found.append({'File': file, 'Row': row_idx + 1, 'Old ID': str(old_id)})
                            file_not_found += 1
                            self.logger.debug(f"NOT FOUND in SSOT: {old_id}")

            # Save to EOIswapped
            output_path = os.path.join(output_subfolder, os.path.basename(file))
            try:
                wb.save(output_path)
            except Exception as e:
                self.logger.error(f"Error saving {output_path}: {e}")
                input("Please close the file in Excel and press Enter to retry.")
                try:
                    wb.save(output_path)
                except Exception as e2:
                    self.logger.error(f"Failed again: {e2}. Skipping save for {file}.")

        except Exception as e:
            self.logger.error(f"Error processing .xls file {file}: {e}")
            self.files_skipped.append(os.path.basename(file))
            shutil.copy(file, os.path.join(skipped_folder, os.path.basename(file)))

        return file_checked, file_matched, file_not_found

    def _generate_report(self, output_subfolder):
        """Generate Excel report."""
        if not self.not_found and not self.files_checked and not self.files_skipped:
            return

        report_wb = Workbook()
        summary_ws = report_wb.active
        summary_ws.title = 'Summary'
        summary_ws.append(['Metric', 'Value'])
        summary_ws.append(['Total Files Processed', len(self.files_checked)])
        summary_ws.append(['Total Matched', self.total_matched])
        summary_ws.append(['Total NOT Matched', len(self.not_found)])
        summary_ws.append(['Note', 'Numbers will be exaggerated, because students may be checked multiple times if they are in multiple files.'])
        summary_ws.append([])
        summary_ws.append(['Files Checked', 'Files Skipped'])
        for i in range(max(len(self.files_checked), len(self.files_skipped))):
            checked = self.files_checked[i] if i < len(self.files_checked) else ''
            skipped = self.files_skipped[i] if i < len(self.files_skipped) else ''
            summary_ws.append([checked, skipped])

        if self.not_found:
            nf_ws = report_wb.create_sheet('Full List')
            nf_ws.append(list(self.not_found[0].keys()))
            for entry in self.not_found:
                nf_ws.append(list(entry.values()))

        report_path = os.path.join(output_subfolder, "EOI_report.xlsx")
        report_wb.save(report_path)
        report_wb.close()


def main():
    """Main entry point."""
    print(r"""
===================================================================================================
 _____ _____ _____            ___________
|  ___|  _  |_   _|          |_   _|  _  \
| |__ | | | | | |    ______    | | | | | |_____      ____ _ _ __  _ __   ___ _ __
|  __|| | | | | |   |______|   | | | | | / __\ \ /\ / / _` | '_ \| '_ \ / _ \ '__|
| |___\ \_/ /_| |_            _| |_| |/ /\__ \\ V  V / (_| | |_) | |_) |  __/ |
\____/ \___/ \___/            \___/|___/ |___/ \_/\_/ \__,_| .__/| .__/ \___|_|
                                                           | |   | |
                                                           |_|   |_|
===================================================================================================
""")

    # Initialize logger
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, "logging.ini")
    logger = THElogger(script_name="EOI", config_file=config_path)

    try:
        # Create and run swapper
        swapper = EOISwapper(logger)
        swapper.run()
    except Exception as e:
        logger.error(f"Fatal error: {e}")
    finally:
        logger.finalize_report()


if __name__ == "__main__":
    main()
