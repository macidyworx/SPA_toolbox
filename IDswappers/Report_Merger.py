import os
import sys

# Add project root to path so Helpers can be imported
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from openpyxl import load_workbook, Workbook
from Helpers.dog_box import select_work_files
from water_logged.the_logger import THElogger


class ReportMerger:
    def __init__(self):
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logging.ini")
        self.logger = THElogger(script_name="Report_Merger", config_file=config_path)

    def run(self):
        """Execute the report merge process."""
        # Select multiple Excel files
        files = select_work_files([".xlsx"])

        if not files:
            self.logger.error("No files selected. Exiting.")
            return

        # Directory to save the merged report (same as the first file's directory)
        output_dir = os.path.dirname(files[0])
        output_path = os.path.join(output_dir, "FULL_Report.xlsx")

        # Lists to collect data
        summaries = []  # list of list-of-rows
        full_lists = []  # list of list-of-rows (including headers)

        for file in files:
            try:
                wb = load_workbook(file, read_only=True, data_only=True)

                if 'Summary' in wb.sheetnames:
                    ws = wb['Summary']
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        rows.append(list(row))
                    summaries.append(rows)

                if 'Full List' in wb.sheetnames:
                    ws = wb['Full List']
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        rows.append(list(row))
                    full_lists.append(rows)

                wb.close()
            except Exception as e:
                self.logger.error(f"Error reading {file}: {e}")
                continue

        if not summaries:
            self.logger.error("No valid Summary sheets found. Exiting.")
            self.logger.finalize_report()
            return

        out_wb = Workbook()

        # Summary sheet - place side by side
        summary_ws = out_wb.active
        summary_ws.title = 'Summary'
        for summary_rows in summaries:
            # Find the next available column (with a gap)
            start_col = summary_ws.max_column + 1 if summary_ws.max_column > 1 else 1
            if start_col > 1:
                start_col += 1  # gap column
            for row_idx, row_data in enumerate(summary_rows, start=1):
                for col_offset, value in enumerate(row_data):
                    summary_ws.cell(row=row_idx, column=start_col + col_offset, value=value)

        # Full List sheet - concatenate
        if full_lists:
            fl_ws = out_wb.create_sheet('Full List')
            header_written = False
            for fl_rows in full_lists:
                for i, row_data in enumerate(fl_rows):
                    if i == 0 and header_written:
                        continue  # skip duplicate headers
                    fl_ws.append(row_data)
                header_written = True

        out_wb.save(output_path)
        out_wb.close()

        self.logger.info(f"Merged report saved to {output_path}")
        self.logger.finalize_report()


def main():
    merger = ReportMerger()
    merger.run()


if __name__ == "__main__":
    main()
