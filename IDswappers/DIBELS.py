import os
import sys

# Add project root to path so Helpers can be imported
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import shutil
import wx
from openpyxl import load_workbook, Workbook
from xlrd import open_workbook
from xlutils.copy import copy
from Helpers.Clean_fields.clean_field import field_cleaner
from Helpers.dog_box import select_sif, select_work_files, select_output_folder
from water_logged.the_logger import THElogger

# Global constants for headers
FILE_FNAME = "First Name"
FILE_LNAME = "Surname"
FILE_ID_HEADER = "Student ID"
FILE_DATE_HEADER = "Test Date"
SIF_SURNAME = "Surname"
SIF_FIRSTNAME = "Firstname"
SIF_STUDENTID = "StudentID"


class DIBELSSwapper:
    """
    Processes DIBELS 8th Edition Excel files and swaps student IDs based on a SIF.

    Steps:
    1. Prompts for working files (.xls/.xlsx)
    2. Prompts for SIF file
    3. Prompts for output directory
    4. Loads SIF into DataFrame
    5. Processes each file and sheet
    6. Replaces Student IDs with SIF matches
    7. Saves modified files to DIBELSswapped folder
    8. Saves report with details
    """

    def __init__(self, logger):
        self.logger = logger
        self.folder = None
        self.sif_path = None
        self.output_dir = None
        self.dibels_folder = None
        self.skipped_folder = None
        self.sif_lookup = None
        self.not_found = []
        self.total_checked = 0
        self.total_matched = 0
        self.files_checked = []
        self.files_skipped = []

    def run(self):
        """Main execution method."""
        try:
            # Get user inputs using dog_box helpers
            sif_path_result = select_sif()
            if sif_path_result is None:
                self.logger.info("User cancelled SIF selection.")
                self.logger.finalize_report()
                return
            self.sif_path = sif_path_result

            work_files = select_work_files([".xlsx", ".xls"])
            if not work_files:
                self.logger.info("User cancelled file selection or no files selected.")
                self.logger.finalize_report()
                return

            # Filter out temp files
            files = [f for f in work_files if not os.path.basename(f).startswith('~$')]
            if not files:
                self.logger.info("No valid files found after filtering.")
                self.logger.finalize_report()
                return

            output_dir_result = select_output_folder("Select output folder for DIBELS")
            if output_dir_result is None:
                self.logger.info("User cancelled output folder selection.")
                self.logger.finalize_report()
                return
            self.output_dir = output_dir_result

            self.logger.info(f"Using SIF: {self.sif_path}")
            self.logger.info(f"Processing {len(files)} file(s)")
            self.logger.info(f"Output directory: {self.output_dir}")

            # Setup output folder
            self.dibels_folder = os.path.join(self.output_dir, "DIBELSswapped")
            if os.path.exists(self.dibels_folder):
                result = wx.MessageBox(
                    f"{self.dibels_folder} already exists.\nRemove it and continue?",
                    "Output Folder Exists", wx.YES_NO | wx.ICON_WARNING)
                if result == wx.YES:
                    shutil.rmtree(self.dibels_folder)
                else:
                    self.logger.info("User cancelled due to existing output folder.")
                    self.logger.finalize_report()
                    return

            os.mkdir(self.dibels_folder)

            self.skipped_folder = os.path.join(self.dibels_folder, "SKIPPED")
            os.makedirs(self.skipped_folder, exist_ok=True)

            # Load SIF into a lookup dict: (cleaned_firstname, cleaned_surname) -> student_id
            sif_wb = load_workbook(self.sif_path, read_only=True, data_only=True)
            sif_ws = sif_wb.active
            self.sif_lookup = {}
            for row in sif_ws.iter_rows(min_row=3, values_only=True):  # data starts row 3, headers row 2
                if row[3] and row[2] and row[4]:  # Firstname=D(idx3), Surname=C(idx2), StudentID=E(idx4)
                    key = (field_cleaner(str(row[3])), field_cleaner(str(row[2])))
                    self.sif_lookup[key] = row[4]
            sif_wb.close()

            self.logger.info(f"Loaded {len(self.sif_lookup)} students from SIF")

            # Process each file
            file_count = 0
            for file in files:
                file_count += 1
                self.logger.info(f"Processing file {file_count}/{len(files)} > {file}")
                self._process_file(file)

            # Save report
            self._save_report()

            self.logger.info(f"Total Students Checked: {self.total_checked}")
            self.logger.info(f"Total Students Matched: {self.total_matched}")
            self.logger.info(f"Total NOT Found: {len(self.not_found)}")
            self.logger.info(f"Processing complete. Files saved in {self.dibels_folder} folder.")

        except Exception as e:
            self.logger.error(f"Error during processing: {e}")
        finally:
            self.logger.finalize_report()

    def _process_file(self, file):
        """Process a single file (.xlsx or .xls)."""
        file_checked = 0
        file_matched = 0
        file_not_found = 0

        if '.xlsx' in file.lower():
            self._process_xlsx(file)
        elif '.xls' in file.lower():
            self._process_xls(file)
        else:
            self.logger.info(f"Unsupported file format: {file}. Skipping.")
            self.files_skipped.append(os.path.basename(file))
            shutil.copy(file, os.path.join(self.skipped_folder, os.path.basename(file)))

    def _process_xlsx(self, file):
        """Process an .xlsx file."""
        wb = load_workbook(file)

        sheets_processed = 0
        for sheet_name in wb.sheetnames:
            if sheet_name == "Main Menu":
                self.logger.info(f"Skipping sheet: {sheet_name}")
                continue

            ws = wb[sheet_name]

            # Find header row and columns
            header_row = None
            fname_col = None
            lname_col = None
            id_col = None
            date_col = None
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == FILE_FNAME:
                        fname_col = cell.column_letter
                        header_row = cell.row
                    elif cell.value == FILE_LNAME:
                        lname_col = cell.column_letter
                    elif cell.value == FILE_ID_HEADER:
                        id_col = cell.column_letter
                    elif cell.value and str(cell.value).startswith("Test Date"):
                        date_col = cell.column_letter
                if fname_col and lname_col and id_col:
                    break

            if not fname_col or not lname_col or not id_col:
                self.logger.info(f"Sheet '{sheet_name}' in {file} does not have required headers. Skipping.")
                continue

            sheets_processed += 1
            self.logger.info(f"Processing sheet: {sheet_name}")

            file_checked = 0
            file_matched = 0
            file_not_found = 0

            # Process each student row
            for row in range(header_row + 1, ws.max_row + 1):
                fname_cell = ws[f"{fname_col}{row}"]
                lname_cell = ws[f"{lname_col}{row}"]
                id_cell = ws[f"{id_col}{row}"]
                fname = fname_cell.value
                lname = lname_cell.value
                if fname and lname and isinstance(fname, str) and isinstance(lname, str):
                    fname = field_cleaner(fname)
                    lname = field_cleaner(lname)
                    self.total_checked += 1
                    file_checked += 1

                    # Find in SIF
                    new_id = self.sif_lookup.get((fname, lname))
                    if new_id is not None:
                        id_cell.value = new_id
                        self.total_matched += 1
                        file_matched += 1
                    else:
                        date_value = ws[f"{date_col}{row}"].value if date_col else None
                        year = self._extract_year(date_value)
                        self.not_found.append({'File': file, 'Sheet': sheet_name, 'Row': row, 'Fname': fname, 'Lname': lname, 'Year': year})
                        file_not_found += 1
                        self.logger.debug(f"NOT FOUND in SIF: {fname} {lname}")

        if sheets_processed > 0:
            # Save to DIBELSswapped
            output_path = os.path.join(self.dibels_folder, os.path.basename(file))
            try:
                wb.save(output_path)
                self.logger.info(f"Students Checked: {file_checked}")
                self.logger.info(f"Students Matched: {file_matched}")
                self.logger.info(f"Students NOT Found: {file_not_found}")
                self.files_checked.append(os.path.basename(file))
            except Exception as e:
                self.logger.error(f"Error saving {output_path}: {e}")
        else:
            self.logger.info(f"No valid sheets found in {file}. Skipping file.")
            self.files_skipped.append(os.path.basename(file))
            shutil.copy(file, os.path.join(self.skipped_folder, os.path.basename(file)))

    def _process_xls(self, file):
        """Process an .xls file."""
        rb = open_workbook(file, formatting_info=True)
        wb = copy(rb)

        sheets_processed = 0
        for sheet_idx in range(rb.nsheets):
            sheet_name = rb.sheet_names()[sheet_idx]
            if sheet_name == "Main Menu":
                self.logger.info(f"Skipping sheet: {sheet_name}")
                continue

            sheet = rb.sheet_by_index(sheet_idx)
            ws = wb.get_sheet(sheet_idx)

            # Find header row and columns (0-based indices)
            header_row = None
            fname_col = None
            lname_col = None
            id_col = None
            date_col = None
            for row_idx in range(sheet.nrows):
                row = sheet.row(row_idx)
                for col_idx, cell in enumerate(row):
                    if cell.value == FILE_FNAME:
                        fname_col = col_idx
                        header_row = row_idx
                    elif cell.value == FILE_LNAME:
                        lname_col = col_idx
                    elif cell.value == FILE_ID_HEADER:
                        id_col = col_idx
                    elif cell.value and str(cell.value).startswith("Test Date"):
                        date_col = col_idx
                if fname_col is not None and lname_col is not None and id_col is not None:
                    break

            if fname_col is None or lname_col is None or id_col is None:
                self.logger.info(f"Sheet '{sheet_name}' in {file} does not have required headers. Skipping.")
                continue

            sheets_processed += 1
            self.logger.info(f"Processing sheet: {sheet_name}")

            file_checked = 0
            file_matched = 0
            file_not_found = 0

            # Process each student row
            for row_idx in range(header_row + 1, sheet.nrows):
                fname = sheet.cell_value(row_idx, fname_col)
                lname = sheet.cell_value(row_idx, lname_col)
                if fname and lname:
                    fname = field_cleaner(fname)
                    lname = field_cleaner(lname)
                    self.total_checked += 1
                    file_checked += 1

                    # Find in SIF
                    new_id = self.sif_lookup.get((fname, lname))
                    if new_id is not None:
                        ws.write(row_idx, id_col, new_id)
                        self.total_matched += 1
                        file_matched += 1
                    else:
                        date_value = sheet.cell_value(row_idx, date_col) if date_col is not None else None
                        year = self._extract_year(date_value)
                        self.not_found.append({'File': file, 'Sheet': sheet_name, 'Row': row_idx + 1, 'Fname': fname, 'Lname': lname, 'Year': year})
                        file_not_found += 1
                        self.logger.debug(f"NOT FOUND in SIF: {fname} {lname}")

        if sheets_processed > 0:
            # Save to DIBELSswapped
            output_path = os.path.join(self.dibels_folder, os.path.basename(file))
            try:
                wb.save(output_path)
                self.logger.info(f"Students Checked: {file_checked}")
                self.logger.info(f"Students Matched: {file_matched}")
                self.logger.info(f"Students NOT Found: {file_not_found}")
                self.files_checked.append(os.path.basename(file))
            except Exception as e:
                self.logger.error(f"Error saving {output_path}: {e}")
        else:
            self.logger.info(f"No valid sheets found in {file}. Skipping file.")
            self.files_skipped.append(os.path.basename(file))
            shutil.copy(file, os.path.join(self.skipped_folder, os.path.basename(file)))

    def _extract_year(self, date_value):
        """Extract year from date value."""
        year = None
        if date_value:
            date_str = str(date_value)
            if '/' in date_str:
                year = date_str.split('/')[-1].split()[0]
            elif '-' in date_str:
                parts = date_str.split('-')
                if len(parts) >= 3:
                    if len(parts[0]) == 4:  # yyyy-mm-dd format
                        year = parts[0]
                    else:  # dd-mm-yyyy format
                        year = parts[2].split()[0]
        return year

    def _save_report(self):
        """Save summary and not-found reports."""
        if not self.not_found and not self.files_checked and not self.files_skipped:
            return

        report_wb = Workbook()

        # Summary sheet
        summary_ws = report_wb.active
        summary_ws.title = 'Summary'
        summary_ws.append(['Metric', 'Value'])
        summary_ws.append(['Total Files Processed', len(self.files_checked)])
        summary_ws.append(['Total Matched', self.total_matched])
        summary_ws.append(['Total NOT Matched', len(self.not_found)])
        summary_ws.append(['Note', 'Numbers will be exaggerated, because students may be checked multiple times if they are in multiple files.'])
        summary_ws.append([])
        summary_ws.append(['Files Checked', 'Files Skipped'])
        for i in range(max(len(self.files_checked), len(self.files_skipped))):
            checked = self.files_checked[i] if i < len(self.files_checked) else ''
            skipped = self.files_skipped[i] if i < len(self.files_skipped) else ''
            summary_ws.append([checked, skipped])

        # Full List sheet
        if self.not_found:
            nf_ws = report_wb.create_sheet('Full List')
            nf_ws.append(list(self.not_found[0].keys()))
            for entry in self.not_found:
                nf_ws.append(list(entry.values()))

        report_wb.save(os.path.join(self.dibels_folder, "DIBELS_report.xlsx"))
        report_wb.close()


def main():
    """Main entry point."""
    print(r"""
===================================================================================================
 _____ ___________ _____ _      _____
|  _  \_   _| ___ \  ___| |    /  ___|
| | | | | | | |_/ / |__ | |    \ `--.   ______
| | | | | | | ___ \  __|| |     `--. \ |______|
| |/ / _| |_| |_/ / |___| |____/\__/ /
|___/  \___/\____/\____/\_____/\____/


  ___________
 |_   _|  _  \
   | | | | | |_____      ____ _ _ __  _ __   ___ _ __
   | | | | | / __\ \ /\ / / _` | '_ \| '_ \ / _ \ '__|
  _| |_| |/ /\__ \\ V  V / (_| | |_) | |_) |  __/ |
  \___/|___/ |___/ \_/\_/ \__,_| .__/| .__/ \___|_|
                               | |   | |
                               |_|   |_|
===================================================================================================
""")

    # Initialize logger
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logging.ini")
    logger = THElogger(script_name="DIBELS", config_file=config_path)

    # Initialize and run swapper
    swapper = DIBELSSwapper(logger)
    swapper.run()


if __name__ == "__main__":
    main()
