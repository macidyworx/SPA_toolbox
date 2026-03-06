"""
ROL.py - ID Swapper for ROL (Rate of Learning) Excel files.

ROL files have 2 tabs. Only the 'ROL Data' tab contains student IDs to swap.
Column headers: Surname, First Name, Student ID.

Can be run standalone or imported as a module.
"""

import os
import sys

# Add project root to path so Helpers can be imported
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import shutil
import wx
from openpyxl import load_workbook, Workbook
from xlrd import open_workbook
from xlutils.copy import copy
from openpyxl.utils import column_index_from_string
from Helpers.Clean_fields.clean_field import field_cleaner
from Helpers.dog_box import select_single_file, select_work_files, select_output_folder
from water_logged.the_logger import THElogger

# Global constants for headers
FILE_FNAME = "First Name"
FILE_LNAME = "Surname"
FILE_ID_HEADER = "Student ID"
SIF_SURNAME = "Surname"
SIF_FIRSTNAME = "Firstname"
SIF_STUDENTID = "StudentID"

# Tab to process
TARGET_SHEET = "ROL Data"


class ROLSwapper:
    """Processes ROL Excel files and swaps student IDs based on SIF or SSOT."""

    def __init__(self):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        config_path = os.path.join(script_dir, "logging.ini")
        self.logger = THElogger(script_name="ROL", config_file=config_path)

    def _find_headers(self, ws):
        """
        Find header row and column positions in an openpyxl worksheet.

        Returns:
            dict with header_row, fname_col, lname_col, id_col or None if not found.
        """
        fname_col = None
        lname_col = None
        id_col = None
        header_row = None

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_text = field_cleaner(str(cell.value), strip_spaces=True)
                if cell_text == field_cleaner(FILE_FNAME, strip_spaces=True):
                    fname_col = cell.column_letter
                    header_row = cell.row
                elif cell_text == field_cleaner(FILE_LNAME, strip_spaces=True):
                    lname_col = cell.column_letter
                elif cell_text == field_cleaner(FILE_ID_HEADER, strip_spaces=True):
                    id_col = cell.column_letter
            if fname_col and lname_col and id_col:
                break

        if not (fname_col and lname_col and id_col and header_row):
            return None

        return {
            'header_row': header_row,
            'fname_col': fname_col,
            'lname_col': lname_col,
            'id_col': id_col,
        }

    def _find_headers_xlrd(self, sheet):
        """
        Find header row and column positions in an xlrd sheet.

        Returns:
            dict with header_row, fname_col, lname_col, id_col (0-based) or None.
        """
        fname_col = None
        lname_col = None
        id_col = None
        header_row = None

        for row_idx in range(sheet.nrows):
            row = sheet.row(row_idx)
            for col_idx, cell in enumerate(row):
                if not cell.value:
                    continue
                cell_text = field_cleaner(str(cell.value), strip_spaces=True)
                if cell_text == field_cleaner(FILE_FNAME, strip_spaces=True):
                    fname_col = col_idx
                    header_row = row_idx
                elif cell_text == field_cleaner(FILE_LNAME, strip_spaces=True):
                    lname_col = col_idx
                elif cell_text == field_cleaner(FILE_ID_HEADER, strip_spaces=True):
                    id_col = col_idx
            if fname_col is not None and lname_col is not None and id_col is not None:
                break

        if fname_col is None or lname_col is None or id_col is None or header_row is None:
            return None

        return {
            'header_row': header_row,
            'fname_col': fname_col,
            'lname_col': lname_col,
            'id_col': id_col,
        }

    def _process_xlsx_sheet(self, ws, mode, lookup, not_found, file_path):
        """
        Process the ROL Data openpyxl worksheet.

        Returns:
            (checked, matched, not_found_count)
        """
        headers = self._find_headers(ws)
        if headers is None:
            self.logger.warning(f"Required headers not found in '{TARGET_SHEET}' of {os.path.basename(file_path)}")
            return (0, 0, 0)

        header_row = headers['header_row']
        fname_col = headers['fname_col']
        lname_col = headers['lname_col']
        id_col = headers['id_col']

        checked = 0
        matched = 0
        nf_count = 0

        for row in range(header_row + 1, ws.max_row + 1):
            fname_cell = ws[f"{fname_col}{row}"]
            lname_cell = ws[f"{lname_col}{row}"]
            id_cell = ws[f"{id_col}{row}"]
            fname = fname_cell.value
            lname = lname_cell.value

            if mode == "sif":
                if fname and lname and isinstance(fname, str) and isinstance(lname, str):
                    cleaned_fname = field_cleaner(fname)
                    cleaned_lname = field_cleaner(lname)
                    checked += 1
                    new_id = lookup.get((cleaned_fname, cleaned_lname))
                    if new_id is not None:
                        id_cell.value = new_id
                        matched += 1
                    else:
                        not_found.append({
                            'File': os.path.basename(file_path),
                            'Row': row,
                            'Fname': fname,
                            'Lname': lname,
                        })
                        nf_count += 1
                        self.logger.debug(f"NOT FOUND in SIF: {fname} {lname}")
            else:  # ssot mode
                old_id = id_cell.value
                if old_id:
                    checked += 1
                    cleaned_id = field_cleaner(str(old_id))
                    new_id = lookup.get(cleaned_id)
                    if new_id is not None:
                        id_cell.value = new_id
                        matched += 1
                    else:
                        not_found.append({
                            'File': os.path.basename(file_path),
                            'Row': row,
                            'Old ID': str(old_id),
                        })
                        nf_count += 1
                        self.logger.debug(f"NOT FOUND in SSOT: {old_id}")

        return (checked, matched, nf_count)

    def _process_xlrd_sheet(self, rb_sheet, wb_sheet, mode, lookup, not_found, file_path):
        """
        Process the ROL Data xlrd sheet.

        Returns:
            (checked, matched, not_found_count)
        """
        headers = self._find_headers_xlrd(rb_sheet)
        if headers is None:
            self.logger.warning(f"Required headers not found in '{TARGET_SHEET}' of {os.path.basename(file_path)}")
            return (0, 0, 0)

        header_row = headers['header_row']
        fname_col = headers['fname_col']
        lname_col = headers['lname_col']
        id_col = headers['id_col']

        checked = 0
        matched = 0
        nf_count = 0

        for row_idx in range(header_row + 1, rb_sheet.nrows):
            fname = rb_sheet.cell_value(row_idx, fname_col)
            lname = rb_sheet.cell_value(row_idx, lname_col)

            if mode == "sif":
                if fname and lname:
                    cleaned_fname = field_cleaner(str(fname))
                    cleaned_lname = field_cleaner(str(lname))
                    checked += 1
                    new_id = lookup.get((cleaned_fname, cleaned_lname))
                    if new_id is not None:
                        wb_sheet.write(row_idx, id_col, new_id)
                        matched += 1
                    else:
                        not_found.append({
                            'File': os.path.basename(file_path),
                            'Row': row_idx + 1,
                            'Fname': fname,
                            'Lname': lname,
                        })
                        nf_count += 1
                        self.logger.debug(f"NOT FOUND in SIF: {fname} {lname}")
            else:  # ssot mode
                old_id = rb_sheet.cell_value(row_idx, id_col)
                if old_id:
                    checked += 1
                    cleaned_id = field_cleaner(str(old_id))
                    new_id = lookup.get(cleaned_id)
                    if new_id is not None:
                        wb_sheet.write(row_idx, id_col, new_id)
                        matched += 1
                    else:
                        not_found.append({
                            'File': os.path.basename(file_path),
                            'Row': row_idx + 1,
                            'Old ID': str(old_id),
                        })
                        nf_count += 1
                        self.logger.debug(f"NOT FOUND in SSOT: {old_id}")

        return (checked, matched, nf_count)

    def run(self):
        """Main workflow: select lookup, files, output, then process."""
        # Get lookup file (SIF or SSOT)
        lookup_result = select_single_file(mode="choose")
        if not lookup_result:
            self.logger.info("User cancelled file selection.")
            self.logger.finalize_report()
            return

        # Determine mode
        if isinstance(lookup_result, str):
            mode = "sif"
            sif_path = lookup_result
        else:
            mode = "ssot"
            ssot_info = lookup_result

        files = select_work_files([".xlsx", ".xls"])
        if not files:
            self.logger.info("No working files selected.")
            self.logger.finalize_report()
            return

        output_dir = select_output_folder("Select output folder for ROL")
        if not output_dir:
            self.logger.info("User cancelled output folder selection.")
            self.logger.finalize_report()
            return

        rol_folder = os.path.join(output_dir, "ROLswapped")

        # Handle existing output subfolder
        if os.path.exists(rol_folder):
            result = wx.MessageBox(
                f"{rol_folder} already exists.\nRemove it and continue?",
                "Output Folder Exists", wx.YES_NO | wx.ICON_WARNING)
            if result == wx.YES:
                shutil.rmtree(rol_folder)
            else:
                self.logger.info("User cancelled due to existing output folder.")
                self.logger.finalize_report()
                return

        os.makedirs(rol_folder, exist_ok=True)
        skipped_folder = os.path.join(rol_folder, "SKIPPED")
        os.makedirs(skipped_folder, exist_ok=True)

        # Build lookup dictionary
        if mode == "sif":
            self.logger.info(f"Using values: SIF={sif_path}, Output={output_dir}")
            sif_wb = load_workbook(sif_path, read_only=True, data_only=True)
            sif_ws = sif_wb.active
            lookup = {}
            for row in sif_ws.iter_rows(min_row=3, values_only=True):
                if row[3] and row[2] and row[4]:  # Firstname, Surname, StudentID
                    key = (field_cleaner(str(row[3])), field_cleaner(str(row[2])))
                    lookup[key] = row[4]
            sif_wb.close()
        else:
            self.logger.info(f"Using values: SSOT={ssot_info['path']}, Output={output_dir}")
            ssot_wb = load_workbook(ssot_info['path'], read_only=True, data_only=True)
            ssot_ws = ssot_wb.active
            lookup = {}
            header_row = ssot_info['header_row']
            old_col = column_index_from_string(ssot_info['old_id_col'])
            new_col = column_index_from_string(ssot_info['new_id_col'])
            for row in ssot_ws.iter_rows(min_row=header_row + 1):
                old_val = row[old_col - 1].value
                new_val = row[new_col - 1].value
                if old_val and new_val:
                    lookup[field_cleaner(str(old_val))] = new_val
            ssot_wb.close()
            self.logger.info(f"Loaded {len(lookup)} ID mappings from SSOT")

        # Filter files (remove temp files)
        files = [f for f in files if not os.path.basename(f).startswith('~$') and os.path.isfile(f)]

        self.logger.info(f"Total files to process: {len(files)}")

        not_found = []
        total_checked = 0
        total_matched = 0
        files_checked = []
        files_skipped = []

        # Process each file
        file_count = 0
        for file in files:
            file_count += 1
            self.logger.info(f"Processing file {file_count}/{len(files)} > {file}")

            file_checked = 0
            file_matched = 0
            file_not_found = 0

            if file.lower().endswith('.xlsx'):
                wb = load_workbook(file)

                # Find target sheet by name
                if TARGET_SHEET not in wb.sheetnames:
                    self.logger.warning(f"'{TARGET_SHEET}' tab not found in {os.path.basename(file)}. Skipping.")
                    files_skipped.append(os.path.basename(file))
                    shutil.copy(file, os.path.join(skipped_folder, os.path.basename(file)))
                    wb.close()
                    continue

                files_checked.append(os.path.basename(file))

                ws = wb[TARGET_SHEET]
                checked, matched, nf = self._process_xlsx_sheet(
                    ws, mode, lookup, not_found, file)
                file_checked += checked
                file_matched += matched
                file_not_found += nf

                total_checked += file_checked
                total_matched += file_matched

                # Save to ROLswapped
                output_path = os.path.join(rol_folder, os.path.basename(file))
                try:
                    wb.save(output_path)
                except Exception as e:
                    self.logger.error(f"Error saving {output_path}: {e}")
                    input("Please close the file in Excel and press Enter to retry.")
                    try:
                        wb.save(output_path)
                    except Exception as e2:
                        self.logger.error(f"Failed again: {e2}. Skipping save for {file}.")

                self.logger.info(f"Checked={file_checked}, Matched={file_matched}, Not Found={file_not_found}")

            elif file.lower().endswith('.xls'):
                rb = open_workbook(file, formatting_info=True)
                wb = copy(rb)

                # Find target sheet by name
                sheet_names = rb.sheet_names()
                if TARGET_SHEET not in sheet_names:
                    self.logger.warning(f"'{TARGET_SHEET}' tab not found in {os.path.basename(file)}. Skipping.")
                    files_skipped.append(os.path.basename(file))
                    shutil.copy(file, os.path.join(skipped_folder, os.path.basename(file)))
                    continue

                files_checked.append(os.path.basename(file))

                sheet_idx = sheet_names.index(TARGET_SHEET)
                rb_sheet = rb.sheet_by_index(sheet_idx)
                wb_sheet = wb.get_sheet(sheet_idx)
                checked, matched, nf = self._process_xlrd_sheet(
                    rb_sheet, wb_sheet, mode, lookup, not_found, file)
                file_checked += checked
                file_matched += matched
                file_not_found += nf

                total_checked += file_checked
                total_matched += file_matched

                # Save to ROLswapped
                output_path = os.path.join(rol_folder, os.path.basename(file))
                try:
                    wb.save(output_path)
                except Exception as e:
                    self.logger.error(f"Error saving {output_path}: {e}")
                    input("Please close the file in Excel and press Enter to retry.")
                    try:
                        wb.save(output_path)
                    except Exception as e2:
                        self.logger.error(f"Failed again: {e2}. Skipping save for {file}.")

                self.logger.info(f"Checked={file_checked}, Matched={file_matched}, Not Found={file_not_found}")
            else:
                self.logger.error(f"Unsupported file format: {file}. Skipping.")
                files_skipped.append(os.path.basename(file))
                shutil.copy(file, os.path.join(skipped_folder, os.path.basename(file)))

        # Save not found log
        if not_found:
            log_wb = Workbook()
            log_ws = log_wb.active
            log_ws.append(list(not_found[0].keys()))
            for entry in not_found:
                log_ws.append(list(entry.values()))
            log_wb.save(os.path.join(rol_folder, "not_found_log.xlsx"))
            log_wb.close()

        # Save report
        if not_found or files_checked or files_skipped:
            report_wb = Workbook()
            summary_ws = report_wb.active
            summary_ws.title = 'Summary'
            summary_ws.append(['Metric', 'Value'])
            summary_ws.append(['Total Files Processed', len(files_checked)])
            summary_ws.append(['Total Matched', total_matched])
            summary_ws.append(['Total NOT Matched', len(not_found)])
            summary_ws.append(['Note', 'Numbers may be exaggerated, because students may appear in multiple files.'])
            summary_ws.append([])
            summary_ws.append(['Files Checked', 'Files Skipped'])
            for i in range(max(len(files_checked), len(files_skipped))):
                checked = files_checked[i] if i < len(files_checked) else ''
                skipped = files_skipped[i] if i < len(files_skipped) else ''
                summary_ws.append([checked, skipped])

            if not_found:
                nf_ws = report_wb.create_sheet('Full List')
                nf_ws.append(list(not_found[0].keys()))
                for entry in not_found:
                    nf_ws.append(list(entry.values()))

            report_wb.save(os.path.join(rol_folder, "ROL_report.xlsx"))
            report_wb.close()

        self.logger.info(f"Total Students Checked --> {total_checked}")
        self.logger.info(f"Total Students Matched --> {total_matched}")
        self.logger.info(f"Total NOT Found --> {len(not_found)}")
        self.logger.info(f"Processing complete. Files saved in {rol_folder} folder.")
        self.logger.finalize_report()


# === STANDALONE EXECUTION ===
def main():
    """Entry point for standalone execution."""
    print(r"""
===================================================================================================
______  _____ _
| ___ \|  _  | |        ___________
| |_/ /| | | | |        |_   _|  _  \
|    / | | | | |          | | | | | |_____      ____ _ _ __  _ __   ___ _ __
| |\ \ \ \_/ / |____     | | | | | / __\ \ /\ / / _` | '_ \| '_ \ / _ \ '__|
\_| \_| \___/\_____/     _| |_| |/ /\__ \\ V  V / (_| | |_) | |_) |  __/ |
                         \___/|___/ |___/ \_/\_/ \__,_| .__/| .__/ \___|_|
                                                       | |   | |
       Rate of Learning                               |_|   |_|
===================================================================================================
""")

    app = wx.App(False)
    swapper = ROLSwapper()
    swapper.run()
    app.Destroy()


if __name__ == "__main__":
    main()
