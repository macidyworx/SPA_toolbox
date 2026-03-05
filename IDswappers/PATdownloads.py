import os
import pandas as pd
import glob
import shutil
import sys
sys.path.insert(0, '.')
from openpyxl import load_workbook
from xlrd import open_workbook
from xlutils.copy import copy
from tool_box.Helpers.Clean_fields.clean_field import field_cleaner

# Determine repo root for relative paths
script_dir = os.path.dirname(os.path.abspath(__file__))
repo_root = os.path.abspath(os.path.join(script_dir, '..', '..', '..'))

TESTING = True
TESTING_FILES_FOLDER = os.path.join(repo_root, 'ZZSample_Files', 'Test_Files', 'PAT_Online_all')
TESTING_SIF_PATH = os.path.join(repo_root, 'ZZSample_Files', 'SIF10Students.xlsx')
TESTING_OUTPUT_FOLDER = os.path.join(repo_root, 'ZZSample_Files', 'TEST_outputs')

"""
===============================================================================
PAT Downloads Swapper Script
===============================================================================

Description:
This script processes PAT Download Excel files (OARS DOWNLOADS ONLY) (.xls and .xlsx) in a specified folder and its
subfolders, swaps student IDs based on a Student Information File (SIF), and
saves modified copies to an output folder named 'PATswapped'.

Steps performed:
1. Prompts the user for the input folder containing .xls/.xlsx files to process.
2. Prompts for the location of the SIF Excel file (with headers in row 2).
3. Prompts for the output directory where 'PATswapped' will be created.
4. Loads the SIF file into a DataFrame.
5. Recursively finds all .xls/.xlsx files in the input folder.
6. For each file:
   - Finds the 'Given name', 'Family name', and 'Unique ID' columns in the active sheet.
   - Matches students against SIF by Firstname and Surname (case-insensitive).
   - Replaces the Unique ID with the SIF StudentID if found; logs to Excel if not.
   - Prints progress to console.
7. Saves modified files to the PATswapped folder.
8. Saves a log of not-found students as 'not_found_log.xlsx' in PATswapped.

Requirements:
- Python with pandas, openpyxl installed.
- SIF file must have columns: 'Surname', 'Firstname', 'StudentID'.
- Input files must have 'Given name', 'Family name', and 'Unique ID' headers in the first sheet.

Usage:
Run the script: python PAToarsDownloads.py
Follow the prompts to enter paths.

===============================================================================
"""

# Global constants for headers
FILE_FNAME = "Given name"
FILE_LNAME = "Family name"
FILE_ID_HEADER = "Unique ID"
FILE_DATE_HEADER = ["Completed", "Date"]
SIF_SURNAME = "Surname"
SIF_FIRSTNAME = "Firstname"
SIF_STUDENTID = "StudentID"


print(r"""
===================================================================================================
______  ___ _____            ___________                                         
| ___ \/ _ \_   _|          |_   _|  _  \                                        
| |_/ / /_\ \| |    ______    | | | | | |_____      ____ _ _ __  _ __   ___ _ __ 
|  __/|  _  || |   |______|   | | | | | / __\ \ /\ / / _` | '_ \| '_ \ / _ \ '__|
| |   | | | || |             _| |_| |/ /\__ \\ V  V / (_| | |_) | |_) |  __/ |   
\_|   \_| |_/\_/             \___/|___/ |___/ \_/\_/ \__,_| .__/| .__/ \___|_|   
                                                          | |   | |              
          OARS files ONLY                                 |_|   |_|                           
===================================================================================================
""")

def get_user_inputs():
    example_folder = os.path.join(os.path.expanduser('~'), 'Desktop', 'SFDS', 'SORTED', 'PAT')
    folder = input(f"Enter the folder location containing .xls and .xlsx files to swap (e.g., {example_folder}): ").strip('"').strip("'")
    example_sif = os.path.join(os.path.expanduser('~'), 'Desktop', 'ZZZ2025SIF.xlsx')
    sif_path = input(f"Enter the location of the SIF file (e.g., {example_sif}): ").strip('"').strip("'")
    example_output = os.path.join(os.path.expanduser('~'), 'Desktop', 'SwappedSFDS')
    output_dir = input(f"Enter the output directory where PATswapped will be created (e.g., {example_output}): ").strip('"').strip("'")
    return folder, sif_path, output_dir

if TESTING:
    print(f"Using test values: Folder={TESTING_FILES_FOLDER}, SIF={TESTING_SIF_PATH}, Output={TESTING_OUTPUT_FOLDER}")
    use_test = input("Do you want to continue using Test data or Enter actual data locations? Y or N: ").strip().upper()
    if use_test == 'Y':
        folder = TESTING_FILES_FOLDER
        sif_path = TESTING_SIF_PATH
        output_dir = TESTING_OUTPUT_FOLDER
    else:
        folder, sif_path, output_dir = get_user_inputs()
else:
    folder, sif_path, output_dir = get_user_inputs()

print(f"Using values: Folder={folder}, SIF={sif_path}, Output={output_dir}")
pat_folder = os.path.join(output_dir, "PATswapped")
while os.path.exists(pat_folder):
    print("")
    print("===================================================================================================")
    print(f"Error >> {pat_folder} already exists.")
    print("")
    choice = input(f"Do you want to (r)emove it, (m)ove to a new location, or (q)uit? (r/m/q): ").lower().strip()
    if choice == 'r':
        shutil.rmtree(pat_folder)
        print(f"Removed {pat_folder}.")
    elif choice == 'm':
        new_dir = input("Enter new output directory: ")
        pat_folder = os.path.join(new_dir, "PATswapped")
        print(f"Changed to {pat_folder}.")
    elif choice == 'q':
        print("Exiting.")
        exit()
    else:
        print("Invalid choice. Please enter r, m, or q.")

# Check if parent directory exists
parent_dir = os.path.dirname(pat_folder)
if not os.path.exists(parent_dir):
    create = input(f"The directory '{parent_dir}' does not exist. Do you want to create it? (y/n): ").lower().strip()
    if create == 'y':
        os.makedirs(parent_dir, exist_ok=True)
        print(f"Created directory: {parent_dir}")
    else:
        print("Please enter a different output directory.")
        new_dir = input("Enter new output directory: ")
        pat_folder = os.path.join(new_dir, "PATswapped")
        print(f"Changed to {pat_folder}.")
        # Recheck the new location
        while os.path.exists(pat_folder):
            print("")
            print("===================================================================================================")
            print(f"Error >> {pat_folder} already exists.")
            print("")
            choice = input(f"Do you want to (r)emove it, (m)ove to a new location, or (q)uit? (r/m/q): ").lower().strip()
            if choice == 'r':
                shutil.rmtree(pat_folder)
                print(f"Removed {pat_folder}.")
            elif choice == 'm':
                new_dir = input("Enter new output directory: ")
                pat_folder = os.path.join(new_dir, "PATswapped")
                print(f"Changed to {pat_folder}.")
            elif choice == 'q':
                print("Exiting.")
                exit()
            else:
                print("Invalid choice. Please enter r, m, or q.")
        # Now check parent again for the new location
        parent_dir = os.path.dirname(pat_folder)
        if not os.path.exists(parent_dir):
            create = input(f"The directory '{parent_dir}' does not exist. Do you want to create it? (y/n): ").lower().strip()
            if create == 'y':
                os.makedirs(parent_dir, exist_ok=True)
                print(f"Created directory: {parent_dir}")
            else:
                print("Exiting due to invalid directory.")
                exit()

os.mkdir(pat_folder)

skipped_folder = os.path.join(pat_folder, "SKIPPED")
os.makedirs(skipped_folder, exist_ok=True)

# Step 4: Load SIF dataframe
sif_df = pd.read_excel(sif_path, header=1)  # Headers in row 2 (0-indexed as 1)
sif_df[SIF_FIRSTNAME] = sif_df[SIF_FIRSTNAME].apply(field_cleaner)
sif_df[SIF_SURNAME] = sif_df[SIF_SURNAME].apply(field_cleaner)

# Get all files (including in subfolders) or single file
if os.path.isfile(folder):
    files = [folder]
else:
    files = glob.glob(os.path.join(folder, "**", "*"), recursive=True)
    # Filter out temporary Excel files and directories
    files = [f for f in files if not os.path.basename(f).startswith('~$') and os.path.isfile(f)]

print(f"Total files to process: {len(files)}")

# List to log not found students
not_found = []

# Counters
total_checked = 0
total_matched = 0

# Track files
files_checked = []
files_skipped = []

# Step 4-7: Process each file
file_count = 0
for file in files:
    file_count += 1
    print(f"Processing file {file_count}/{len(files)} > {file}")
    
    # Per-file counters
    file_checked = 0
    file_matched = 0
    file_not_found = 0
    
    if file.lower().endswith('.xlsx'):
        wb = load_workbook(file)
        ws = wb.active  # Assume first sheet

        # Find header row and columns
        header_row = None
        fname_col = None
        lname_col = None
        id_col = None
        date_col = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == FILE_FNAME:
                    fname_col = cell.column_letter
                    header_row = cell.row
                elif cell.value == FILE_LNAME:
                    lname_col = cell.column_letter
                elif cell.value == FILE_ID_HEADER:
                    id_col = cell.column_letter
                elif cell.value in FILE_DATE_HEADER:
                    date_col = cell.column_letter
            if fname_col and lname_col and id_col:
                break

        if not fname_col or not lname_col or not id_col:
            print(f"Warning: Required headers not found in {file}. Skipping.")
            files_skipped.append(os.path.basename(file))
            shutil.copy(file, os.path.join(skipped_folder, os.path.basename(file)))
            continue

        files_checked.append(os.path.basename(file))

        # Process each student row
        for row in range(header_row + 1, ws.max_row + 1):
            fname_cell = ws[f"{fname_col}{row}"]
            lname_cell = ws[f"{lname_col}{row}"]
            id_cell = ws[f"{id_col}{row}"]
            fname = fname_cell.value
            lname = lname_cell.value
            if fname and lname and isinstance(fname, str) and isinstance(lname, str):
                fname = field_cleaner(fname)
                lname = field_cleaner(lname)
                #print(f"Checking student: {fname} {lname}")
                total_checked += 1
                file_checked += 1
                # Find in SIF
                match = sif_df[(sif_df[SIF_FIRSTNAME] == fname) & (sif_df[SIF_SURNAME] == lname)]
                if not match.empty:
                    new_id = match[SIF_STUDENTID].iloc[0]
                    id_cell.value = new_id
                    total_matched += 1
                    file_matched += 1
                    #print("Found")
                else:
                    date_value = ws[f"{date_col}{row}"].value if date_col else None
                    year = None
                    if date_value:
                        date_str = str(date_value)
                        if '/' in date_str:
                            year = date_str.split('/')[-1].split()[0]
                        elif '-' in date_str:
                            parts = date_str.split('-')
                            if len(parts) >= 3:
                                if len(parts[0]) == 4:  # yyyy-mm-dd format
                                    year = parts[0]
                                else:  # dd-mm-yyyy format
                                    year = parts[2].split()[0]
                    not_found.append({'File': file, 'Row': row, 'Fname': fname, 'Lname': lname, 'Year': year})
                    file_not_found += 1
                    print(f"NOT FOUND in SIF dataFrame: {fname} {lname}")

        # Save to PATswapped
        output_path = os.path.join(pat_folder, os.path.basename(file))
        try:
            wb.save(output_path)
        except Exception as e:
            print(f"Error saving {output_path}: {e}")
            input("Please close the file in Excel and press Enter to retry.")
            try:
                wb.save(output_path)
            except Exception as e2:
                print(f"Failed again: {e2}. Skipping save for {file}.")
        print(f"Students Checked: {file_checked}")
        print(f"Students Matched: {file_matched}")
        print(f"Students NOT Found: {file_not_found}")
        
    elif file.lower().endswith('.xls'):
        rb = open_workbook(file, formatting_info=True)
        wb = copy(rb)
        ws = wb.get_sheet(0)

        # Find header row and columns (0-based indices)
        header_row = None
        fname_col = None
        lname_col = None
        id_col = None
        date_col = None
        sheet = rb.sheet_by_index(0)
        for row_idx in range(sheet.nrows):
            row = sheet.row(row_idx)
            for col_idx, cell in enumerate(row):
                if cell.value == FILE_FNAME:
                    fname_col = col_idx
                    header_row = row_idx
                elif cell.value == FILE_LNAME:
                    lname_col = col_idx
                elif cell.value == FILE_ID_HEADER:
                    id_col = col_idx
                elif cell.value in FILE_DATE_HEADER:
                    date_col = col_idx
            if fname_col is not None and lname_col is not None and id_col is not None:
                break

        if fname_col is None or lname_col is None or id_col is None:
            print(f"Warning: Required headers not found in {file}. Skipping.")
            files_skipped.append(os.path.basename(file))
            shutil.copy(file, os.path.join(skipped_folder, os.path.basename(file)))
            continue

        files_checked.append(os.path.basename(file))

        # Process each student row
        for row_idx in range(header_row + 1, sheet.nrows):
            fname = sheet.cell_value(row_idx, fname_col)
            lname = sheet.cell_value(row_idx, lname_col)
            if fname and lname:
                fname = field_cleaner(fname)
                lname = field_cleaner(lname)
                #print(f"Checking student: {fname} {lname}")
                total_checked += 1
                file_checked += 1
                # Find in SIF
                match = sif_df[(sif_df[SIF_FIRSTNAME] == fname) & (sif_df[SIF_SURNAME] == lname)]
                if not match.empty:
                    new_id = match[SIF_STUDENTID].iloc[0]
                    ws.write(row_idx, id_col, new_id)
                    total_matched += 1
                    file_matched += 1
                    #print("Found")
                else:
                    date_value = sheet.cell_value(row_idx, date_col) if date_col is not None else None
                    year = None
                    if date_value:
                        date_str = str(date_value)
                        if '/' in date_str:
                            year = date_str.split('/')[-1].split()[0]
                        elif '-' in date_str:
                            parts = date_str.split('-')
                            if len(parts) >= 3:
                                if len(parts[0]) == 4:  # yyyy-mm-dd format
                                    year = parts[0]
                                else:  # dd-mm-yyyy format
                                    year = parts[2].split()[0]
                    not_found.append({'File': file, 'Row': row_idx + 1, 'Fname': fname, 'Lname': lname, 'Year': year})
                    file_not_found += 1
                    print(f"NOT FOUND in SIF dataFrame: {fname} {lname}")

        # Save to PATswapped
        output_path = os.path.join(pat_folder, os.path.basename(file))
        try:
            wb.save(output_path)
        except Exception as e:
            print(f"Error saving {output_path}: {e}")
            input("Please close the file in Excel and press Enter to retry.")
            try:
                wb.save(output_path)
            except Exception as e2:
                print(f"Failed again: {e2}. Skipping save for {file}.")
        print(f"Students Checked: {file_checked}")
        print(f"Students Matched: {file_matched}")
        print(f"Students NOT Found: {file_not_found}")
    else:
        print(f"Unsupported file format: {file}. Skipping.")
        files_skipped.append(os.path.basename(file))
        shutil.copy(file, os.path.join(skipped_folder, os.path.basename(file)))

# Save not found log
if not_found:
    log_df = pd.DataFrame(not_found)
    log_df.to_excel(os.path.join(pat_folder, "not_found_log.xlsx"), index=False)

# Save report
if not_found or files_checked or files_skipped:
    # Create summary data
    summary_data = [
        {'Metric': 'Total Files Processed', 'Value': len(files_checked)},
        {'Metric': 'Total Matched', 'Value': total_matched},
        {'Metric': 'Total NOT Matched', 'Value': len(not_found)},
        {'Metric': 'Note', 'Value': 'Numbers will be exaggerated, because students may be checked multiple times if they are in multiple files.'},
    ]
    summary_df = pd.DataFrame(summary_data)
    
    not_found_df = pd.DataFrame(not_found)
    
    with pd.ExcelWriter(os.path.join(pat_folder, "PAT_report.xlsx")) as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        not_found_df.to_excel(writer, sheet_name='Full List', index=False)
        
        # Add files lists to Summary sheet
        sheet = writer.sheets['Summary']
        sheet.cell(row=7, column=1).value = "Files Checked"
        for i, file in enumerate(files_checked, start=8):
            sheet.cell(row=i, column=1).value = file
        
        sheet.cell(row=7, column=2).value = "Files Skipped"
        for i, file in enumerate(files_skipped, start=8):
            sheet.cell(row=i, column=2).value = file

print(f"Total Students Checked --> {total_checked}")
print(f"Total Students Matched --> {total_matched}")
print(f"Total NOT Found --> {len(not_found)}")

print(f"Processing complete. Files saved in {pat_folder} folder.")